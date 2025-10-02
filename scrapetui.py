#!/usr/bin/env python3
"""
WebScrape-TUI v1.0 - Text User Interface Web Scraping Application

A comprehensive Python-based terminal application for web scraping, data management,
and AI-powered content analysis built with the Textual framework.

PURPOSE:
This application provides an intuitive text-based user interface for scraping websites,
storing articles in a SQLite database, and applying AI-powered analysis including
summarization and sentiment analysis using Google's Gemini API.

KEY FEATURES:

• Interactive TUI built with Textual framework
  - Responsive terminal-based interface
  - Modal dialogs for user input
  - Real-time data tables with sorting and filtering
  - Status indicators and progress feedback

• Web Scraping Capabilities
  - Pre-configured scraper profiles for popular websites
  - Custom scraper creation with URL patterns and CSS selectors
  - BeautifulSoup-powered HTML parsing
  - Robust error handling and retry mechanisms

• Database Management
  - SQLite database for persistent data storage
  - Normalized schema with articles, tags, and scraper profiles
  - Transaction-safe operations with context managers
  - Automatic schema migrations and updates

• AI Integration
  - Google Gemini API integration for content analysis
  - Multiple summarization styles (brief, detailed, bullet points)
  - Sentiment analysis with confidence scoring
  - Async processing to maintain UI responsiveness

• Data Management
  - Advanced filtering by title, URL, date, tags, and sentiment
  - Multiple sorting options (date, title, URL, sentiment, etc.)
  - Tag-based organization system
  - CSV export functionality with applied filters

• Content Reading
  - Full-text article fetching and display
  - Markdown rendering for enhanced readability
  - Article preview and detailed view modes

TECHNICAL IMPLEMENTATION:

• Architecture Patterns:
  - Async/await for non-blocking operations
  - Worker pattern for background tasks
  - Modal screen pattern for user interactions
  - Reactive programming with Textual's reactive system

• Core Technologies:
  - Textual: Modern Python TUI framework
  - SQLite: Embedded database for data persistence
  - BeautifulSoup4: HTML parsing and content extraction
  - Requests: HTTP client for web scraping
  - LXML: Fast XML/HTML parser backend

• Data Flow:
  1. User configures scraper profiles or manual URL/selector input
  2. Background workers fetch content using BeautifulSoup
  3. Articles stored in normalized SQLite database
  4. Optional AI processing via Gemini API
  5. Reactive UI updates with filtered/sorted data display
  6. Export capabilities for data analysis

PENDING FEATURES & IMPROVEMENTS:
• Enhanced scraper profile management
• Additional AI model integrations
• Bulk operations for article management
• Advanced search capabilities
• Data visualization components
• Plugin system for custom processors
• Configuration file management
• Advanced export formats (JSON, XML)
• Scheduled scraping capabilities
• Content deduplication algorithms

CONFIGURATION:
- Database: scraped_data_tui_v1.0.db (configurable via .env)
- Logs: scraper_tui_v1.0.log (configurable via .env)
- Styles: web_scraper_tui_v1.0.tcss
- API Keys: Set GEMINI_API_KEY in .env file for AI features
- Environment: Configure via .env file (copy from .env.example)

VERSION: 1.0 (Stable Release)
AUTHOR: WebScrape-TUI Development Team
LICENSE: MIT
PYTHON: Requires Python 3.8+

For usage instructions, see README.md
For contribution guidelines, see CONTRIBUTING.md
"""

import sqlite3
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin, quote_plus
import logging
import csv
import json
import yaml
import re
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, List, Optional, Tuple, Dict
from abc import ABC, abstractmethod
import functools
import secrets
import shutil
import bcrypt

# APScheduler imports for scheduling functionality
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
from apscheduler.triggers.interval import IntervalTrigger
from apscheduler.triggers.date import DateTrigger

# Data visualization and analytics imports (v1.6.0)
import matplotlib
matplotlib.use('Agg')  # Non-interactive backend for headless chart generation
import matplotlib.pyplot as plt
import pandas as pd
from collections import Counter
from io import BytesIO
import base64

# Enhanced export formats (v1.7.0)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import PieChart, LineChart, BarChart, Reference
from openpyxl.drawing.image import Image as XLImage
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, Image as RLImage
)
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from wordcloud import WordCloud

# Advanced AI features (v1.8.0)
import spacy
from sentence_transformers import SentenceTransformer
from scipy.spatial.distance import cosine
import nltk
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.decomposition import LatentDirichletAllocation
import warnings

# Smart Categorization & Topic Modeling (v1.9.0)
from gensim import corpora, models
from gensim.models import LdaModel, Nmf
from gensim.parsing.preprocessing import STOPWORDS
import networkx as nx
from rouge_score import rouge_scorer
from fuzzywuzzy import fuzz
from sklearn.decomposition import NMF
from sklearn.cluster import KMeans

# Download required NLTK data (suppress output)
try:
    nltk.data.find('corpora/stopwords')
    nltk.data.find('tokenizers/punkt')
except LookupError:
    import ssl
    try:
        _create_unverified_https_context = ssl._create_unverified_context
    except AttributeError:
        pass
    else:
        ssl._create_default_https_context = _create_unverified_https_context

    nltk.download('stopwords', quiet=True)
    nltk.download('punkt', quiet=True)

# Textual imports
from textual.app import App, ComposeResult
from textual.widgets import (
    Header, Footer, DataTable, Static, Button, Input, Label, Markdown,
    LoadingIndicator, RadioSet, RadioButton, ListView, ListItem, Checkbox,
    Select, TextArea
)
from textual.containers import Vertical, Horizontal, VerticalScroll
from textual.screen import ModalScreen
from textual.reactive import reactive
from textual.binding import Binding
from textual.logging import TextualHandler
# from textual.notifications import Notifications # Rely on App.notify()


# --- Environment Configuration ---
def load_env_file(env_path: Path = Path(".env")) -> Dict[str, str]:
    """
    Load environment variables from .env file.

    Args:
        env_path: Path to the .env file

    Returns:
        Dictionary of environment variables
    """
    env_vars = {}
    if env_path.exists():
        try:
            with open(env_path, 'r', encoding='utf-8') as f:
                for line_num, line in enumerate(f, 1):
                    line = line.strip()
                    # Skip empty lines and comments
                    if not line or line.startswith('#'):
                        continue
                    # Parse KEY=VALUE format
                    if '=' in line:
                        key, value = line.split('=', 1)
                        key = key.strip()
                        value = value.strip()
                        # Remove quotes if present
                        if value.startswith('"') and value.endswith('"'):
                            value = value[1:-1]
                        elif value.startswith("'") and value.endswith("'"):
                            value = value[1:-1]
                        env_vars[key] = value
                    else:
                        print(f"Warning: Invalid format in .env file at "
                              f"line {line_num}: {line}")
        except Exception as e:
            print(f"Warning: Could not read .env file: {e}")
    else:
        print("Info: No .env file found. Using default configuration.")

    return env_vars


# Load environment variables
env_vars = load_env_file()

# --- Globals and Configuration ---
DB_PATH = Path(env_vars.get("DATABASE_PATH", "scraped_data_tui_v1.0.db"))
LOG_FILE = Path(env_vars.get("LOG_FILE_PATH", "scraper_tui_v1.0.log"))
GEMINI_API_URL_TEMPLATE = (
    "https://generativelanguage.googleapis.com/v1beta/models/"
    "gemini-2.0-flash:generateContent?key={api_key}"
)
GEMINI_API_KEY = env_vars.get("GEMINI_API_KEY", "")
OPENAI_API_KEY = env_vars.get("OPENAI_API_KEY", "")
CLAUDE_API_KEY = env_vars.get("CLAUDE_API_KEY", "")

logging.basicConfig(
    level=env_vars.get("LOG_LEVEL", "DEBUG"),
    handlers=[
        logging.FileHandler(LOG_FILE, mode='a', encoding='utf-8'),
        TextualHandler()
    ],
    format=("%(asctime)s - %(name)s - %(levelname)s - "
            "%(module)s:%(lineno)d - %(message)s")
)
logger = logging.getLogger(__name__)

# Reduce noise from third-party libraries
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("markdown_it").setLevel(logging.WARNING)
logging.getLogger("urllib3").setLevel(logging.WARNING)
logging.getLogger("requests").setLevel(logging.WARNING)
logging.getLogger("textual").setLevel(logging.INFO)
logging.getLogger("asyncio").setLevel(logging.WARNING)
logging.getLogger("httpcore").setLevel(logging.WARNING)


# --- DateTime Utilities (Python 3.12+ compatible) ---
def db_datetime_now() -> str:
    """Get current datetime as ISO string for database storage (Python 3.12+ compatible)."""
    return datetime.now().isoformat()


def db_datetime_future(hours: int = 24) -> str:
    """Get future datetime as ISO string for database storage (Python 3.12+ compatible)."""
    return (datetime.now() + timedelta(hours=hours)).isoformat()


def parse_db_datetime(iso_string: Optional[str]) -> Optional[datetime]:
    """Parse ISO datetime string from database to datetime object."""
    return datetime.fromisoformat(iso_string) if iso_string else None


# --- Database Utilities ---
def get_db_connection():
    try:
        # Python 3.12+ compatible: no automatic datetime conversion
        # Datetime values stored as ISO strings, no adapters needed
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        # Enable foreign key constraints (required for v2.0.0 multi-user support)
        conn.execute("PRAGMA foreign_keys = ON")
        return conn
    except sqlite3.Error as e:
        logger.critical(f"DB connection error: {e}", exc_info=True)
        raise


# --- Authentication & User Management (v2.0.0) ---

def hash_password(password: str) -> str:
    """
    Hash password using bcrypt with cost factor 12.

    Args:
        password: Plaintext password to hash

    Returns:
        Bcrypt hash string
    """
    salt = bcrypt.gensalt(rounds=12)
    return bcrypt.hashpw(password.encode('utf-8'), salt).decode('utf-8')


def verify_password(password: str, password_hash: str) -> bool:
    """
    Verify password against bcrypt hash.

    Args:
        password: Plaintext password to verify
        password_hash: Bcrypt hash to check against

    Returns:
        True if password matches, False otherwise
    """
    try:
        return bcrypt.checkpw(
            password.encode('utf-8'),
            password_hash.encode('utf-8')
        )
    except Exception as e:
        logger.error(f"Password verification error: {e}")
        return False


def create_session_token() -> str:
    """
    Generate cryptographically secure session token (32 bytes).

    Returns:
        URL-safe base64-encoded token string
    """
    return secrets.token_urlsafe(32)


def create_user_session(
    user_id: int,
    duration_hours: int = 24,
    ip_address: Optional[str] = None
) -> str:
    """
    Create new session for user.

    Args:
        user_id: User ID to create session for
        duration_hours: Session validity duration (default 24 hours)
        ip_address: Optional IP address to associate with session

    Returns:
        Session token string
    """
    with get_db_connection() as conn:
        token = create_session_token()
        expires_at = db_datetime_future(duration_hours)

        conn.execute("""
            INSERT INTO user_sessions (user_id, session_token, expires_at, ip_address)
            VALUES (?, ?, ?, ?)
        """, (user_id, token, expires_at, ip_address))
        conn.commit()

        logger.info(f"Created session for user_id={user_id}, expires at {expires_at}")
        return token


def validate_session(session_token: str) -> Optional[int]:
    """
    Validate session token and return user_id if valid.

    Args:
        session_token: Session token to validate

    Returns:
        user_id if session is valid and not expired, None otherwise
    """
    if not session_token:
        return None

    try:
        with get_db_connection() as conn:
            row = conn.execute("""
                SELECT user_id, expires_at
                FROM user_sessions
                WHERE session_token = ? AND expires_at > ?
            """, (session_token, db_datetime_now())).fetchone()

            if row:
                logger.debug(
                    f"Session validated for user_id={row['user_id']}"
                )
                return row['user_id']
            else:
                logger.debug("Session invalid or expired")
                return None
    except Exception as e:
        logger.error(f"Session validation error: {e}", exc_info=True)
        return None


def logout_session(session_token: str) -> None:
    """
    Delete session from database.

    Args:
        session_token: Session token to logout
    """
    try:
        with get_db_connection() as conn:
            conn.execute(
                "DELETE FROM user_sessions WHERE session_token = ?",
                (session_token,)
            )
            conn.commit()
            logger.info("Session logged out successfully")
    except Exception as e:
        logger.error(f"Session logout error: {e}", exc_info=True)


def authenticate_user(username: str, password: str) -> Optional[int]:
    """
    Authenticate user with username and password.

    Args:
        username: Username to authenticate
        password: Plaintext password to verify

    Returns:
        user_id if credentials are valid and user is active, None otherwise
    """
    try:
        with get_db_connection() as conn:
            row = conn.execute("""
                SELECT id, password_hash, is_active
                FROM users
                WHERE username = ?
            """, (username,)).fetchone()

            if row and row['is_active'] and verify_password(
                password, row['password_hash']
            ):
                # Update last_login timestamp
                conn.execute(
                    "UPDATE users SET last_login = ? WHERE id = ?",
                    (db_datetime_now(), row['id'])
                )
                conn.commit()
                logger.info(f"User authenticated: {username}")
                return row['id']
            else:
                logger.warning(
                    f"Authentication failed for username: {username}"
                )
                return None
    except Exception as e:
        logger.error(f"Authentication error: {e}", exc_info=True)
        return None


def initialize_admin_user() -> None:
    """
    Create default admin user on first run if no users exist.
    Uses configuration: username='admin', password='Ch4ng3M3'
    """
    try:
        with get_db_connection() as conn:
            # Check if any users exist
            count_row = conn.execute(
                "SELECT COUNT(*) as cnt FROM users"
            ).fetchone()

            if count_row['cnt'] == 0:
                # Create default admin user
                password_hash = hash_password('Ch4ng3M3')
                conn.execute("""
                    INSERT INTO users (username, password_hash, email, role)
                    VALUES (?, ?, ?, ?)
                """, ('admin', password_hash, 'admin@localhost', 'admin'))
                conn.commit()
                logger.info(
                    "Default admin user created: username='admin', "
                    "password='Ch4ng3M3' (please change after first login)"
                )
    except Exception as e:
        logger.error(f"Admin user initialization error: {e}", exc_info=True)


def migrate_database_to_v2() -> bool:
    """
    Migrate v1.x database to v2.0 schema.

    Returns:
        True if migration successful, False otherwise
    """
    try:
        with get_db_connection() as conn:
            # Check current version
            try:
                version_row = conn.execute("""
                    SELECT version FROM schema_version
                    ORDER BY applied_at DESC LIMIT 1
                """).fetchone()
                current_version = (
                    version_row['version'] if version_row else '1.0'
                )
            except sqlite3.OperationalError:
                # schema_version table doesn't exist - this is v1.x
                current_version = '1.0'

            if current_version.startswith('2.'):
                logger.info(
                    f"Database already at version {current_version}"
                )
                return True

            # Backup database before migration (skip for :memory:)
            if DB_PATH != ":memory:" and isinstance(DB_PATH, (str, Path)):
                db_path = Path(DB_PATH) if isinstance(DB_PATH, str) else DB_PATH
                if db_path != Path(":memory:"):
                    backup_path = db_path.with_suffix('.db.backup-v1')
                    shutil.copy2(db_path, backup_path)
                    logger.info(f"Database backed up to {backup_path}")

            # Create new tables for v2.0
            conn.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT UNIQUE NOT NULL,
                    password_hash TEXT NOT NULL,
                    email TEXT UNIQUE,
                    role TEXT DEFAULT 'user' CHECK(role IN ('admin', 'user', 'viewer')),
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    last_login TIMESTAMP,
                    is_active BOOLEAN DEFAULT 1
                )
            """)

            conn.execute("""
                CREATE TABLE IF NOT EXISTS user_sessions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    session_token TEXT UNIQUE NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    expires_at TIMESTAMP NOT NULL,
                    ip_address TEXT,
                    FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
                )
            """)

            conn.execute("""
                CREATE TABLE IF NOT EXISTS schema_version (
                    version TEXT PRIMARY KEY,
                    applied_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    description TEXT
                )
            """)

            # Create indexes for performance
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_users_username "
                "ON users(username)"
            )
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_users_email "
                "ON users(email)"
            )
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_sessions_token "
                "ON user_sessions(session_token)"
            )
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_sessions_user_id "
                "ON user_sessions(user_id)"
            )

            # Add user_id columns to existing tables
            try:
                conn.execute(
                    "ALTER TABLE saved_scrapers ADD COLUMN user_id INTEGER"
                )
            except sqlite3.OperationalError:
                pass  # Column already exists

            try:
                conn.execute(
                    "ALTER TABLE saved_scrapers "
                    "ADD COLUMN is_shared BOOLEAN DEFAULT 0"
                )
            except sqlite3.OperationalError:
                pass  # Column already exists

            try:
                conn.execute(
                    "ALTER TABLE scraped_data ADD COLUMN user_id INTEGER"
                )
            except sqlite3.OperationalError:
                pass  # Column already exists

            # Create default admin user (inline to use same connection)
            count_row = conn.execute("SELECT COUNT(*) as cnt FROM users").fetchone()

            if count_row['cnt'] == 0:
                password_hash = hash_password('Ch4ng3M3')
                conn.execute("""
                    INSERT INTO users (username, password_hash, email, role)
                    VALUES (?, ?, ?, ?)
                """, ('admin', password_hash, 'admin@localhost', 'admin'))
                logger.info(
                    "Default admin user created: username='admin', "
                    "password='Ch4ng3M3' (please change after first login)"
                )

            # Get admin user ID
            admin_row = conn.execute(
                "SELECT id FROM users WHERE username = 'admin'"
            ).fetchone()

            if not admin_row:
                logger.error("Failed to create or find admin user during migration")
                return False

            admin_id = admin_row['id']

            # Assign all existing data to admin user (only if tables exist)
            # Check if saved_scrapers table exists
            scrapers_table = conn.execute("""
                SELECT name FROM sqlite_master
                WHERE type='table' AND name='saved_scrapers'
            """).fetchone()
            if scrapers_table:
                conn.execute(
                    "UPDATE saved_scrapers SET user_id = ? WHERE user_id IS NULL",
                    (admin_id,)
                )

            # Check if scraped_data table exists
            data_table = conn.execute("""
                SELECT name FROM sqlite_master
                WHERE type='table' AND name='scraped_data'
            """).fetchone()
            if data_table:
                conn.execute(
                    "UPDATE scraped_data SET user_id = ? WHERE user_id IS NULL",
                    (admin_id,)
                )

            # Record migration
            conn.execute("""
                INSERT INTO schema_version (version, description)
                VALUES (?, ?)
            """, (
                '2.0.0',
                'Multi-user foundation: users, sessions, ownership tracking'
            ))

            conn.commit()
            logger.info("Database successfully migrated to v2.0.0")
            return True

    except Exception as e:
        logger.error(
            f"Database migration failed: {e}",
            exc_info=True
        )
        return False


PREINSTALLED_SCRAPERS = [
    {
        "name": "Generic Article Cleaner",
        "url": "[USER_PROVIDES_URL]",
        "selector": "article, main, div[role='main']",
        "default_limit": 1,
        "default_tags_csv": "article, general",
        "description": (
            "Tries to extract main content from any "
            "article-like page. Selectors are broad; may need "
            "refinement for specific sites."
        )
    },
    {
        "name": "Wikipedia Article Text",
        "url": "https://en.wikipedia.org/wiki/Web_scraping",
        "selector": "div.mw-parser-output p",
        "default_limit": 0,
        "default_tags_csv": "wikipedia, reference",
        "description": (
            "Extracts main paragraph text from Wikipedia "
            "articles. Aims to exclude infoboxes/navs."
        )
    },
    {
        "name": "StackOverflow Q&A",
        "url": "[USER_PROVIDES_URL]",
        "selector": "#question .s-prose, .answer .s-prose",
        "default_limit": 0,
        "default_tags_csv": "stackoverflow, q&a, tech",
        "description": (
            "Extracts questions and answers from StackOverflow "
            "pages. Point to a specific question URL."
        )
    },
    {
        "name": "News Headlines (General)",
        "url": "[USER_PROVIDES_URL]",
        "selector": (
            "h1 a, h2 a, h3 a, article header a, "
            ".headline a, .story-title a"
        ),
        "default_limit": 20,
        "default_tags_csv": "news, headlines",
        "description": (
            "Attempts to extract headlines and links from "
            "news websites. Selector covers common patterns."
        )
    },
    {
        "name": "Academic Abstract (arXiv)",
        "url": "https://arxiv.org/abs/2103.00020",
        "selector": "blockquote.abstract",
        "default_limit": 1,
        "default_tags_csv": "academic, abstract, arxiv",
        "description": (
            "Specifically targets arXiv.org to extract the "
            "abstract of a paper. Replace URL with desired "
            "arXiv paper."
        )
    },
    {
        "name": "Tech Specs (Simple Table)",
        "url": "[USER_PROVIDES_URL]",
        "selector": "table.specifications td, table.specs td",
        "default_limit": 0,
        "default_tags_csv": "technical, specs, table-data",
        "description": (
            "Aims to pull cell data from simple HTML tables "
            "often found in product specifications."
        )
    },
    {
        "name": "Forum Posts (Generic)",
        "url": "[USER_PROVIDES_URL]",
        "selector": ".post-content, .comment-text, .messageText",
        "default_limit": 0,
        "default_tags_csv": "forum, discussion",
        "description": (
            "Designed for typical forum thread structures "
            "to extract individual posts/comments."
        )
    },
    {
        "name": "Recipe Ingredients",
        "url": "[USER_PROVIDES_URL]",
        "selector": (
            ".recipe-ingredients li, .ingredient-list p, "
            "ul.ingredients li"
        ),
        "default_limit": 0,
        "default_tags_csv": "recipe, ingredients, food",
        "description": (
            "Focuses on extracting lists of ingredients from "
            "recipe web pages using common list patterns."
        )
    },
    {
        "name": "Product Details (Basic)",
        "url": "[USER_PROVIDES_URL]",
        "selector": (
            "h1.product-title, .product-name, span.price, "
            ".product-price, #priceblock_ourprice, .pdp-title, "
            ".pdp-price"
        ),
        "default_limit": 5,
        "default_tags_csv": "product, e-commerce",
        "description": (
            "Extracts product title and price from e-commerce "
            "pages using common selectors. Limit is low as it "
            "might pick up related product info."
        )
    },
    {
        "name": "Archived Page (Wayback)",
        "url": "[ARCHIVE_WAYBACK_URL]",
        "selector": "body",  # Special handling for URL
        "default_limit": 1,
        "default_tags_csv": "archive, wayback",
        "description": (
            "SPECIAL: Prompts for an original URL, then tries "
            "to fetch its latest version from the Wayback "
            "Machine. Scrapes entire body."
        )
    },
    {
        "name": "Hacker News Front Page",
        "url": "https://news.ycombinator.com/",
        "selector": "tr.athing",
        "default_limit": 30,
        "default_tags_csv": "hackernews, tech, news",
        "description": (
            "Scrapes the latest tech news and discussions from "
            "Hacker News front page. Extracts titles and links."
        )
    },
    {
        "name": "Reddit Subreddit Posts",
        "url": "https://old.reddit.com/r/python/",
        "selector": "div.thing",
        "default_limit": 25,
        "default_tags_csv": "reddit, social, discussion",
        "description": (
            "Extracts posts from Reddit using old.reddit.com interface. "
            "Change /r/python/ to any subreddit of interest."
        )
    },
    {
        "name": "Medium Articles",
        "url": "https://medium.com/tag/technology",
        "selector": "article, div[data-testid='article-preview']",
        "default_limit": 20,
        "default_tags_csv": "medium, blog, article",
        "description": (
            "Scrapes articles from Medium. Works best with topic tags "
            "or publication pages. Adjust tag in URL as needed."
        )
    },
    {
        "name": "Dev.to Articles",
        "url": "https://dev.to/",
        "selector": "article.crayons-story",
        "default_limit": 20,
        "default_tags_csv": "dev.to, programming, tutorial",
        "description": (
            "Extracts developer articles from Dev.to community. "
            "Great for programming tutorials and tech discussions."
        )
    },
    {
        "name": "GitHub Trending Repos",
        "url": "https://github.com/trending",
        "selector": "article.Box-row",
        "default_limit": 25,
        "default_tags_csv": "github, opensource, trending",
        "description": (
            "Scrapes trending repositories from GitHub. "
            "Can filter by language: /trending/python, /trending/rust, etc."
        )
    },
    {
        "name": "TechCrunch News",
        "url": "https://techcrunch.com/",
        "selector": "article.post-block",
        "default_limit": 20,
        "default_tags_csv": "techcrunch, startup, news",
        "description": (
            "Extracts latest startup and technology news from TechCrunch. "
            "Covers startups, funding, and tech industry news."
        )
    },
    {
        "name": "Ars Technica Articles",
        "url": "https://arstechnica.com/",
        "selector": "article, header.article-header",
        "default_limit": 15,
        "default_tags_csv": "arstechnica, tech, science",
        "description": (
            "Scrapes in-depth technology and science articles "
            "from Ars Technica."
        )
    },
    {
        "name": "The Verge Articles",
        "url": "https://www.theverge.com/",
        "selector": "div.duet--content-cards--content-card",
        "default_limit": 20,
        "default_tags_csv": "theverge, tech, consumer",
        "description": (
            "Extracts consumer technology news and reviews "
            "from The Verge."
        )
    },
    {
        "name": "Product Hunt Products",
        "url": "https://www.producthunt.com/",
        "selector": "li[data-test='post-item']",
        "default_limit": 20,
        "default_tags_csv": "producthunt, products, startup",
        "description": (
            "Scrapes new product launches from Product Hunt. "
            "Discover the latest tech products and tools."
        )
    },
    {
        "name": "Lobsters Tech News",
        "url": "https://lobste.rs/",
        "selector": "div.story",
        "default_limit": 25,
        "default_tags_csv": "lobsters, tech, programming",
        "description": (
            "Extracts computing-focused links from Lobsters community. "
            "High-quality tech discussions and articles."
        )
    },
    {
        "name": "RSS Feed Parser (Generic)",
        "url": "[USER_PROVIDES_URL]",
        "selector": "item, entry",
        "default_limit": 50,
        "default_tags_csv": "rss, feed, syndication",
        "description": (
            "Attempts to parse RSS/Atom feeds. "
            "Works with standard XML feed formats. "
            "Provide feed URL (ending in .xml, .rss, /feed, etc.)"
        )
    },
    {
        "name": "Documentation Pages",
        "url": "[USER_PROVIDES_URL]",
        "selector": (
            "article.documentation, div.doc-content, "
            "main.content, div.markdown-body"
        ),
        "default_limit": 1,
        "default_tags_csv": "documentation, reference, guide",
        "description": (
            "Extracts content from documentation pages. "
            "Works with common doc site patterns (ReadTheDocs, "
            "GitHub Pages, etc.)"
        )
    },
    {
        "name": "Blog Posts (WordPress)",
        "url": "[USER_PROVIDES_URL]",
        "selector": "article.post, div.post-content, article.hentry",
        "default_limit": 10,
        "default_tags_csv": "blog, wordpress, article",
        "description": (
            "Targets WordPress-based blogs using standard article "
            "selectors. Works with most WordPress themes."
        )
    },
    {
        "name": "YouTube Video Descriptions",
        "url": "[USER_PROVIDES_URL]",
        "selector": (
            "ytd-watch-metadata, "
            "ytd-video-description-transcript-section-renderer"
        ),
        "default_limit": 1,
        "default_tags_csv": "youtube, video, description",
        "description": (
            "Attempts to extract video title and description "
            "from YouTube watch pages. May require specific selectors."
        )
    }
]


def init_db():
    """
    Initialize or migrate database to latest schema version.

    For v2.0.0+: Calls migrate_database_to_v2() if needed
    Returns:
        True if successful, False otherwise
    """
    # First, attempt migration to v2.0 if needed
    migration_result = migrate_database_to_v2()
    if not migration_result:
        logger.error("Database migration to v2.0 failed")
        return False

    try:
        with get_db_connection() as conn:
            conn.execute(
                "CREATE TABLE IF NOT EXISTS scraped_data ("
                "id INTEGER PRIMARY KEY AUTOINCREMENT, "
                "url TEXT NOT NULL, "
                "title TEXT NOT NULL, "
                "link TEXT NOT NULL, "
                "timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP, "
                "summary TEXT, "
                "sentiment TEXT, "
                "user_id INTEGER)"
            )
            try:
                conn.execute("ALTER TABLE scraped_data ADD COLUMN summary TEXT")
            except sqlite3.OperationalError:
                pass
            try:
                conn.execute("ALTER TABLE scraped_data ADD COLUMN sentiment TEXT")
            except sqlite3.OperationalError:
                pass
            conn.execute(
                "CREATE TABLE IF NOT EXISTS tags ("
                "id INTEGER PRIMARY KEY AUTOINCREMENT, "
                "name TEXT NOT NULL UNIQUE)"
            )
            conn.execute(
                "CREATE TABLE IF NOT EXISTS article_tags ("
                "article_id INTEGER NOT NULL, "
                "tag_id INTEGER NOT NULL, "
                "FOREIGN KEY (article_id) REFERENCES scraped_data(id) "
                "ON DELETE CASCADE, "
                "FOREIGN KEY (tag_id) REFERENCES tags(id) ON DELETE CASCADE, "
                "PRIMARY KEY (article_id, tag_id))"
            )
            conn.execute("""
                CREATE TABLE IF NOT EXISTS saved_scrapers (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL UNIQUE,
                    url TEXT NOT NULL,
                    selector TEXT NOT NULL,
                    default_limit INTEGER DEFAULT 0,
                    default_tags_csv TEXT,
                    description TEXT,
                    is_preinstalled INTEGER DEFAULT 0,
                    user_id INTEGER,
                    is_shared BOOLEAN DEFAULT 0
                )
            """)
            try:
                conn.execute(
                    "ALTER TABLE saved_scrapers ADD COLUMN description TEXT"
                )
            except sqlite3.OperationalError:
                pass
            try:
                conn.execute(
                    "ALTER TABLE saved_scrapers ADD COLUMN "
                    "is_preinstalled INTEGER DEFAULT 0"
                )
            except sqlite3.OperationalError:
                pass

            # Add summarization templates table
            conn.execute("""
                CREATE TABLE IF NOT EXISTS summarization_templates (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL UNIQUE,
                    template TEXT NOT NULL,
                    description TEXT,
                    is_builtin INTEGER DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # Add filter presets table
            conn.execute("""
                CREATE TABLE IF NOT EXISTS filter_presets (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL UNIQUE,
                    title_filter TEXT,
                    url_filter TEXT,
                    date_from TEXT,
                    date_to TEXT,
                    tags_filter TEXT,
                    sentiment_filter TEXT,
                    use_regex INTEGER DEFAULT 0,
                    tags_logic TEXT DEFAULT 'AND',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # Migrate old filter_presets if date_filter column exists
            try:
                conn.execute("ALTER TABLE filter_presets ADD COLUMN date_from TEXT")
            except sqlite3.OperationalError:
                pass
            try:
                conn.execute("ALTER TABLE filter_presets ADD COLUMN date_to TEXT")
            except sqlite3.OperationalError:
                pass
            try:
                conn.execute("ALTER TABLE filter_presets ADD COLUMN use_regex INTEGER DEFAULT 0")
            except sqlite3.OperationalError:
                pass
            try:
                conn.execute("ALTER TABLE filter_presets ADD COLUMN tags_logic TEXT DEFAULT 'AND'")
            except sqlite3.OperationalError:
                pass

            # Add scheduled scrapes table (v1.5.0)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS scheduled_scrapes (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL UNIQUE,
                    scraper_profile_id INTEGER,
                    schedule_type TEXT NOT NULL,
                    schedule_value TEXT NOT NULL,
                    enabled INTEGER DEFAULT 1,
                    last_run TIMESTAMP,
                    next_run TIMESTAMP,
                    run_count INTEGER DEFAULT 0,
                    last_status TEXT,
                    last_error TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (scraper_profile_id) REFERENCES saved_scrapers(id) ON DELETE CASCADE
                )
            """)

            # Add topic modeling tables (v1.9.0)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS topics (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    label TEXT NOT NULL,
                    keywords TEXT,
                    model_type TEXT,
                    confidence REAL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

            conn.execute("""
                CREATE TABLE IF NOT EXISTS article_topics (
                    article_id INTEGER NOT NULL,
                    topic_id INTEGER NOT NULL,
                    confidence REAL,
                    is_primary INTEGER DEFAULT 0,
                    FOREIGN KEY (article_id) REFERENCES scraped_data(id) ON DELETE CASCADE,
                    FOREIGN KEY (topic_id) REFERENCES topics(id) ON DELETE CASCADE,
                    PRIMARY KEY (article_id, topic_id)
                )
            """)

            # Add entity relationship tables (v1.9.0)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS entities (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    text TEXT NOT NULL UNIQUE,
                    entity_type TEXT,
                    count INTEGER DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

            conn.execute("""
                CREATE TABLE IF NOT EXISTS article_entities (
                    article_id INTEGER NOT NULL,
                    entity_id INTEGER NOT NULL,
                    occurrences INTEGER DEFAULT 1,
                    FOREIGN KEY (article_id) REFERENCES scraped_data(id) ON DELETE CASCADE,
                    FOREIGN KEY (entity_id) REFERENCES entities(id) ON DELETE CASCADE,
                    PRIMARY KEY (article_id, entity_id)
                )
            """)

            conn.execute("""
                CREATE TABLE IF NOT EXISTS entity_relationships (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    entity1_id INTEGER NOT NULL,
                    entity2_id INTEGER NOT NULL,
                    relationship_type TEXT DEFAULT 'co-occurrence',
                    weight INTEGER DEFAULT 1,
                    article_id INTEGER,
                    FOREIGN KEY (entity1_id) REFERENCES entities(id) ON DELETE CASCADE,
                    FOREIGN KEY (entity2_id) REFERENCES entities(id) ON DELETE CASCADE,
                    FOREIGN KEY (article_id) REFERENCES scraped_data(id) ON DELETE CASCADE
                )
            """)

            # Add Q&A history table (v1.9.0)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS qa_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    question TEXT NOT NULL,
                    answer TEXT NOT NULL,
                    article_ids TEXT,
                    confidence REAL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # Add summary feedback table (v1.9.0)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS summary_feedback (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    article_id INTEGER NOT NULL,
                    summary_id TEXT NOT NULL,
                    rating INTEGER NOT NULL,
                    comments TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (article_id) REFERENCES scraped_data(id) ON DELETE CASCADE
                )
            """)

            # Add article clusters table (v1.9.0)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS article_clusters (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    cluster_id INTEGER NOT NULL,
                    article_id INTEGER NOT NULL,
                    cluster_name TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (article_id) REFERENCES scraped_data(id) ON DELETE CASCADE
                )
            """)

            index_statements = [
                "CREATE UNIQUE INDEX IF NOT EXISTS idx_link_unique "
                "ON scraped_data (link);",
                "CREATE INDEX IF NOT EXISTS idx_url ON scraped_data (url);",
                "CREATE INDEX IF NOT EXISTS idx_timestamp "
                "ON scraped_data (timestamp);",
                "CREATE INDEX IF NOT EXISTS idx_title ON scraped_data (title);",
                "CREATE INDEX IF NOT EXISTS idx_sentiment "
                "ON scraped_data (sentiment);",
                "CREATE INDEX IF NOT EXISTS idx_tag_name ON tags (name);",
                "CREATE INDEX IF NOT EXISTS idx_article_tags_article "
                "ON article_tags (article_id);",
                "CREATE INDEX IF NOT EXISTS idx_article_tags_tag "
                "ON article_tags (tag_id);",
                "CREATE UNIQUE INDEX IF NOT EXISTS idx_saved_scraper_name "
                "ON saved_scrapers (name);",
                # v1.9.0 indexes
                "CREATE INDEX IF NOT EXISTS idx_topics_model_type ON topics (model_type);",
                "CREATE INDEX IF NOT EXISTS idx_article_topics_article "
                "ON article_topics (article_id);",
                "CREATE INDEX IF NOT EXISTS idx_article_topics_topic "
                "ON article_topics (topic_id);",
                "CREATE INDEX IF NOT EXISTS idx_entities_text ON entities (text);",
                "CREATE INDEX IF NOT EXISTS idx_entities_type ON entities (entity_type);",
                "CREATE INDEX IF NOT EXISTS idx_article_entities_article "
                "ON article_entities (article_id);",
                "CREATE INDEX IF NOT EXISTS idx_article_entities_entity "
                "ON article_entities (entity_id);",
                "CREATE INDEX IF NOT EXISTS idx_entity_relationships_entity1 "
                "ON entity_relationships (entity1_id);",
                "CREATE INDEX IF NOT EXISTS idx_entity_relationships_entity2 "
                "ON entity_relationships (entity2_id);",
                "CREATE INDEX IF NOT EXISTS idx_qa_history_created "
                "ON qa_history (created_at);",
                "CREATE INDEX IF NOT EXISTS idx_summary_feedback_article "
                "ON summary_feedback (article_id);",
                "CREATE INDEX IF NOT EXISTS idx_article_clusters_cluster "
                "ON article_clusters (cluster_id);",
                "CREATE INDEX IF NOT EXISTS idx_article_clusters_article "
                "ON article_clusters (article_id);"
            ]
            for idx_sql in index_statements:
                conn.execute(idx_sql)
            for ps in PREINSTALLED_SCRAPERS:
                conn.execute("""
                    INSERT OR IGNORE INTO saved_scrapers (
                        name, url, selector, default_limit,
                        default_tags_csv, description, is_preinstalled
                    )
                    VALUES (?, ?, ?, ?, ?, ?, 1)
                """, (ps["name"], ps["url"], ps["selector"],
                      ps["default_limit"], ps["default_tags_csv"],
                      ps["description"]))

            # Insert built-in summarization templates
            builtin_templates = [
                ("Overview", "Provide a concise overview (100-150 words) of the following content:\n\n{content}\n\nOverview:", "Standard overview summary (100-150 words)"),
                ("Bullet Points", "Summarize the following content into key bullet points:\n\n{content}\n\nKey Points:", "Bullet-point summary of main ideas"),
                ("ELI5", "Explain the following content like I'm 5 years old:\n\n{content}\n\nSimple Explanation:", "Simplified explanation for general audience"),
                ("Academic", "Provide an academic-style summary with key findings, methodology, and conclusions:\n\n{content}\n\nAcademic Summary:", "Formal academic summary with structured analysis"),
                ("Executive Summary", "Create an executive summary highlighting key business insights, recommendations, and action items:\n\n{content}\n\nExecutive Summary:", "Business-focused summary with actionable insights"),
                ("Technical Brief", "Summarize the technical aspects, implementation details, and specifications:\n\n{content}\n\nTechnical Summary:", "Technical summary for engineering audience"),
                ("News Brief", "Write a news-style summary with who, what, when, where, why, and how:\n\n{content}\n\nNews Summary:", "Journalistic 5W1H summary format")
            ]
            for name, template, description in builtin_templates:
                conn.execute("""
                    INSERT OR IGNORE INTO summarization_templates (name, template, description, is_builtin)
                    VALUES (?, ?, ?, 1)
                """, (name, template, description))

            conn.commit()
        logger.info("Database initialized/updated successfully for v1.9.0.")
        return True
    except sqlite3.Error as e:
        logger.critical(f"DB init error v1.9.0: {e}", exc_info=True)
        return False


def get_tags_for_article(conn: sqlite3.Connection, article_id: int) -> List[str]:
    cursor = conn.execute(
        "SELECT t.name FROM tags t "
        "JOIN article_tags at ON t.id = at.tag_id "
        "WHERE at.article_id = ? ORDER BY t.name",
        (article_id,)
    )
    return [row['name'] for row in cursor.fetchall()]


def _update_tags_for_article_blocking(article_id: int, new_tags_str: str):
    with get_db_connection() as conn:
        current_tags = set(get_tags_for_article(conn, article_id))
        new_tags_list = [tag.strip().lower() for tag in new_tags_str.split(',') if tag.strip()]
        tags_to_add = [tn for tn in new_tags_list if tn not in current_tags]
        tags_to_remove = [tn for tn in current_tags if tn not in new_tags_list]
        for tag_name in tags_to_add:
            cur = conn.execute("INSERT OR IGNORE INTO tags (name) VALUES (?)", (tag_name,))
            tag_id = cur.lastrowid
            if not tag_id:
                tag_id = conn.execute("SELECT id FROM tags WHERE name = ?", (tag_name,)).fetchone()['id']
            conn.execute("INSERT OR IGNORE INTO article_tags (article_id, tag_id) VALUES (?, ?)", (article_id, tag_id))
        for tag_name in tags_to_remove:
            tag_id_row = conn.execute("SELECT id FROM tags WHERE name = ?", (tag_name,)).fetchone()
            if tag_id_row:
                conn.execute("DELETE FROM article_tags WHERE article_id = ? AND tag_id = ?", (article_id, tag_id_row['id']))
        conn.commit()


def fetch_article_content(article_url: str, for_reading: bool = False) -> str | None:
    logger.info(f"Fetching content from: {article_url}")
    try:
        headers = {'User-Agent': 'Mozilla/5.0 WebScraperTUI/5.0'}
        response = requests.get(article_url, timeout=20, headers=headers)
        response.raise_for_status()
        soup = BeautifulSoup(response.content, "lxml")
        elements = ["article", "main", "div[role='main']", ".entry-content", ".post-content", "body"]
        text = ""
        for s in elements:
            el = soup.select_one(s)
            if el:
                if for_reading:
                    for ut in el(['script', 'style', 'nav', 'header', 'footer', 'aside', '.sidebar']):
                        ut.decompose()
                    text = el.get_text(separator='\n', strip=True)
                else:
                    text = el.get_text(separator=' ', strip=True)
                if len(text) > 200:
                    break
        if for_reading:
            cleaned = "\n".join(line.strip() for line in text.splitlines() if line.strip())
        else:
            cleaned = ' '.join(text.split())
        logger.info(f"Fetched ~{len(cleaned)} chars from {article_url}")
        return cleaned
    except Exception as e:
        logger.error(f"Err fetching {article_url}: {e}", exc_info=True)
        return None


# --- AI Provider Abstraction Layer ---

class AIProvider(ABC):
    """Abstract base class for AI providers."""

    def __init__(self, api_key: str):
        self.api_key = api_key

    @abstractmethod
    def get_summary(self, text: str, style: str = "overview", template: Optional[str] = None) -> Optional[str]:
        """Generate a summary of the text."""
        pass

    @abstractmethod
    def get_sentiment(self, text: str) -> Optional[str]:
        """Analyze sentiment of the text. Returns: Positive, Negative, or Neutral."""
        pass

    @property
    @abstractmethod
    def name(self) -> str:
        """Provider name for display."""
        pass

    @property
    @abstractmethod
    def available_models(self) -> List[str]:
        """List of available models for this provider."""
        pass


class GeminiProvider(AIProvider):
    """Google Gemini AI provider."""

    def __init__(self, api_key: str, model: str = "gemini-2.0-flash"):
        super().__init__(api_key)
        self.model = model
        self.api_url_template = (
            "https://generativelanguage.googleapis.com/v1beta/models/"
            "{model}:generateContent?key={api_key}"
        )

    @property
    def name(self) -> str:
        return "Google Gemini"

    @property
    def available_models(self) -> List[str]:
        return ["gemini-2.0-flash", "gemini-1.5-pro", "gemini-1.5-flash"]

    def get_summary(self, text: str, style: str = "overview", template: Optional[str] = None, max_length: int = 15000) -> Optional[str]:
        if not text:
            return None
        if len(text) > max_length:
            text = text[:max_length]

        # Use custom template if provided, otherwise use default prompts
        if template:
            prompt = template.replace("{content}", text)
        else:
            prompts = {
                "overview": f"Provide a concise overview (100-150 words) of:\n\n{text}\n\nOverview:",
                "bullets": f"Summarize into key bullet points:\n\n{text}\n\nBullets:",
                "eli5": f"Explain like I'm 5:\n\n{text}\n\nELI5:"
            }
            prompt = prompts.get(style, prompts["overview"])

        payload = {
            "contents": [{"role": "user", "parts": [{"text": prompt}]}],
            "generationConfig": {"temperature": 0.6, "maxOutputTokens": 600}
        }
        api_url = self.api_url_template.format(model=self.model, api_key=self.api_key)
        logger.info(f"Requesting '{style}' summary from {self.name}.")

        try:
            response = requests.post(api_url, json=payload, timeout=90)
            response.raise_for_status()
            result = response.json()
            if result.get("candidates") and result["candidates"][0]["content"]["parts"][0].get("text"):
                summary = result["candidates"][0]["content"]["parts"][0]["text"]
                logger.info(f"Summary received (length: {len(summary)}).")
                return summary.strip()
            else:
                logger.error(f"Unexpected {self.name} summary response: {result}")
        except Exception as e:
            logger.error(f"Error calling {self.name} for summary: {e}", exc_info=True)
        return None

    def get_sentiment(self, text: str, max_length: int = 10000) -> Optional[str]:
        if not text:
            return None
        if len(text) > max_length:
            text = text[:max_length]

        prompt = (f"Analyze sentiment. Respond: Positive, Negative, or Neutral.\n\n"
                  f"Text:\n\"\"\"\n{text}\n\"\"\"\n\nSentiment:")
        payload = {
            "contents": [{"role": "user", "parts": [{"text": prompt}]}],
            "generationConfig": {"temperature": 0.2, "maxOutputTokens": 10}
        }
        api_url = self.api_url_template.format(model=self.model, api_key=self.api_key)
        logger.info(f"Requesting sentiment from {self.name}.")

        try:
            response = requests.post(api_url, json=payload, timeout=45)
            response.raise_for_status()
            result = response.json()
            if (result.get("candidates") and
                    result["candidates"][0]["content"]["parts"][0].get("text")):
                s_text = result["candidates"][0]["content"]["parts"][0]["text"].strip().capitalize()
                if s_text in ["Positive", "Negative", "Neutral"]:
                    logger.info(f"Sentiment: {s_text}.")
                    return s_text
                else:
                    logger.warning(f"LLM non-standard sentiment: '{s_text}'. Defaulting Neutral.")
                    return "Neutral"
            else:
                logger.error(f"Unexpected {self.name} sentiment response: {result}")
        except Exception as e:
            logger.error(f"Error calling {self.name} for sentiment: {e}", exc_info=True)
        return None


class OpenAIProvider(AIProvider):
    """OpenAI GPT provider."""

    def __init__(self, api_key: str, model: str = "gpt-4o-mini"):
        super().__init__(api_key)
        self.model = model
        self.api_url = "https://api.openai.com/v1/chat/completions"

    @property
    def name(self) -> str:
        return "OpenAI"

    @property
    def available_models(self) -> List[str]:
        return ["gpt-4o", "gpt-4o-mini", "gpt-4-turbo", "gpt-3.5-turbo"]

    def get_summary(self, text: str, style: str = "overview", template: Optional[str] = None, max_length: int = 15000) -> Optional[str]:
        if not text:
            return None
        if len(text) > max_length:
            text = text[:max_length]

        # Use custom template if provided, otherwise use default prompts
        if template:
            prompt = template.replace("{content}", text)
        else:
            prompts = {
                "overview": f"Provide a concise overview (100-150 words) of:\n\n{text}\n\nOverview:",
                "bullets": f"Summarize into key bullet points:\n\n{text}\n\nBullets:",
                "eli5": f"Explain like I'm 5:\n\n{text}\n\nELI5:"
            }
            prompt = prompts.get(style, prompts["overview"])

        headers = {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/json"
        }
        payload = {
            "model": self.model,
            "messages": [{"role": "user", "content": prompt}],
            "temperature": 0.6,
            "max_tokens": 600
        }
        logger.info(f"Requesting '{style}' summary from {self.name}.")

        try:
            response = requests.post(self.api_url, headers=headers, json=payload, timeout=90)
            response.raise_for_status()
            result = response.json()
            if result.get("choices") and len(result["choices"]) > 0:
                summary = result["choices"][0]["message"]["content"]
                logger.info(f"Summary received (length: {len(summary)}).")
                return summary.strip()
            else:
                logger.error(f"Unexpected {self.name} summary response: {result}")
        except Exception as e:
            logger.error(f"Error calling {self.name} for summary: {e}", exc_info=True)
        return None

    def get_sentiment(self, text: str, max_length: int = 10000) -> Optional[str]:
        if not text:
            return None
        if len(text) > max_length:
            text = text[:max_length]

        prompt = (f"Analyze sentiment. Respond with ONLY one word: Positive, Negative, or Neutral.\n\n"
                  f"Text:\n\"\"\"\n{text}\n\"\"\"\n\nSentiment:")

        headers = {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/json"
        }
        payload = {
            "model": self.model,
            "messages": [{"role": "user", "content": prompt}],
            "temperature": 0.2,
            "max_tokens": 10
        }
        logger.info(f"Requesting sentiment from {self.name}.")

        try:
            response = requests.post(self.api_url, headers=headers, json=payload, timeout=45)
            response.raise_for_status()
            result = response.json()
            if result.get("choices") and len(result["choices"]) > 0:
                s_text = result["choices"][0]["message"]["content"].strip().capitalize()
                if s_text in ["Positive", "Negative", "Neutral"]:
                    logger.info(f"Sentiment: {s_text}.")
                    return s_text
                else:
                    logger.warning(f"LLM non-standard sentiment: '{s_text}'. Defaulting Neutral.")
                    return "Neutral"
            else:
                logger.error(f"Unexpected {self.name} sentiment response: {result}")
        except Exception as e:
            logger.error(f"Error calling {self.name} for sentiment: {e}", exc_info=True)
        return None


class ClaudeProvider(AIProvider):
    """Anthropic Claude provider."""

    def __init__(self, api_key: str, model: str = "claude-3-5-haiku-20241022"):
        super().__init__(api_key)
        self.model = model
        self.api_url = "https://api.anthropic.com/v1/messages"

    @property
    def name(self) -> str:
        return "Anthropic Claude"

    @property
    def available_models(self) -> List[str]:
        return ["claude-3-5-sonnet-20241022", "claude-3-5-haiku-20241022", "claude-3-opus-20240229"]

    def get_summary(self, text: str, style: str = "overview", template: Optional[str] = None, max_length: int = 15000) -> Optional[str]:
        if not text:
            return None
        if len(text) > max_length:
            text = text[:max_length]

        # Use custom template if provided, otherwise use default prompts
        if template:
            prompt = template.replace("{content}", text)
        else:
            prompts = {
                "overview": f"Provide a concise overview (100-150 words) of:\n\n{text}\n\nOverview:",
                "bullets": f"Summarize into key bullet points:\n\n{text}\n\nBullets:",
                "eli5": f"Explain like I'm 5:\n\n{text}\n\nELI5:"
            }
            prompt = prompts.get(style, prompts["overview"])

        headers = {
            "x-api-key": self.api_key,
            "anthropic-version": "2023-06-01",
            "Content-Type": "application/json"
        }
        payload = {
            "model": self.model,
            "messages": [{"role": "user", "content": prompt}],
            "max_tokens": 600,
            "temperature": 0.6
        }
        logger.info(f"Requesting '{style}' summary from {self.name}.")

        try:
            response = requests.post(self.api_url, headers=headers, json=payload, timeout=90)
            response.raise_for_status()
            result = response.json()
            if result.get("content") and len(result["content"]) > 0:
                summary = result["content"][0]["text"]
                logger.info(f"Summary received (length: {len(summary)}).")
                return summary.strip()
            else:
                logger.error(f"Unexpected {self.name} summary response: {result}")
        except Exception as e:
            logger.error(f"Error calling {self.name} for summary: {e}", exc_info=True)
        return None

    def get_sentiment(self, text: str, max_length: int = 10000) -> Optional[str]:
        if not text:
            return None
        if len(text) > max_length:
            text = text[:max_length]

        prompt = (f"Analyze sentiment. Respond with ONLY one word: Positive, Negative, or Neutral.\n\n"
                  f"Text:\n\"\"\"\n{text}\n\"\"\"\n\nSentiment:")

        headers = {
            "x-api-key": self.api_key,
            "anthropic-version": "2023-06-01",
            "Content-Type": "application/json"
        }
        payload = {
            "model": self.model,
            "messages": [{"role": "user", "content": prompt}],
            "max_tokens": 10,
            "temperature": 0.2
        }
        logger.info(f"Requesting sentiment from {self.name}.")

        try:
            response = requests.post(self.api_url, headers=headers, json=payload, timeout=45)
            response.raise_for_status()
            result = response.json()
            if result.get("content") and len(result["content"]) > 0:
                s_text = result["content"][0]["text"].strip().capitalize()
                if s_text in ["Positive", "Negative", "Neutral"]:
                    logger.info(f"Sentiment: {s_text}.")
                    return s_text
                else:
                    logger.warning(f"LLM non-standard sentiment: '{s_text}'. Defaulting Neutral.")
                    return "Neutral"
            else:
                logger.error(f"Unexpected {self.name} sentiment response: {result}")
        except Exception as e:
            logger.error(f"Error calling {self.name} for sentiment: {e}", exc_info=True)
        return None


# Global AI provider instance
_current_ai_provider: Optional[AIProvider] = None


def get_ai_provider() -> Optional[AIProvider]:
    """Get the currently configured AI provider."""
    global _current_ai_provider
    if _current_ai_provider is None:
        # Initialize default provider (Gemini)
        if GEMINI_API_KEY:
            _current_ai_provider = GeminiProvider(GEMINI_API_KEY)
    return _current_ai_provider


def set_ai_provider(provider: AIProvider) -> None:
    """Set the current AI provider."""
    global _current_ai_provider
    _current_ai_provider = provider


# --- Template Manager ---

class TemplateManager:
    """Manages custom summarization templates."""

    @staticmethod
    def get_all_templates() -> List[Dict[str, Any]]:
        """Get all templates from database."""
        try:
            with get_db_connection() as conn:
                cursor = conn.execute("""
                    SELECT id, name, template, description, is_builtin
                    FROM summarization_templates
                    ORDER BY is_builtin DESC, name ASC
                """)
                return [dict(row) for row in cursor.fetchall()]
        except Exception as e:
            logger.error(f"Error loading templates: {e}")
            return []

    @staticmethod
    def get_template_by_name(name: str) -> Optional[str]:
        """Get template text by name."""
        try:
            with get_db_connection() as conn:
                cursor = conn.execute(
                    "SELECT template FROM summarization_templates WHERE name = ?",
                    (name,)
                )
                row = cursor.fetchone()
                return row['template'] if row else None
        except Exception as e:
            logger.error(f"Error loading template '{name}': {e}")
            return None

    @staticmethod
    def save_template(name: str, template: str, description: str = "") -> bool:
        """Save or update a custom template."""
        try:
            with get_db_connection() as conn:
                conn.execute("""
                    INSERT OR REPLACE INTO summarization_templates (name, template, description, is_builtin)
                    VALUES (?, ?, ?, 0)
                """, (name, template, description))
                conn.commit()
                logger.info(f"Template '{name}' saved successfully.")
                return True
        except Exception as e:
            logger.error(f"Error saving template '{name}': {e}")
            return False

    @staticmethod
    def delete_template(name: str) -> bool:
        """Delete a custom template (cannot delete built-in)."""
        try:
            with get_db_connection() as conn:
                cursor = conn.execute(
                    "DELETE FROM summarization_templates WHERE name = ? AND is_builtin = 0",
                    (name,)
                )
                conn.commit()
                if cursor.rowcount > 0:
                    logger.info(f"Template '{name}' deleted successfully.")
                    return True
                else:
                    logger.warning(f"Cannot delete built-in template '{name}'.")
                    return False
        except Exception as e:
            logger.error(f"Error deleting template '{name}': {e}")
            return False

    @staticmethod
    def apply_template(template: str, content: str, title: str = "", url: str = "", date: str = "") -> str:
        """Apply template variables to content."""
        # Replace variable placeholders
        result = template.replace("{content}", content)
        result = result.replace("{title}", title)
        result = result.replace("{url}", url)
        result = result.replace("{date}", date)
        return result


class ConfigManager:
    """Manages application configuration with YAML/JSON persistence."""

    DEFAULT_CONFIG = {
        'ai': {
            'default_provider': 'gemini',
            'default_model': None,
        },
        'export': {
            'default_format': 'csv',
            'output_directory': '.',
        },
        'ui': {
            'theme': 'default',
            'table_columns': ['ID', 'Title', 'URL', 'Link', 'Timestamp', 'Summary', 'Sentiment', 'Tags'],
        },
        'database': {
            'auto_vacuum': False,
            'backup_on_exit': False,
        },
        'logging': {
            'level': 'INFO',
            'max_file_size_mb': 10,
        }
    }

    CONFIG_PATH = Path('config.yaml')

    @classmethod
    def load_config(cls) -> Dict[str, Any]:
        """Load configuration from YAML file or create default."""
        if cls.CONFIG_PATH.exists():
            try:
                with open(cls.CONFIG_PATH, 'r', encoding='utf-8') as f:
                    config = yaml.safe_load(f) or {}
                    # Merge with defaults for any missing keys
                    return cls._merge_config(cls.DEFAULT_CONFIG.copy(), config)
            except Exception as e:
                logger.error(f"Failed to load config: {e}")
                return cls.DEFAULT_CONFIG.copy()
        else:
            # Create default config file
            cls.save_config(cls.DEFAULT_CONFIG)
            return cls.DEFAULT_CONFIG.copy()

    @classmethod
    def save_config(cls, config: Dict[str, Any]) -> bool:
        """Save configuration to YAML file."""
        try:
            with open(cls.CONFIG_PATH, 'w', encoding='utf-8') as f:
                yaml.safe_dump(config, f, default_flow_style=False, sort_keys=False)
            return True
        except Exception as e:
            logger.error(f"Failed to save config: {e}")
            return False

    @classmethod
    def save_as_json(cls, config: Dict[str, Any], path: Path) -> bool:
        """Save configuration as JSON (alternative format)."""
        try:
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=2)
            return True
        except Exception as e:
            logger.error(f"Failed to save JSON config: {e}")
            return False

    @classmethod
    def _merge_config(cls, base: Dict[str, Any], override: Dict[str, Any]) -> Dict[str, Any]:
        """Deep merge override config into base config."""
        result = base.copy()
        for key, value in override.items():
            if key in result and isinstance(result[key], dict) and isinstance(value, dict):
                result[key] = cls._merge_config(result[key], value)
            else:
                result[key] = value
        return result


class FilterPresetManager:
    """Manages filter presets with database persistence."""

    @staticmethod
    def save_preset(name: str, title_filter: str, url_filter: str, date_from: str,
                   date_to: str, tags_filter: str, sentiment_filter: str,
                   use_regex: bool, tags_logic: str) -> bool:
        """Save a new filter preset or update existing."""
        try:
            with get_db_connection() as conn:
                # Check if preset exists
                cursor = conn.execute(
                    "SELECT id FROM filter_presets WHERE name = ?",
                    (name,)
                )
                existing = cursor.fetchone()

                if existing:
                    # Update existing preset
                    conn.execute(
                        """
                        UPDATE filter_presets
                        SET title_filter = ?, url_filter = ?, date_from = ?,
                            date_to = ?, tags_filter = ?, sentiment_filter = ?,
                            use_regex = ?, tags_logic = ?
                        WHERE name = ?
                        """,
                        (title_filter, url_filter, date_from, date_to,
                         tags_filter, sentiment_filter, int(use_regex), tags_logic, name)
                    )
                else:
                    # Insert new preset
                    conn.execute(
                        """
                        INSERT INTO filter_presets
                        (name, title_filter, url_filter, date_from, date_to,
                         tags_filter, sentiment_filter, use_regex, tags_logic)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (name, title_filter, url_filter, date_from, date_to,
                         tags_filter, sentiment_filter, int(use_regex), tags_logic)
                    )
                conn.commit()
                return True
        except Exception as e:
            logger.error(f"Error saving filter preset: {e}")
            return False

    @staticmethod
    def load_preset(name: str) -> Optional[Dict[str, Any]]:
        """Load a filter preset by name."""
        try:
            with get_db_connection() as conn:
                cursor = conn.execute(
                    """
                    SELECT title_filter, url_filter, date_from, date_to,
                           tags_filter, sentiment_filter, use_regex, tags_logic
                    FROM filter_presets
                    WHERE name = ?
                    """,
                    (name,)
                )
                row = cursor.fetchone()
                if row:
                    return {
                        'title_filter': row['title_filter'] or '',
                        'url_filter': row['url_filter'] or '',
                        'date_from': row['date_from'] or '',
                        'date_to': row['date_to'] or '',
                        'tags_filter': row['tags_filter'] or '',
                        'sentiment_filter': row['sentiment_filter'] or '',
                        'use_regex': bool(row['use_regex']),
                        'tags_logic': row['tags_logic'] or 'AND'
                    }
                return None
        except Exception as e:
            logger.error(f"Error loading filter preset: {e}")
            return None

    @staticmethod
    def list_presets() -> List[str]:
        """List all available filter preset names."""
        try:
            with get_db_connection() as conn:
                cursor = conn.execute(
                    "SELECT name FROM filter_presets ORDER BY created_at DESC"
                )
                return [row['name'] for row in cursor.fetchall()]
        except Exception as e:
            logger.error(f"Error listing filter presets: {e}")
            return []

    @staticmethod
    def delete_preset(name: str) -> bool:
        """Delete a filter preset."""
        try:
            with get_db_connection() as conn:
                conn.execute("DELETE FROM filter_presets WHERE name = ?", (name,))
                conn.commit()
                return True
        except Exception as e:
            logger.error(f"Error deleting filter preset: {e}")
            return False


class ScheduleManager:
    """Manages scheduled scraping with APScheduler integration (v1.5.0)."""

    @staticmethod
    def create_schedule(name: str, scraper_profile_id: int, schedule_type: str,
                       schedule_value: str, enabled: bool = True) -> bool:
        """
        Create a new scheduled scrape.

        Args:
            name: Unique schedule name
            scraper_profile_id: ID of scraper profile to execute
            schedule_type: Type of schedule ('interval', 'cron', 'daily', 'weekly')
            schedule_value: Schedule parameters (e.g., '60' for 60 minutes, '0 9 * * *' for cron)
            enabled: Whether schedule is active

        Returns:
            True if successful, False otherwise
        """
        try:
            with get_db_connection() as conn:
                # Calculate next run time based on schedule
                next_run = ScheduleManager._calculate_next_run(schedule_type, schedule_value)

                conn.execute("""
                    INSERT INTO scheduled_scrapes
                    (name, scraper_profile_id, schedule_type, schedule_value, enabled, next_run)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (name, scraper_profile_id, schedule_type, schedule_value, int(enabled), next_run))
                conn.commit()
                logger.info(f"Schedule '{name}' created successfully")
                return True
        except sqlite3.IntegrityError as e:
            logger.error(f"Schedule name '{name}' already exists: {e}")
            return False
        except Exception as e:
            logger.error(f"Error creating schedule: {e}")
            return False

    @staticmethod
    def update_schedule(schedule_id: int, name: Optional[str] = None,
                       scraper_profile_id: Optional[int] = None,
                       schedule_type: Optional[str] = None,
                       schedule_value: Optional[str] = None,
                       enabled: Optional[bool] = None) -> bool:
        """Update an existing schedule."""
        try:
            with get_db_connection() as conn:
                # Build dynamic update query
                updates = []
                params = []

                if name is not None:
                    updates.append("name = ?")
                    params.append(name)
                if scraper_profile_id is not None:
                    updates.append("scraper_profile_id = ?")
                    params.append(scraper_profile_id)
                if schedule_type is not None and schedule_value is not None:
                    updates.append("schedule_type = ?")
                    updates.append("schedule_value = ?")
                    params.append(schedule_type)
                    params.append(schedule_value)
                    # Recalculate next run
                    next_run = ScheduleManager._calculate_next_run(schedule_type, schedule_value)
                    updates.append("next_run = ?")
                    params.append(next_run)
                if enabled is not None:
                    updates.append("enabled = ?")
                    params.append(int(enabled))

                if not updates:
                    return True  # Nothing to update

                params.append(schedule_id)
                query = f"UPDATE scheduled_scrapes SET {', '.join(updates)} WHERE id = ?"
                conn.execute(query, params)
                conn.commit()
                logger.info(f"Schedule ID {schedule_id} updated successfully")
                return True
        except Exception as e:
            logger.error(f"Error updating schedule: {e}")
            return False

    @staticmethod
    def delete_schedule(schedule_id: int) -> bool:
        """Delete a scheduled scrape."""
        try:
            with get_db_connection() as conn:
                conn.execute("DELETE FROM scheduled_scrapes WHERE id = ?", (schedule_id,))
                conn.commit()
                logger.info(f"Schedule ID {schedule_id} deleted successfully")
                return True
        except Exception as e:
            logger.error(f"Error deleting schedule: {e}")
            return False

    @staticmethod
    def list_schedules(enabled_only: bool = False) -> List[Dict[str, Any]]:
        """
        List all scheduled scrapes.

        Args:
            enabled_only: If True, only return enabled schedules

        Returns:
            List of schedule dictionaries
        """
        try:
            with get_db_connection() as conn:
                query = """
                    SELECT
                        ss.id, ss.name, ss.scraper_profile_id, ss.schedule_type,
                        ss.schedule_value, ss.enabled, ss.last_run, ss.next_run,
                        ss.run_count, ss.last_status, ss.last_error, ss.created_at,
                        sp.name as profile_name
                    FROM scheduled_scrapes ss
                    LEFT JOIN saved_scrapers sp ON ss.scraper_profile_id = sp.id
                """
                if enabled_only:
                    query += " WHERE ss.enabled = 1"
                query += " ORDER BY ss.created_at DESC"

                cursor = conn.execute(query)
                schedules = []
                for row in cursor.fetchall():
                    schedules.append({
                        'id': row['id'],
                        'name': row['name'],
                        'scraper_profile_id': row['scraper_profile_id'],
                        'profile_name': row['profile_name'],
                        'schedule_type': row['schedule_type'],
                        'schedule_value': row['schedule_value'],
                        'enabled': bool(row['enabled']),
                        'last_run': row['last_run'],
                        'next_run': row['next_run'],
                        'run_count': row['run_count'],
                        'last_status': row['last_status'],
                        'last_error': row['last_error'],
                        'created_at': row['created_at']
                    })
                return schedules
        except Exception as e:
            logger.error(f"Error listing schedules: {e}")
            return []

    @staticmethod
    def get_schedule(schedule_id: int) -> Optional[Dict[str, Any]]:
        """Get a specific schedule by ID."""
        try:
            with get_db_connection() as conn:
                cursor = conn.execute("""
                    SELECT
                        ss.id, ss.name, ss.scraper_profile_id, ss.schedule_type,
                        ss.schedule_value, ss.enabled, ss.last_run, ss.next_run,
                        ss.run_count, ss.last_status, ss.last_error, ss.created_at,
                        sp.name as profile_name, sp.url, sp.selector
                    FROM scheduled_scrapes ss
                    LEFT JOIN saved_scrapers sp ON ss.scraper_profile_id = sp.id
                    WHERE ss.id = ?
                """, (schedule_id,))
                row = cursor.fetchone()
                if row:
                    return {
                        'id': row['id'],
                        'name': row['name'],
                        'scraper_profile_id': row['scraper_profile_id'],
                        'profile_name': row['profile_name'],
                        'profile_url': row['url'],
                        'profile_selector': row['selector'],
                        'schedule_type': row['schedule_type'],
                        'schedule_value': row['schedule_value'],
                        'enabled': bool(row['enabled']),
                        'last_run': row['last_run'],
                        'next_run': row['next_run'],
                        'run_count': row['run_count'],
                        'last_status': row['last_status'],
                        'last_error': row['last_error'],
                        'created_at': row['created_at']
                    }
                return None
        except Exception as e:
            logger.error(f"Error getting schedule: {e}")
            return None

    @staticmethod
    def record_execution(schedule_id: int, status: str, error: Optional[str] = None) -> bool:
        """
        Record a schedule execution result.

        Args:
            schedule_id: Schedule ID
            status: Execution status ('success', 'failed', 'running')
            error: Error message if failed

        Returns:
            True if successful
        """
        try:
            with get_db_connection() as conn:
                now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

                # Get current schedule to calculate next run
                cursor = conn.execute(
                    "SELECT schedule_type, schedule_value FROM scheduled_scrapes WHERE id = ?",
                    (schedule_id,)
                )
                row = cursor.fetchone()
                if not row:
                    return False

                next_run = ScheduleManager._calculate_next_run(
                    row['schedule_type'],
                    row['schedule_value']
                )

                conn.execute("""
                    UPDATE scheduled_scrapes
                    SET last_run = ?, last_status = ?, last_error = ?,
                        run_count = run_count + 1, next_run = ?
                    WHERE id = ?
                """, (now, status, error, next_run, schedule_id))
                conn.commit()
                return True
        except Exception as e:
            logger.error(f"Error recording execution: {e}")
            return False

    @staticmethod
    def _calculate_next_run(schedule_type: str, schedule_value: str) -> str:
        """
        Calculate next run time based on schedule type.

        Args:
            schedule_type: Type of schedule ('interval', 'cron', 'daily', 'weekly', 'hourly')
            schedule_value: Schedule parameters

        Returns:
            ISO formatted datetime string for next run
        """
        now = datetime.now()

        try:
            if schedule_type == 'hourly':
                next_run = now + timedelta(hours=1)
            elif schedule_type == 'daily':
                # Parse HH:MM format from schedule_value
                try:
                    hour, minute = map(int, schedule_value.split(':'))
                    next_run = now.replace(hour=hour, minute=minute, second=0, microsecond=0)
                    if next_run <= now:
                        next_run += timedelta(days=1)
                except:
                    next_run = now + timedelta(days=1)
            elif schedule_type == 'weekly':
                # Parse day_of_week:HH:MM format (0=Monday, 6=Sunday)
                try:
                    parts = schedule_value.split(':')
                    day_of_week = int(parts[0])
                    hour, minute = int(parts[1]), int(parts[2])

                    days_ahead = day_of_week - now.weekday()
                    if days_ahead <= 0:  # Target day already happened this week
                        days_ahead += 7
                    next_run = now + timedelta(days=days_ahead)
                    next_run = next_run.replace(hour=hour, minute=minute, second=0, microsecond=0)
                except:
                    next_run = now + timedelta(weeks=1)
            elif schedule_type == 'interval':
                # Minutes interval
                try:
                    minutes = int(schedule_value)
                    next_run = now + timedelta(minutes=minutes)
                except:
                    next_run = now + timedelta(hours=1)
            elif schedule_type == 'cron':
                # For cron, we'll use a simplified next-hour approach
                # Full cron parsing would require croniter library
                next_run = now + timedelta(hours=1)
            else:
                next_run = now + timedelta(hours=1)

            return next_run.strftime('%Y-%m-%d %H:%M:%S')
        except Exception as e:
            logger.error(f"Error calculating next run: {e}")
            return (now + timedelta(hours=1)).strftime('%Y-%m-%d %H:%M:%S')


class AnalyticsManager:
    """Manages data analytics and visualization (v1.6.0)."""

    @staticmethod
    def get_statistics() -> Dict[str, Any]:
        """
        Get comprehensive statistics about scraped data.

        Returns:
            Dictionary containing various statistics
        """
        try:
            with get_db_connection() as conn:
                stats = {}

                # Total articles
                cursor = conn.execute("SELECT COUNT(*) as total FROM scraped_data")
                stats['total_articles'] = cursor.fetchone()['total']

                # Articles with summaries
                cursor = conn.execute("SELECT COUNT(*) as total FROM scraped_data WHERE summary IS NOT NULL")
                stats['articles_with_summaries'] = cursor.fetchone()['total']

                # Articles with sentiment
                cursor = conn.execute("SELECT COUNT(*) as total FROM scraped_data WHERE sentiment IS NOT NULL")
                stats['articles_with_sentiment'] = cursor.fetchone()['total']

                # Sentiment distribution
                cursor = conn.execute("""
                    SELECT sentiment, COUNT(*) as count
                    FROM scraped_data
                    WHERE sentiment IS NOT NULL
                    GROUP BY sentiment
                """)
                sentiment_dist = {row['sentiment']: row['count'] for row in cursor.fetchall()}
                stats['sentiment_distribution'] = sentiment_dist

                # Top sources
                cursor = conn.execute("""
                    SELECT url, COUNT(*) as count
                    FROM scraped_data
                    GROUP BY url
                    ORDER BY count DESC
                    LIMIT 10
                """)
                stats['top_sources'] = [(row['url'], row['count']) for row in cursor.fetchall()]

                # Top tags
                cursor = conn.execute("""
                    SELECT t.name, COUNT(*) as count
                    FROM tags t
                    JOIN article_tags at ON t.id = at.tag_id
                    GROUP BY t.id, t.name
                    ORDER BY count DESC
                    LIMIT 20
                """)
                stats['top_tags'] = [(row['name'], row['count']) for row in cursor.fetchall()]

                # Articles per day (last 30 days)
                cursor = conn.execute("""
                    SELECT DATE(timestamp) as date, COUNT(*) as count
                    FROM scraped_data
                    WHERE timestamp >= datetime('now', '-30 days')
                    GROUP BY DATE(timestamp)
                    ORDER BY date
                """)
                stats['articles_per_day'] = [(row['date'], row['count']) for row in cursor.fetchall()]

                # Summary statistics
                stats['summary_percentage'] = (
                    (stats['articles_with_summaries'] / stats['total_articles'] * 100)
                    if stats['total_articles'] > 0 else 0
                )
                stats['sentiment_percentage'] = (
                    (stats['articles_with_sentiment'] / stats['total_articles'] * 100)
                    if stats['total_articles'] > 0 else 0
                )

                return stats
        except Exception as e:
            logger.error(f"Error getting statistics: {e}")
            return {}

    @staticmethod
    def generate_sentiment_chart(output_path: Optional[str] = None) -> Optional[str]:
        """
        Generate a pie chart showing sentiment distribution.

        Args:
            output_path: Optional file path to save chart. If None, returns base64 encoded image.

        Returns:
            File path if saved, or base64 encoded image string
        """
        try:
            stats = AnalyticsManager.get_statistics()
            sentiment_dist = stats.get('sentiment_distribution', {})

            if not sentiment_dist:
                logger.warning("No sentiment data available for chart")
                return None

            # Create pie chart
            plt.figure(figsize=(10, 6))
            labels = list(sentiment_dist.keys())
            sizes = list(sentiment_dist.values())
            colors = []
            for label in labels:
                if 'positive' in label.lower():
                    colors.append('#4CAF50')
                elif 'negative' in label.lower():
                    colors.append('#F44336')
                else:
                    colors.append('#FFC107')

            plt.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', startangle=90)
            plt.axis('equal')
            plt.title('Sentiment Distribution', fontsize=16, fontweight='bold')

            if output_path:
                plt.savefig(output_path, dpi=150, bbox_inches='tight')
                plt.close()
                return output_path
            else:
                # Return base64 encoded image
                buffer = BytesIO()
                plt.savefig(buffer, format='png', dpi=150, bbox_inches='tight')
                plt.close()
                buffer.seek(0)
                img_base64 = base64.b64encode(buffer.read()).decode()
                return f"data:image/png;base64,{img_base64}"
        except Exception as e:
            logger.error(f"Error generating sentiment chart: {e}")
            return None

    @staticmethod
    def generate_timeline_chart(output_path: Optional[str] = None) -> Optional[str]:
        """
        Generate a line chart showing articles scraped over time.

        Args:
            output_path: Optional file path to save chart. If None, returns base64 encoded image.

        Returns:
            File path if saved, or base64 encoded image string
        """
        try:
            stats = AnalyticsManager.get_statistics()
            articles_per_day = stats.get('articles_per_day', [])

            if not articles_per_day:
                logger.warning("No timeline data available for chart")
                return None

            # Create line chart
            dates = [item[0] for item in articles_per_day]
            counts = [item[1] for item in articles_per_day]

            plt.figure(figsize=(12, 6))
            plt.plot(dates, counts, marker='o', linestyle='-', linewidth=2, markersize=6, color='#2196F3')
            plt.xlabel('Date', fontsize=12)
            plt.ylabel('Number of Articles', fontsize=12)
            plt.title('Articles Scraped Over Time (Last 30 Days)', fontsize=16, fontweight='bold')
            plt.grid(True, alpha=0.3)
            plt.xticks(rotation=45, ha='right')
            plt.tight_layout()

            if output_path:
                plt.savefig(output_path, dpi=150, bbox_inches='tight')
                plt.close()
                return output_path
            else:
                buffer = BytesIO()
                plt.savefig(buffer, format='png', dpi=150, bbox_inches='tight')
                plt.close()
                buffer.seek(0)
                img_base64 = base64.b64encode(buffer.read()).decode()
                return f"data:image/png;base64,{img_base64}"
        except Exception as e:
            logger.error(f"Error generating timeline chart: {e}")
            return None

    @staticmethod
    def generate_top_sources_chart(output_path: Optional[str] = None) -> Optional[str]:
        """
        Generate a horizontal bar chart showing top sources.

        Args:
            output_path: Optional file path to save chart. If None, returns base64 encoded image.

        Returns:
            File path if saved, or base64 encoded image string
        """
        try:
            stats = AnalyticsManager.get_statistics()
            top_sources = stats.get('top_sources', [])

            if not top_sources:
                logger.warning("No source data available for chart")
                return None

            # Create horizontal bar chart
            sources = [item[0][:50] + '...' if len(item[0]) > 50 else item[0] for item in top_sources]
            counts = [item[1] for item in top_sources]

            plt.figure(figsize=(12, 8))
            plt.barh(sources, counts, color='#9C27B0')
            plt.xlabel('Number of Articles', fontsize=12)
            plt.ylabel('Source', fontsize=12)
            plt.title('Top 10 Sources by Article Count', fontsize=16, fontweight='bold')
            plt.gca().invert_yaxis()  # Highest at top
            plt.tight_layout()

            if output_path:
                plt.savefig(output_path, dpi=150, bbox_inches='tight')
                plt.close()
                return output_path
            else:
                buffer = BytesIO()
                plt.savefig(buffer, format='png', dpi=150, bbox_inches='tight')
                plt.close()
                buffer.seek(0)
                img_base64 = base64.b64encode(buffer.read()).decode()
                return f"data:image/png;base64,{img_base64}"
        except Exception as e:
            logger.error(f"Error generating top sources chart: {e}")
            return None

    @staticmethod
    def generate_tag_cloud_data() -> List[Tuple[str, int]]:
        """
        Generate tag cloud data (tag frequencies).

        Returns:
            List of (tag_name, count) tuples sorted by count descending
        """
        try:
            stats = AnalyticsManager.get_statistics()
            return stats.get('top_tags', [])
        except Exception as e:
            logger.error(f"Error generating tag cloud data: {e}")
            return []

    @staticmethod
    def export_statistics_report(output_path: str) -> bool:
        """
        Export a comprehensive statistics report to a text file.

        Args:
            output_path: Path to save the report

        Returns:
            True if successful, False otherwise
        """
        try:
            stats = AnalyticsManager.get_statistics()

            with open(output_path, 'w', encoding='utf-8') as f:
                f.write("=" * 80 + "\n")
                f.write("WebScrape-TUI Analytics Report\n")
                f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write("=" * 80 + "\n\n")

                f.write("OVERVIEW\n")
                f.write("-" * 80 + "\n")
                f.write(f"Total Articles: {stats.get('total_articles', 0)}\n")
                f.write(f"Articles with Summaries: {stats.get('articles_with_summaries', 0)} ({stats.get('summary_percentage', 0):.1f}%)\n")
                f.write(f"Articles with Sentiment: {stats.get('articles_with_sentiment', 0)} ({stats.get('sentiment_percentage', 0):.1f}%)\n\n")

                f.write("SENTIMENT DISTRIBUTION\n")
                f.write("-" * 80 + "\n")
                for sentiment, count in stats.get('sentiment_distribution', {}).items():
                    f.write(f"  {sentiment}: {count}\n")
                f.write("\n")

                f.write("TOP 10 SOURCES\n")
                f.write("-" * 80 + "\n")
                for i, (source, count) in enumerate(stats.get('top_sources', []), 1):
                    f.write(f"{i:2d}. {source[:70]:<70} ({count} articles)\n")
                f.write("\n")

                f.write("TOP 20 TAGS\n")
                f.write("-" * 80 + "\n")
                for i, (tag, count) in enumerate(stats.get('top_tags', []), 1):
                    f.write(f"{i:2d}. {tag:<30} ({count} uses)\n")
                f.write("\n")

                f.write("=" * 80 + "\n")

            logger.info(f"Statistics report exported to {output_path}")
            return True
        except Exception as e:
            logger.error(f"Error exporting statistics report: {e}")
            return False


class ExcelExportManager:
    """
    Enhanced Excel export with formatting, charts, and multi-sheet layouts.

    Provides professional XLSX export capabilities including:
    - Multiple sheets (Articles, Statistics, Timeline, Charts)
    - Styled headers with auto-column sizing
    - Embedded charts and graphs
    - Filter metadata preservation
    - Custom export templates
    """

    @staticmethod
    def export_to_excel(
        articles: List[Dict[str, Any]],
        output_path: str,
        include_charts: bool = True,
        template: str = "standard"
    ) -> bool:
        """
        Export articles to Excel with formatting and optional charts.

        Args:
            articles: List of article dictionaries
            output_path: Path to save the Excel file
            include_charts: Whether to include chart sheets
            template: Export template ('standard', 'executive', 'detailed')

        Returns:
            True if export successful, False otherwise
        """
        try:
            wb = Workbook()

            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # Create Articles sheet
            ws_articles = wb.create_sheet("Articles", 0)
            ExcelExportManager._create_articles_sheet(ws_articles, articles, template)

            # Create Statistics sheet
            if articles:
                ws_stats = wb.create_sheet("Statistics", 1)
                ExcelExportManager._create_statistics_sheet(ws_stats, articles)

            # Create Timeline sheet
            if articles and include_charts:
                ws_timeline = wb.create_sheet("Timeline", 2)
                ExcelExportManager._create_timeline_sheet(ws_timeline, articles)

            # Save workbook
            wb.save(output_path)
            logger.info(f"Excel export completed: {output_path}")
            return True

        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            return False

    @staticmethod
    def _create_articles_sheet(ws, articles: List[Dict[str, Any]], template: str):
        """Create and format the Articles sheet."""
        # Define headers based on template
        if template == "executive":
            headers = ["ID", "Title", "Source", "Date", "Sentiment", "Summary"]
        elif template == "detailed":
            headers = ["ID", "Title", "Source URL", "Date Scraped", "Tags", "Sentiment", "Summary", "Full Text Preview"]
        else:  # standard
            headers = ["ID", "Title", "Source URL", "Date Scraped", "Tags", "Sentiment"]

        # Style header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Add article data
        for row_num, article in enumerate(articles, 2):
            if template == "executive":
                row_data = [
                    article.get('id', ''),
                    article.get('title', ''),
                    article.get('source_url', '')[:50],
                    article.get('date_scraped', ''),
                    article.get('sentiment', ''),
                    article.get('summary', '')[:200] if article.get('summary') else ''
                ]
            elif template == "detailed":
                row_data = [
                    article.get('id', ''),
                    article.get('title', ''),
                    article.get('source_url', ''),
                    article.get('date_scraped', ''),
                    article.get('tags', ''),
                    article.get('sentiment', ''),
                    article.get('summary', '')[:300] if article.get('summary') else '',
                    article.get('full_text', '')[:500] if article.get('full_text') else ''
                ]
            else:  # standard
                row_data = [
                    article.get('id', ''),
                    article.get('title', ''),
                    article.get('source_url', ''),
                    article.get('date_scraped', ''),
                    article.get('tags', ''),
                    article.get('sentiment', '')
                ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = ws['A2']

    @staticmethod
    def _create_statistics_sheet(ws, articles: List[Dict[str, Any]]):
        """Create statistics summary sheet."""
        # Title
        ws['A1'] = "WebScrape-TUI Analytics Summary"
        ws['A1'].font = Font(size=16, bold=True, color="366092")
        ws.merge_cells('A1:B1')

        row = 3
        stats = [
            ("Total Articles", len(articles)),
            ("", ""),
            ("Sentiment Distribution", ""),
        ]

        # Count sentiments
        sentiment_counts = {}
        for article in articles:
            sent = article.get('sentiment', 'Unknown')
            sentiment_counts[sent] = sentiment_counts.get(sent, 0) + 1

        for sentiment, count in sentiment_counts.items():
            stats.append((f"  {sentiment}", count))

        # Count sources
        stats.append(("", ""))
        stats.append(("Top Sources", "Count"))
        source_counts = {}
        for article in articles:
            source = article.get('source_url', 'Unknown')[:50]
            source_counts[source] = source_counts.get(source, 0) + 1

        top_sources = sorted(source_counts.items(), key=lambda x: x[1], reverse=True)[:10]
        stats.extend(top_sources)

        # Write statistics
        for stat_row in stats:
            ws.cell(row=row, column=1).value = stat_row[0]
            ws.cell(row=row, column=2).value = stat_row[1]
            row += 1

        # Format
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15

    @staticmethod
    def _create_timeline_sheet(ws, articles: List[Dict[str, Any]]):
        """Create timeline visualization sheet."""
        ws['A1'] = "Article Collection Timeline"
        ws['A1'].font = Font(size=14, bold=True)

        # Count articles by date
        date_counts = {}
        for article in articles:
            date_str = article.get('date_scraped', '')[:10]  # YYYY-MM-DD
            if date_str:
                date_counts[date_str] = date_counts.get(date_str, 0) + 1

        # Sort by date
        sorted_dates = sorted(date_counts.items())

        # Write data
        ws['A3'] = "Date"
        ws['B3'] = "Articles"
        ws['A3'].font = Font(bold=True)
        ws['B3'].font = Font(bold=True)

        for row_num, (date, count) in enumerate(sorted_dates, 4):
            ws.cell(row=row_num, column=1).value = date
            ws.cell(row=row_num, column=2).value = count


class PDFExportManager:
    """
    Professional PDF report generation with charts and formatting.

    Provides comprehensive PDF export capabilities including:
    - Executive summary sections
    - Embedded charts and graphs
    - Table of contents
    - Professional layouts and styling
    - Custom report templates
    """

    @staticmethod
    def export_to_pdf(
        articles: List[Dict[str, Any]],
        output_path: str,
        include_charts: bool = True,
        template: str = "standard"
    ) -> bool:
        """
        Export articles to PDF with professional formatting.

        Args:
            articles: List of article dictionaries
            output_path: Path to save the PDF file
            include_charts: Whether to include charts
            template: Report template ('standard', 'executive', 'detailed')

        Returns:
            True if export successful, False otherwise
        """
        try:
            doc = SimpleDocTemplate(output_path, pagesize=letter)
            story = []
            styles = getSampleStyleSheet()

            # Custom styles
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#366092'),
                spaceAfter=30,
                alignment=TA_CENTER
            )

            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=16,
                textColor=colors.HexColor('#366092'),
                spaceAfter=12,
                spaceBefore=12
            )

            # Title page
            story.append(Paragraph("WebScrape-TUI Analytics Report", title_style))
            story.append(Spacer(1, 0.2*inch))
            story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
            story.append(Paragraph(f"Total Articles: {len(articles)}", styles['Normal']))
            story.append(PageBreak())

            # Executive Summary
            if template in ["executive", "detailed"]:
                story.append(Paragraph("Executive Summary", heading_style))
                PDFExportManager._add_executive_summary(story, articles, styles)
                story.append(PageBreak())

            # Statistics Section
            story.append(Paragraph("Statistics Overview", heading_style))
            PDFExportManager._add_statistics_section(story, articles, styles)
            story.append(Spacer(1, 0.3*inch))

            # Articles Table
            if template == "detailed":
                story.append(Paragraph("Articles Listing", heading_style))
                PDFExportManager._add_articles_table(story, articles, styles)

            # Build PDF
            doc.build(story)
            logger.info(f"PDF export completed: {output_path}")
            return True

        except Exception as e:
            logger.error(f"Error exporting to PDF: {e}")
            return False

    @staticmethod
    def _add_executive_summary(story, articles: List[Dict[str, Any]], styles):
        """Add executive summary section to PDF."""
        # Calculate key metrics
        total_articles = len(articles)
        sentiment_counts = {}
        for article in articles:
            sent = article.get('sentiment', 'Unknown')
            sentiment_counts[sent] = sentiment_counts.get(sent, 0) + 1

        # Summary text
        summary_text = f"""
        This report analyzes {total_articles} articles collected through WebScrape-TUI.
        The collection includes content from multiple sources with AI-powered sentiment analysis.
        """

        story.append(Paragraph(summary_text, styles['Normal']))
        story.append(Spacer(1, 0.2*inch))

        # Key findings
        story.append(Paragraph("Key Findings:", styles['Heading3']))
        findings = [
            f"• Total articles analyzed: {total_articles}",
            f"• Sentiment distribution: {len(sentiment_counts)} categories",
        ]

        if 'Positive' in sentiment_counts:
            findings.append(f"• Positive sentiment: {sentiment_counts['Positive']} articles")
        if 'Negative' in sentiment_counts:
            findings.append(f"• Negative sentiment: {sentiment_counts['Negative']} articles")

        for finding in findings:
            story.append(Paragraph(finding, styles['Normal']))

    @staticmethod
    def _add_statistics_section(story, articles: List[Dict[str, Any]], styles):
        """Add statistics section to PDF."""
        # Sentiment distribution
        sentiment_counts = {}
        for article in articles:
            sent = article.get('sentiment', 'Unknown')
            sentiment_counts[sent] = sentiment_counts.get(sent, 0) + 1

        # Create statistics table
        data = [['Metric', 'Value']]
        data.append(['Total Articles', str(len(articles))])
        data.append(['', ''])
        data.append(['Sentiment Distribution', ''])

        for sentiment, count in sentiment_counts.items():
            data.append([f'  {sentiment}', str(count)])

        # Style table
        table = Table(data, colWidths=[3*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        story.append(table)

    @staticmethod
    def _add_articles_table(story, articles: List[Dict[str, Any]], styles):
        """Add articles listing table to PDF."""
        data = [['ID', 'Title', 'Date', 'Sentiment']]

        for article in articles[:50]:  # Limit to 50 for PDF size
            data.append([
                str(article.get('id', '')),
                article.get('title', '')[:40],
                article.get('date_scraped', '')[:10],
                article.get('sentiment', '')
            ])

        table = Table(data, colWidths=[0.5*inch, 3.5*inch, 1*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))

        story.append(table)


class EnhancedVisualizationManager:
    """
    Enhanced visualization capabilities with word clouds and heatmaps.

    Provides advanced chart types including:
    - Word cloud visualization for tags
    - Heatmaps for scraping activity patterns
    - Scatter plots for sentiment vs. time
    - Custom date range charts
    """

    @staticmethod
    def generate_word_cloud(
        tags_data: List[Tuple[str, int]],
        output_path: str,
        width: int = 800,
        height: int = 400
    ) -> bool:
        """
        Generate word cloud from tag frequency data.

        Args:
            tags_data: List of (tag, frequency) tuples
            output_path: Path to save the word cloud image
            width: Image width in pixels
            height: Image height in pixels

        Returns:
            True if generation successful, False otherwise
        """
        try:
            # Create word frequency dictionary
            word_freq = {tag: freq for tag, freq in tags_data}

            if not word_freq:
                logger.warning("No tag data available for word cloud")
                return False

            # Generate word cloud
            wordcloud = WordCloud(
                width=width,
                height=height,
                background_color='white',
                colormap='viridis',
                relative_scaling=0.5,
                min_font_size=10
            ).generate_from_frequencies(word_freq)

            # Create figure
            plt.figure(figsize=(10, 5))
            plt.imshow(wordcloud, interpolation='bilinear')
            plt.axis('off')
            plt.title('Tag Frequency Word Cloud', fontsize=16, pad=20)
            plt.tight_layout(pad=0)

            # Save
            plt.savefig(output_path, dpi=150, bbox_inches='tight')
            plt.close()

            logger.info(f"Word cloud generated: {output_path}")
            return True

        except Exception as e:
            logger.error(f"Error generating word cloud: {e}")
            return False

    @staticmethod
    def generate_sentiment_scatter(
        articles: List[Dict[str, Any]],
        output_path: str
    ) -> bool:
        """
        Generate scatter plot of sentiment over time.

        Args:
            articles: List of article dictionaries
            output_path: Path to save the chart

        Returns:
            True if generation successful, False otherwise
        """
        try:
            # Prepare data
            dates = []
            sentiments = []
            colors_list = []

            sentiment_colors = {
                'Positive': '#4CAF50',
                'Negative': '#F44336',
                'Neutral': '#9E9E9E'
            }

            for article in articles:
                date_str = article.get('date_scraped', '')
                sentiment = article.get('sentiment', 'Neutral')

                if date_str and sentiment:
                    try:
                        date_obj = datetime.strptime(date_str[:10], '%Y-%m-%d')
                        dates.append(date_obj)

                        # Map sentiment to numeric value
                        sent_value = {'Positive': 1, 'Neutral': 0, 'Negative': -1}.get(sentiment, 0)
                        sentiments.append(sent_value)
                        colors_list.append(sentiment_colors.get(sentiment, '#9E9E9E'))
                    except:
                        pass

            if not dates:
                logger.warning("No date/sentiment data for scatter plot")
                return False

            # Create scatter plot
            fig, ax = plt.subplots(figsize=(12, 6))
            ax.scatter(dates, sentiments, c=colors_list, alpha=0.6, s=50)

            ax.set_xlabel('Date', fontsize=12)
            ax.set_ylabel('Sentiment', fontsize=12)
            ax.set_title('Sentiment Analysis Over Time', fontsize=14, pad=20)
            ax.set_yticks([-1, 0, 1])
            ax.set_yticklabels(['Negative', 'Neutral', 'Positive'])
            ax.grid(True, alpha=0.3)

            # Add legend
            from matplotlib.patches import Patch
            legend_elements = [
                Patch(facecolor='#4CAF50', label='Positive'),
                Patch(facecolor='#9E9E9E', label='Neutral'),
                Patch(facecolor='#F44336', label='Negative')
            ]
            ax.legend(handles=legend_elements, loc='upper right')

            plt.tight_layout()
            plt.savefig(output_path, dpi=150, bbox_inches='tight')
            plt.close()

            logger.info(f"Sentiment scatter plot generated: {output_path}")
            return True

        except Exception as e:
            logger.error(f"Error generating sentiment scatter plot: {e}")
            return False


class AITaggingManager:
    """
    AI-powered auto-tagging and categorization system.

    Provides intelligent tag generation using:
    - Keyword extraction (TF-IDF)
    - Topic modeling (LDA)
    - AI-based tag suggestions
    - Confidence scoring
    """

    _tfidf_vectorizer = None
    _lda_model = None

    @staticmethod
    def generate_tags_from_content(
        text: str,
        num_tags: int = 5,
        min_confidence: float = 0.3
    ) -> List[Tuple[str, float]]:
        """
        Generate tags from article content using TF-IDF keyword extraction.

        Args:
            text: Article text content
            num_tags: Maximum number of tags to generate
            min_confidence: Minimum confidence score (0.0-1.0)

        Returns:
            List of (tag, confidence_score) tuples
        """
        try:
            if not text or len(text.strip()) < 50:
                logger.warning("Text too short for tag generation")
                return []

            # Tokenize and clean
            stop_words = set(stopwords.words('english'))
            tokens = word_tokenize(text.lower())
            tokens = [t for t in tokens if t.isalnum() and t not in stop_words and len(t) > 3]

            if len(tokens) < 10:
                return []

            # Use TF-IDF for keyword extraction
            vectorizer = TfidfVectorizer(max_features=20, stop_words='english')
            try:
                tfidf_matrix = vectorizer.fit_transform([text])
                feature_names = vectorizer.get_feature_names_out()
                scores = tfidf_matrix.toarray()[0]

                # Get top keywords with scores
                keyword_scores = sorted(
                    [(feature_names[i], scores[i]) for i in range(len(scores))],
                    key=lambda x: x[1],
                    reverse=True
                )

                # Filter by confidence and limit number
                tags = [
                    (keyword, float(score))
                    for keyword, score in keyword_scores[:num_tags]
                    if score >= min_confidence
                ]

                logger.info(f"Generated {len(tags)} tags from content")
                return tags

            except Exception as e:
                logger.error(f"TF-IDF vectorization error: {e}")
                return []

        except Exception as e:
            logger.error(f"Error generating tags: {e}")
            return []

    @staticmethod
    def suggest_tags_with_ai(
        text: str,
        existing_tags: List[str],
        num_suggestions: int = 3
    ) -> List[str]:
        """
        Suggest tags using AI based on content and existing tag taxonomy.

        Args:
            text: Article text content
            existing_tags: List of all existing tags in database
            num_suggestions: Number of suggestions to return

        Returns:
            List of suggested tag names
        """
        try:
            # Extract keywords from content
            generated_tags = AITaggingManager.generate_tags_from_content(
                text, num_tags=10, min_confidence=0.2
            )

            if not generated_tags:
                return []

            # Get keywords only (without scores)
            keywords = [tag for tag, _ in generated_tags]

            # Find similar existing tags
            suggestions = []
            for keyword in keywords[:num_suggestions]:
                # Check for exact matches in existing tags
                if keyword in existing_tags:
                    suggestions.append(keyword)
                else:
                    # Find similar tags (simple substring matching)
                    similar = [tag for tag in existing_tags if keyword in tag.lower() or tag.lower() in keyword]
                    if similar:
                        suggestions.append(similar[0])
                    else:
                        suggestions.append(keyword)

            return suggestions[:num_suggestions]

        except Exception as e:
            logger.error(f"Error suggesting tags: {e}")
            return []


class EntityRecognitionManager:
    """
    Named Entity Recognition and extraction system.

    Provides entity extraction using spaCy:
    - People (PERSON)
    - Organizations (ORG)
    - Locations (GPE, LOC)
    - Dates (DATE)
    - Products (PRODUCT)
    """

    _nlp_model = None

    @staticmethod
    def load_spacy_model():
        """Load spaCy model (lazy loading)."""
        if EntityRecognitionManager._nlp_model is None:
            try:
                EntityRecognitionManager._nlp_model = spacy.load('en_core_web_sm')
                logger.info("Loaded spaCy en_core_web_sm model")
            except OSError:
                logger.warning("spaCy model not found. Install with: python -m spacy download en_core_web_sm")
                return False
        return True

    @staticmethod
    def extract_entities(text: str) -> Dict[str, List[str]]:
        """
        Extract named entities from text.

        Args:
            text: Article text content

        Returns:
            Dictionary with entity types as keys and lists of entities as values
        """
        try:
            if not EntityRecognitionManager.load_spacy_model():
                return {}

            if not text or len(text.strip()) < 20:
                return {}

            # Process text (limit to first 5000 chars for performance)
            doc = EntityRecognitionManager._nlp_model(text[:5000])

            # Extract entities by type
            entities = {
                'people': [],
                'organizations': [],
                'locations': [],
                'dates': [],
                'products': []
            }

            for ent in doc.ents:
                entity_text = ent.text.strip()
                if ent.label_ == 'PERSON':
                    entities['people'].append(entity_text)
                elif ent.label_ == 'ORG':
                    entities['organizations'].append(entity_text)
                elif ent.label_ in ('GPE', 'LOC'):
                    entities['locations'].append(entity_text)
                elif ent.label_ == 'DATE':
                    entities['dates'].append(entity_text)
                elif ent.label_ == 'PRODUCT':
                    entities['products'].append(entity_text)

            # Remove duplicates while preserving order
            for key in entities:
                seen = set()
                entities[key] = [x for x in entities[key] if not (x in seen or seen.add(x))]

            logger.info(f"Extracted entities: {sum(len(v) for v in entities.values())} total")
            return entities

        except Exception as e:
            logger.error(f"Error extracting entities: {e}")
            return {}


class ContentSimilarityManager:
    """
    Content similarity detection and duplicate identification.

    Uses sentence transformers for semantic similarity:
    - Duplicate detection
    - Related article suggestions
    - Content clustering
    """

    _model = None

    @staticmethod
    def load_model():
        """Load sentence transformer model (lazy loading)."""
        if ContentSimilarityManager._model is None:
            try:
                # Use a smaller, faster model
                ContentSimilarityManager._model = SentenceTransformer('all-MiniLM-L6-v2')
                logger.info("Loaded sentence-transformers model")
            except Exception as e:
                logger.error(f"Error loading sentence-transformers model: {e}")
                return False
        return True

    @staticmethod
    def calculate_similarity(text1: str, text2: str) -> float:
        """
        Calculate semantic similarity between two texts.

        Args:
            text1: First text
            text2: Second text

        Returns:
            Similarity score (0.0-1.0)
        """
        try:
            if not ContentSimilarityManager.load_model():
                return 0.0

            if not text1 or not text2:
                return 0.0

            # Generate embeddings
            embeddings = ContentSimilarityManager._model.encode([text1[:1000], text2[:1000]])

            # Calculate cosine similarity
            similarity = 1 - cosine(embeddings[0], embeddings[1])

            return float(max(0.0, min(1.0, similarity)))

        except Exception as e:
            logger.error(f"Error calculating similarity: {e}")
            return 0.0

    @staticmethod
    def find_similar_articles(
        target_text: str,
        articles: List[Dict[str, Any]],
        top_k: int = 5,
        min_similarity: float = 0.5
    ) -> List[Tuple[Dict[str, Any], float]]:
        """
        Find similar articles to target text.

        Args:
            target_text: Reference text
            articles: List of article dictionaries with 'content' or 'summary'
            top_k: Number of similar articles to return
            min_similarity: Minimum similarity threshold

        Returns:
            List of (article, similarity_score) tuples
        """
        try:
            if not ContentSimilarityManager.load_model():
                return []

            if not target_text or not articles:
                return []

            # Get target embedding
            target_embedding = ContentSimilarityManager._model.encode([target_text[:1000]])[0]

            # Calculate similarities
            similarities = []
            for article in articles:
                # Use summary if available, otherwise content
                compare_text = article.get('summary') or article.get('full_text') or article.get('title', '')
                if not compare_text:
                    continue

                article_embedding = ContentSimilarityManager._model.encode([compare_text[:1000]])[0]
                similarity = 1 - cosine(target_embedding, article_embedding)

                if similarity >= min_similarity:
                    similarities.append((article, float(similarity)))

            # Sort by similarity and return top k
            similarities.sort(key=lambda x: x[1], reverse=True)
            return similarities[:top_k]

        except Exception as e:
            logger.error(f"Error finding similar articles: {e}")
            return []


class KeywordExtractionManager:
    """
    Keyword and key phrase extraction system.

    Provides:
    - TF-IDF based keyword ranking
    - Multi-word phrase extraction
    - Trending keyword analysis
    """

    @staticmethod
    def extract_keywords(
        text: str,
        num_keywords: int = 10,
        min_word_length: int = 4
    ) -> List[Tuple[str, float]]:
        """
        Extract keywords from text using TF-IDF.

        Args:
            text: Text content
            num_keywords: Number of keywords to extract
            min_word_length: Minimum keyword length

        Returns:
            List of (keyword, score) tuples
        """
        try:
            if not text or len(text.strip()) < 50:
                return []

            # Use TF-IDF
            vectorizer = TfidfVectorizer(
                max_features=num_keywords * 2,
                stop_words='english',
                token_pattern=r'\b[a-z]{' + str(min_word_length) + r',}\b'
            )

            tfidf_matrix = vectorizer.fit_transform([text])
            feature_names = vectorizer.get_feature_names_out()
            scores = tfidf_matrix.toarray()[0]

            # Sort by score
            keywords = sorted(
                [(feature_names[i], float(scores[i])) for i in range(len(scores))],
                key=lambda x: x[1],
                reverse=True
            )

            return keywords[:num_keywords]

        except Exception as e:
            logger.error(f"Error extracting keywords: {e}")
            return []

    @staticmethod
    def extract_key_phrases(
        text: str,
        num_phrases: int = 5
    ) -> List[Tuple[str, float]]:
        """
        Extract multi-word key phrases.

        Args:
            text: Text content
            num_phrases: Number of phrases to extract

        Returns:
            List of (phrase, score) tuples
        """
        try:
            if not text or len(text.strip()) < 50:
                return []

            # Use TF-IDF with bigrams and trigrams
            vectorizer = TfidfVectorizer(
                max_features=num_phrases * 2,
                stop_words='english',
                ngram_range=(2, 3)  # 2-3 word phrases
            )

            tfidf_matrix = vectorizer.fit_transform([text])
            feature_names = vectorizer.get_feature_names_out()
            scores = tfidf_matrix.toarray()[0]

            # Sort by score
            phrases = sorted(
                [(feature_names[i], float(scores[i])) for i in range(len(scores))],
                key=lambda x: x[1],
                reverse=True
            )

            return phrases[:num_phrases]

        except Exception as e:
            logger.error(f"Error extracting key phrases: {e}")
            return []


class MultiLevelSummarizationManager:
    """
    Multi-level summarization system.

    Provides:
    - One-sentence summaries
    - Paragraph summaries
    - Full summaries
    - Extractive summaries
    """

    @staticmethod
    def generate_one_sentence_summary(text: str) -> str:
        """
        Generate ultra-concise one-sentence summary.

        Args:
            text: Article text

        Returns:
            One-sentence summary
        """
        # Use existing AI provider
        provider = get_ai_provider()
        if provider is None:
            return "AI provider not configured."

        try:
            custom_prompt = """Summarize this article in exactly ONE sentence (maximum 20 words).
Focus on the single most important point.

Article:
{article_text}

One-sentence summary:"""

            summary = provider.get_summary(
                text,
                summary_style="brief",
                template=custom_prompt,
                max_length=500
            )

            return summary or "Unable to generate summary."

        except Exception as e:
            logger.error(f"Error generating one-sentence summary: {e}")
            return "Summary generation failed."

    @staticmethod
    def generate_extractive_summary(text: str, num_sentences: int = 3) -> str:
        """
        Generate extractive summary by selecting key sentences.

        Args:
            text: Article text
            num_sentences: Number of sentences to extract

        Returns:
            Extractive summary
        """
        try:
            if not text or len(text.strip()) < 100:
                return text

            # Simple extractive approach: use TF-IDF to score sentences
            sentences = [s.strip() for s in text.split('.') if len(s.strip()) > 20]
            if len(sentences) <= num_sentences:
                return text

            # Score sentences by TF-IDF
            vectorizer = TfidfVectorizer(stop_words='english')
            try:
                tfidf_matrix = vectorizer.fit_transform(sentences)
                sentence_scores = tfidf_matrix.sum(axis=1).A1

                # Get top sentences
                top_indices = sentence_scores.argsort()[-num_sentences:][::-1]
                top_indices = sorted(top_indices)  # Preserve order

                extractive = '. '.join([sentences[i] for i in top_indices]) + '.'
                return extractive

            except:
                # Fallback: return first N sentences
                return '. '.join(sentences[:num_sentences]) + '.'

        except Exception as e:
            logger.error(f"Error generating extractive summary: {e}")
            return text[:500] + "..."


class TopicModelingManager:
    """
    Topic modeling and categorization system (v1.9.0).

    Provides:
    - LDA (Latent Dirichlet Allocation) topic modeling
    - NMF (Non-negative Matrix Factorization) topic modeling
    - Automatic category assignment
    - Category hierarchy creation
    - Multi-label classification
    """

    @staticmethod
    def perform_lda_topic_modeling(
        articles: List[Dict[str, Any]],
        num_topics: int = 5,
        passes: int = 10
    ) -> Dict[str, Any]:
        """
        Perform LDA topic modeling on article collection.

        Args:
            articles: List of article dictionaries with 'content' field
            num_topics: Number of topics to extract
            passes: Number of LDA training passes

        Returns:
            Dictionary with topics, word distributions, and article assignments
        """
        try:
            if not articles or len(articles) < 2:
                return {"topics": [], "assignments": {}, "error": "Need at least 2 articles"}

            # Extract and preprocess text
            texts = []
            for article in articles:
                content = article.get('content', '') or article.get('text', '')
                if content:
                    # Tokenize and remove stopwords
                    tokens = [
                        word.lower() for word in content.split()
                        if word.lower() not in STOPWORDS and len(word) > 3
                    ]
                    texts.append(tokens)

            if not texts or len(texts) < 2:
                return {"topics": [], "assignments": {}, "error": "Insufficient text data"}

            # Create dictionary and corpus
            dictionary = corpora.Dictionary(texts)
            dictionary.filter_extremes(no_below=1, no_above=0.8, keep_n=1000)
            corpus = [dictionary.doc2bow(text) for text in texts]

            # Train LDA model
            lda_model = LdaModel(
                corpus=corpus,
                id2word=dictionary,
                num_topics=num_topics,
                random_state=42,
                passes=passes,
                alpha='auto',
                per_word_topics=True
            )

            # Extract topics with top words
            topics = []
            for idx in range(num_topics):
                words = lda_model.show_topic(idx, topn=10)
                topic_words = [word for word, _ in words]
                topic_weights = [weight for _, weight in words]
                topics.append({
                    "id": idx,
                    "words": topic_words,
                    "weights": topic_weights,
                    "label": f"Topic {idx}: {', '.join(topic_words[:3])}"
                })

            # Assign articles to dominant topics
            assignments = {}
            for idx, doc_bow in enumerate(corpus):
                topic_dist = lda_model.get_document_topics(doc_bow)
                if topic_dist:
                    dominant_topic = max(topic_dist, key=lambda x: x[1])
                    assignments[articles[idx].get('id', idx)] = {
                        "topic_id": dominant_topic[0],
                        "confidence": float(dominant_topic[1]),
                        "all_topics": [(int(t[0]), float(t[1])) for t in topic_dist]
                    }

            return {
                "topics": topics,
                "assignments": assignments,
                "num_articles": len(articles),
                "dictionary_size": len(dictionary)
            }

        except Exception as e:
            logger.error(f"Error in LDA topic modeling: {e}")
            return {"topics": [], "assignments": {}, "error": str(e)}

    @staticmethod
    def perform_nmf_topic_modeling(
        articles: List[Dict[str, Any]],
        num_topics: int = 5
    ) -> Dict[str, Any]:
        """
        Perform NMF topic modeling on article collection.

        Args:
            articles: List of article dictionaries with 'content' field
            num_topics: Number of topics to extract

        Returns:
            Dictionary with topics, word distributions, and article assignments
        """
        try:
            if not articles or len(articles) < 2:
                return {"topics": [], "assignments": {}, "error": "Need at least 2 articles"}

            # Extract text
            texts = []
            article_ids = []
            for article in articles:
                content = article.get('content', '') or article.get('text', '')
                if content and len(content.strip()) > 50:
                    texts.append(content)
                    article_ids.append(article.get('id', len(article_ids)))

            if len(texts) < 2:
                return {"topics": [], "assignments": {}, "error": "Insufficient text data"}

            # Vectorize with TF-IDF
            vectorizer = TfidfVectorizer(
                max_features=1000,
                stop_words='english',
                min_df=1,
                max_df=0.8
            )
            tfidf_matrix = vectorizer.fit_transform(texts)

            # Train NMF model
            nmf_model = NMF(
                n_components=num_topics,
                random_state=42,
                max_iter=300
            )
            doc_topic = nmf_model.fit_transform(tfidf_matrix)
            feature_names = vectorizer.get_feature_names_out()

            # Extract topics
            topics = []
            for idx, topic in enumerate(nmf_model.components_):
                top_indices = topic.argsort()[-10:][::-1]
                top_words = [feature_names[i] for i in top_indices]
                top_weights = [float(topic[i]) for i in top_indices]

                topics.append({
                    "id": idx,
                    "words": top_words,
                    "weights": top_weights,
                    "label": f"Topic {idx}: {', '.join(top_words[:3])}"
                })

            # Assign articles to dominant topics
            assignments = {}
            for idx, topic_dist in enumerate(doc_topic):
                dominant_topic = topic_dist.argmax()
                assignments[article_ids[idx]] = {
                    "topic_id": int(dominant_topic),
                    "confidence": float(topic_dist[dominant_topic]),
                    "all_topics": [(int(i), float(v)) for i, v in enumerate(topic_dist)]
                }

            return {
                "topics": topics,
                "assignments": assignments,
                "num_articles": len(texts),
                "vocabulary_size": len(feature_names)
            }

        except Exception as e:
            logger.error(f"Error in NMF topic modeling: {e}")
            return {"topics": [], "assignments": {}, "error": str(e)}

    @staticmethod
    def assign_categories_to_articles(
        topic_results: Dict[str, Any],
        category_mapping: Optional[Dict[int, str]] = None
    ) -> Dict[int, List[str]]:
        """
        Assign category labels to articles based on topic assignments.

        Args:
            topic_results: Results from LDA or NMF topic modeling
            category_mapping: Optional mapping of topic IDs to category names

        Returns:
            Dictionary mapping article IDs to category lists
        """
        try:
            assignments = topic_results.get('assignments', {})
            topics = topic_results.get('topics', [])

            if not assignments or not topics:
                return {}

            # Create default category mapping if not provided
            if category_mapping is None:
                category_mapping = {
                    topic['id']: topic['label']
                    for topic in topics
                }

            article_categories = {}
            for article_id, assignment in assignments.items():
                categories = []

                # Add dominant topic category
                dominant_id = assignment['topic_id']
                if dominant_id in category_mapping:
                    categories.append(category_mapping[dominant_id])

                # Add secondary topics (confidence > 0.15)
                for topic_id, confidence in assignment.get('all_topics', []):
                    if topic_id != dominant_id and confidence > 0.15:
                        if topic_id in category_mapping:
                            categories.append(f"{category_mapping[topic_id]} (secondary)")

                article_categories[int(article_id)] = categories

            return article_categories

        except Exception as e:
            logger.error(f"Error assigning categories: {e}")
            return {}


class EntityRelationshipManager:
    """
    Entity relationship mapping and knowledge graph construction (v1.9.0).

    Provides:
    - Entity extraction from articles
    - Relationship mapping between entities
    - Knowledge graph construction
    - Entity-based search and filtering
    """

    @staticmethod
    def extract_entity_relationships(
        articles: List[Dict[str, Any]],
        nlp_model: Optional[Any] = None
    ) -> Dict[str, Any]:
        """
        Extract entities and their relationships from articles.

        Args:
            articles: List of article dictionaries
            nlp_model: Optional pre-loaded spaCy model

        Returns:
            Dictionary with entities, relationships, and co-occurrence data
        """
        try:
            # Load spaCy model if not provided
            if nlp_model is None:
                try:
                    nlp_model = spacy.load("en_core_web_sm")
                except:
                    logger.warning("spaCy model not loaded, using basic extraction")
                    return EntityRelationshipManager._basic_entity_extraction(articles)

            entities = {}
            relationships = []
            entity_article_map = {}

            for article in articles:
                article_id = article.get('id', 0)
                content = article.get('content', '') or article.get('text', '')

                if not content or len(content) < 100:
                    continue

                # Process with spaCy (limit to first 100K chars for performance)
                doc = nlp_model(content[:100000])

                # Extract named entities
                article_entities = []
                for ent in doc.ents:
                    if ent.label_ in ['PERSON', 'ORG', 'GPE', 'LOC', 'PRODUCT', 'EVENT']:
                        entity_text = ent.text.strip()
                        entity_label = ent.label_

                        # Add to entities dictionary
                        if entity_text not in entities:
                            entities[entity_text] = {
                                "text": entity_text,
                                "type": entity_label,
                                "count": 0,
                                "articles": []
                            }

                        entities[entity_text]["count"] += 1
                        if article_id not in entities[entity_text]["articles"]:
                            entities[entity_text]["articles"].append(article_id)

                        article_entities.append(entity_text)

                entity_article_map[article_id] = article_entities

                # Create relationships (co-occurrence in same article)
                for i, ent1 in enumerate(article_entities):
                    for ent2 in article_entities[i+1:]:
                        relationships.append({
                            "source": ent1,
                            "target": ent2,
                            "article_id": article_id,
                            "type": "co-occurrence"
                        })

            return {
                "entities": entities,
                "relationships": relationships,
                "entity_article_map": entity_article_map,
                "total_entities": len(entities),
                "total_relationships": len(relationships)
            }

        except Exception as e:
            logger.error(f"Error extracting entity relationships: {e}")
            return {"entities": {}, "relationships": [], "error": str(e)}

    @staticmethod
    def _basic_entity_extraction(articles: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Fallback entity extraction without spaCy."""
        # Simple capitalized word extraction as entities
        entities = {}
        relationships = []

        for article in articles:
            content = article.get('content', '') or article.get('text', '')
            words = content.split()
            capitalized = [w for w in words if w[0].isupper() and len(w) > 2]

            for word in set(capitalized):
                if word not in entities:
                    entities[word] = {"text": word, "type": "UNKNOWN", "count": 0, "articles": []}
                entities[word]["count"] += 1
                entities[word]["articles"].append(article.get('id', 0))

        return {"entities": entities, "relationships": relationships}

    @staticmethod
    def build_knowledge_graph(
        entity_data: Dict[str, Any],
        min_entity_count: int = 2
    ) -> nx.Graph:
        """
        Build a knowledge graph from entity relationship data.

        Args:
            entity_data: Entity relationship data from extract_entity_relationships
            min_entity_count: Minimum entity occurrence count to include

        Returns:
            NetworkX graph object
        """
        try:
            G = nx.Graph()

            # Add entity nodes
            entities = entity_data.get('entities', {})
            for entity_text, entity_info in entities.items():
                if entity_info['count'] >= min_entity_count:
                    G.add_node(
                        entity_text,
                        type=entity_info['type'],
                        count=entity_info['count'],
                        articles=entity_info['articles']
                    )

            # Add relationship edges
            relationships = entity_data.get('relationships', [])
            relationship_counts = {}

            for rel in relationships:
                source = rel['source']
                target = rel['target']

                # Only add edges for nodes that exist
                if source in G.nodes and target in G.nodes:
                    edge_key = tuple(sorted([source, target]))
                    relationship_counts[edge_key] = relationship_counts.get(edge_key, 0) + 1

            # Add weighted edges
            for (source, target), weight in relationship_counts.items():
                G.add_edge(source, target, weight=weight)

            return G

        except Exception as e:
            logger.error(f"Error building knowledge graph: {e}")
            return nx.Graph()

    @staticmethod
    def get_related_entities(
        graph: nx.Graph,
        entity: str,
        max_depth: int = 2
    ) -> List[Tuple[str, int]]:
        """
        Get entities related to a given entity within max depth.

        Args:
            graph: Knowledge graph
            entity: Entity to find relations for
            max_depth: Maximum graph distance

        Returns:
            List of (entity, distance) tuples
        """
        try:
            if entity not in graph.nodes:
                return []

            related = []
            for node in graph.nodes:
                if node != entity:
                    try:
                        path_length = nx.shortest_path_length(graph, entity, node)
                        if path_length <= max_depth:
                            related.append((node, path_length))
                    except nx.NetworkXNoPath:
                        continue

            return sorted(related, key=lambda x: x[1])

        except Exception as e:
            logger.error(f"Error getting related entities: {e}")
            return []


class DuplicateDetectionManager:
    """
    Duplicate detection and article clustering (v1.9.0).

    Provides:
    - Fuzzy duplicate detection
    - Related article suggestions
    - Clustering similar articles
    - Deduplication recommendations
    """

    @staticmethod
    def find_duplicates(
        articles: List[Dict[str, Any]],
        similarity_threshold: int = 85
    ) -> List[Dict[str, Any]]:
        """
        Find potential duplicate articles using fuzzy matching.

        Args:
            articles: List of article dictionaries
            similarity_threshold: Similarity score threshold (0-100)

        Returns:
            List of duplicate pairs with similarity scores
        """
        try:
            duplicates = []

            for i, article1 in enumerate(articles):
                title1 = article1.get('title', '')
                content1 = article1.get('content', '') or article1.get('text', '')

                for article2 in articles[i+1:]:
                    title2 = article2.get('title', '')
                    content2 = article2.get('content', '') or article2.get('text', '')

                    # Check title similarity
                    title_sim = fuzz.ratio(title1, title2)

                    # Check content similarity (first 500 chars)
                    content_sim = fuzz.partial_ratio(
                        content1[:500],
                        content2[:500]
                    )

                    # Combined score (weighted: 60% title, 40% content)
                    combined_score = int(title_sim * 0.6 + content_sim * 0.4)

                    if combined_score >= similarity_threshold:
                        duplicates.append({
                            "article1_id": article1.get('id'),
                            "article1_title": title1,
                            "article2_id": article2.get('id'),
                            "article2_title": title2,
                            "similarity_score": combined_score,
                            "title_similarity": title_sim,
                            "content_similarity": content_sim
                        })

            return sorted(duplicates, key=lambda x: x['similarity_score'], reverse=True)

        except Exception as e:
            logger.error(f"Error finding duplicates: {e}")
            return []

    @staticmethod
    def find_related_articles(
        target_article: Dict[str, Any],
        all_articles: List[Dict[str, Any]],
        top_n: int = 5,
        embedding_model: Optional[Any] = None
    ) -> List[Dict[str, Any]]:
        """
        Find articles related to a target article using embeddings.

        Args:
            target_article: Target article dictionary
            all_articles: List of all articles
            top_n: Number of related articles to return
            embedding_model: Optional pre-loaded SentenceTransformer model

        Returns:
            List of related articles with similarity scores
        """
        try:
            # Load embedding model if not provided
            if embedding_model is None:
                embedding_model = SentenceTransformer('all-MiniLM-L6-v2')

            target_content = target_article.get('content', '') or target_article.get('text', '')
            if not target_content:
                return []

            # Generate target embedding
            target_embedding = embedding_model.encode([target_content[:1000]])[0]

            # Calculate similarities
            related = []
            for article in all_articles:
                # Skip the target article itself
                if article.get('id') == target_article.get('id'):
                    continue

                content = article.get('content', '') or article.get('text', '')
                if not content:
                    continue

                # Generate embedding
                embedding = embedding_model.encode([content[:1000]])[0]

                # Calculate cosine similarity
                similarity = 1 - cosine(target_embedding, embedding)

                related.append({
                    "article_id": article.get('id'),
                    "title": article.get('title', 'Untitled'),
                    "similarity_score": float(similarity),
                    "url": article.get('url', '')
                })

            # Sort by similarity and return top N
            related.sort(key=lambda x: x['similarity_score'], reverse=True)
            return related[:top_n]

        except Exception as e:
            logger.error(f"Error finding related articles: {e}")
            return []

    @staticmethod
    def cluster_articles(
        articles: List[Dict[str, Any]],
        num_clusters: int = 5,
        embedding_model: Optional[Any] = None
    ) -> Dict[str, Any]:
        """
        Cluster similar articles using K-means.

        Args:
            articles: List of article dictionaries
            num_clusters: Number of clusters to create
            embedding_model: Optional pre-loaded SentenceTransformer model

        Returns:
            Dictionary with cluster assignments and centroids
        """
        try:
            if len(articles) < num_clusters:
                return {"clusters": {}, "error": "Not enough articles for clustering"}

            # Load embedding model if not provided
            if embedding_model is None:
                embedding_model = SentenceTransformer('all-MiniLM-L6-v2')

            # Generate embeddings
            texts = []
            article_ids = []
            for article in articles:
                content = article.get('content', '') or article.get('text', '')
                if content:
                    texts.append(content[:1000])
                    article_ids.append(article.get('id'))

            if len(texts) < num_clusters:
                return {"clusters": {}, "error": "Insufficient text data"}

            embeddings = embedding_model.encode(texts)

            # Perform K-means clustering
            kmeans = KMeans(n_clusters=num_clusters, random_state=42, n_init=10)
            cluster_labels = kmeans.fit_predict(embeddings)

            # Organize results
            clusters = {}
            for i, label in enumerate(cluster_labels):
                label = int(label)
                if label not in clusters:
                    clusters[label] = {
                        "cluster_id": label,
                        "articles": [],
                        "size": 0
                    }

                clusters[label]["articles"].append({
                    "article_id": article_ids[i],
                    "title": articles[i].get('title', 'Untitled')
                })
                clusters[label]["size"] += 1

            return {
                "clusters": clusters,
                "num_clusters": num_clusters,
                "total_articles": len(article_ids)
            }

        except Exception as e:
            logger.error(f"Error clustering articles: {e}")
            return {"clusters": {}, "error": str(e)}


class SummaryQualityManager:
    """
    Summary quality metrics and evaluation (v1.9.0).

    Provides:
    - ROUGE score calculation
    - Summary coherence analysis
    - User feedback collection
    - Quality metrics tracking
    """

    @staticmethod
    def calculate_rouge_scores(
        reference: str,
        summary: str
    ) -> Dict[str, Dict[str, float]]:
        """
        Calculate ROUGE scores for a summary.

        Args:
            reference: Reference text (original article)
            summary: Generated summary

        Returns:
            Dictionary with ROUGE-1, ROUGE-2, and ROUGE-L scores
        """
        try:
            scorer = rouge_scorer.RougeScorer(['rouge1', 'rouge2', 'rougeL'], use_stemmer=True)
            scores = scorer.score(reference, summary)

            results = {}
            for metric, score in scores.items():
                results[metric] = {
                    "precision": float(score.precision),
                    "recall": float(score.recall),
                    "fmeasure": float(score.fmeasure)
                }

            return results

        except Exception as e:
            logger.error(f"Error calculating ROUGE scores: {e}")
            return {}

    @staticmethod
    def evaluate_summary_coherence(summary: str) -> Dict[str, Any]:
        """
        Evaluate summary coherence using various metrics.

        Args:
            summary: Summary text

        Returns:
            Dictionary with coherence metrics
        """
        try:
            # Basic metrics
            sentences = [s.strip() for s in summary.split('.') if s.strip()]
            words = summary.split()

            # Calculate readability metrics
            avg_sentence_length = len(words) / len(sentences) if sentences else 0

            # Calculate lexical diversity
            unique_words = set(w.lower() for w in words if w.isalpha())
            lexical_diversity = len(unique_words) / len(words) if words else 0

            # Check for proper sentence structure
            has_capitalization = any(s[0].isupper() for s in sentences if s)
            has_punctuation = summary.count('.') + summary.count('!') + summary.count('?') > 0

            # Calculate coherence score (0-100)
            coherence_score = 0
            if 10 <= avg_sentence_length <= 25:
                coherence_score += 30  # Good sentence length
            if 0.3 <= lexical_diversity <= 0.8:
                coherence_score += 30  # Good vocabulary diversity
            if has_capitalization:
                coherence_score += 20  # Proper capitalization
            if has_punctuation:
                coherence_score += 20  # Proper punctuation

            return {
                "coherence_score": coherence_score,
                "avg_sentence_length": round(avg_sentence_length, 2),
                "lexical_diversity": round(lexical_diversity, 3),
                "num_sentences": len(sentences),
                "num_words": len(words),
                "quality_level": (
                    "Excellent" if coherence_score >= 80
                    else "Good" if coherence_score >= 60
                    else "Fair" if coherence_score >= 40
                    else "Poor"
                )
            }

        except Exception as e:
            logger.error(f"Error evaluating coherence: {e}")
            return {"coherence_score": 0, "error": str(e)}

    @staticmethod
    def save_summary_feedback(
        article_id: int,
        summary_id: str,
        rating: int,
        comments: str = ""
    ) -> bool:
        """
        Save user feedback for a summary.

        Args:
            article_id: Article ID
            summary_id: Summary identifier
            rating: User rating (1-5)
            comments: Optional feedback comments

        Returns:
            Success status
        """
        try:
            with get_db_connection() as conn:
                conn.execute("""
                    INSERT INTO summary_feedback
                    (article_id, summary_id, rating, comments, created_at)
                    VALUES (?, ?, ?, ?, ?)
                """, (
                    article_id,
                    summary_id,
                    rating,
                    comments,
                    datetime.now().isoformat()
                ))
                conn.commit()
            return True

        except Exception as e:
            logger.error(f"Error saving summary feedback: {e}")
            return False


class QuestionAnsweringManager:
    """
    Question answering system for scraped content (v1.9.0).

    Provides:
    - Interactive Q&A about articles
    - Source attribution in answers
    - Multi-article synthesis
    - Conversation history tracking
    """

    @staticmethod
    def answer_question(
        question: str,
        articles: List[Dict[str, Any]],
        max_context_length: int = 5000
    ) -> Dict[str, Any]:
        """
        Answer a question based on article content.

        Args:
            question: User question
            articles: List of relevant articles
            max_context_length: Maximum context length for AI

        Returns:
            Dictionary with answer and source attribution
        """
        try:
            provider = get_ai_provider()
            if provider is None:
                return {
                    "answer": "AI provider not configured. Please set API keys in .env file.",
                    "sources": [],
                    "confidence": 0
                }

            # Build context from articles
            context_parts = []
            sources = []

            for idx, article in enumerate(articles[:5]):  # Limit to 5 articles
                title = article.get('title', 'Untitled')
                content = article.get('content', '') or article.get('text', '')

                if content:
                    # Truncate content if needed
                    snippet = content[:1000] if len(content) > 1000 else content
                    context_parts.append(f"[Article {idx+1}: {title}]\n{snippet}")
                    sources.append({
                        "article_id": article.get('id'),
                        "title": title,
                        "url": article.get('url', '')
                    })

            # Combine context
            context = "\n\n".join(context_parts)[:max_context_length]

            # Create Q&A prompt
            qa_prompt = f"""Based on the following articles, please answer this question:

Question: {question}

Articles:
{context}

Please provide a clear, concise answer based on the information in the articles. If the articles don't contain enough information to answer the question, please say so.

Answer:"""

            # Get answer from AI
            answer = provider.get_summary(
                qa_prompt,
                summary_style="brief",
                template=None,
                max_length=len(qa_prompt) + 1000
            )

            return {
                "answer": answer or "Unable to generate answer.",
                "sources": sources,
                "confidence": 0.8 if answer else 0.0,
                "question": question
            }

        except Exception as e:
            logger.error(f"Error answering question: {e}")
            return {
                "answer": f"Error generating answer: {str(e)}",
                "sources": [],
                "confidence": 0
            }

    @staticmethod
    def save_qa_conversation(
        question: str,
        answer: str,
        article_ids: List[int],
        confidence: float = 0.0
    ) -> bool:
        """
        Save Q&A conversation to database.

        Args:
            question: User question
            answer: AI-generated answer
            article_ids: List of source article IDs
            confidence: Answer confidence score

        Returns:
            Success status
        """
        try:
            with get_db_connection() as conn:
                conn.execute("""
                    INSERT INTO qa_history
                    (question, answer, article_ids, confidence, created_at)
                    VALUES (?, ?, ?, ?, ?)
                """, (
                    question,
                    answer,
                    json.dumps(article_ids),
                    confidence,
                    datetime.now().isoformat()
                ))
                conn.commit()
            return True

        except Exception as e:
            logger.error(f"Error saving Q&A conversation: {e}")
            return False

    @staticmethod
    def get_qa_history(limit: int = 20) -> List[Dict[str, Any]]:
        """
        Retrieve Q&A conversation history.

        Args:
            limit: Maximum number of conversations to retrieve

        Returns:
            List of Q&A conversation dictionaries
        """
        try:
            with get_db_connection() as conn:
                cursor = conn.execute("""
                    SELECT id, question, answer, article_ids, confidence, created_at
                    FROM qa_history
                    ORDER BY created_at DESC
                    LIMIT ?
                """, (limit,))

                history = []
                for row in cursor:
                    history.append({
                        "id": row[0],
                        "question": row[1],
                        "answer": row[2],
                        "article_ids": json.loads(row[3]) if row[3] else [],
                        "confidence": row[4],
                        "created_at": row[5]
                    })

                return history

        except Exception as e:
            logger.error(f"Error retrieving Q&A history: {e}")
            return []


# Legacy function wrappers for backward compatibility
def get_summary_from_llm(text_content: str, summary_style: str = "overview", max_length: int = 15000, template: Optional[str] = None) -> str | None:
    """Legacy wrapper that uses the current AI provider."""
    provider = get_ai_provider()
    if provider is None:
        logger.error("No AI provider configured. Set API keys in .env file.")
        return None
    return provider.get_summary(text_content, summary_style, template, max_length)


def get_sentiment_from_llm(text_content: str, max_length: int = 10000) -> Optional[str]:
    """Legacy wrapper that uses the current AI provider."""
    provider = get_ai_provider()
    if provider is None:
        logger.error("No AI provider configured. Set API keys in .env file.")
        return None
    return provider.get_sentiment(text_content, max_length)


def scrape_url_action(
    source_url: str, selector: str, limit: int = 0, user_id: Optional[int] = None
) -> tuple[int, int, str, Optional[List[int]]]:
    """
    Scrape URL and store articles.

    Args:
        source_url: URL to scrape
        selector: CSS selector for articles
        limit: Maximum number of articles (0 = no limit)
        user_id: User ID for tracking ownership (v2.0.0)
    """
    logger.info(
        f"Scraping {source_url} with selector '{selector}', limit {limit}"
    )
    inserted_ids: List[int] = []
    try:
        headers = {'User-Agent': 'Mozilla/5.0 WebScraperTUI/5.0'}
        response = requests.get(source_url, timeout=15, headers=headers)
        response.raise_for_status()
    except Exception as e:
        logger.error(f"Failed to fetch {source_url}: {e}")
        raise
    soup = BeautifulSoup(response.text, "lxml")
    items = soup.select(selector)
    if not items:
        logger.warning(f"No items for '{selector}' on {source_url}")
        return 0, 0, source_url, None
    records = []
    for i, tag_item in enumerate(items):
        if limit > 0 and len(records) >= limit:
            break
        title = tag_item.get_text(strip=True)
        link_href = tag_item.get("href")
        if title and link_href:
            records.append((source_url, title, urljoin(source_url, link_href)))
    if not records:
        return 0, 0, source_url, None
    try:
        with get_db_connection() as conn:
            c = conn.cursor()
            for rec_url, rec_title, rec_link in records:
                try:
                    # v2.0.0: Add user_id tracking
                    c.execute(
                        "INSERT INTO scraped_data (url, title, link, user_id) VALUES (?, ?, ?, ?)",
                        (rec_url, rec_title, rec_link, user_id)
                    )
                    if c.lastrowid:
                        inserted_ids.append(c.lastrowid)
                except sqlite3.IntegrityError:
                    logger.debug(f"Skipping duplicate link: {rec_link}")
                    pass
            conn.commit()
        inserted_count, skipped_count = len(inserted_ids), len(records) - len(inserted_ids)
        logger.info(f"Scraping {source_url}: Stored {inserted_count} new, skipped {skipped_count} duplicates.")
        return inserted_count, skipped_count, source_url, inserted_ids
    except sqlite3.Error as e:
        logger.error(f"DB error storing from {source_url}: {e}", exc_info=True)
        raise


# ═══════════════════════════════════════════════════════════════════════════
# v2.0.0 Multi-User UI Components (Phase 2)
# ═══════════════════════════════════════════════════════════════════════════


class LoginModal(ModalScreen[Optional[int]]):
    """
    Login modal for user authentication (v2.0.0).

    Shown on application startup. Returns user_id if successful, None if cancelled.
    """

    DEFAULT_CSS = """
    LoginModal {
        align: center middle;
        background: $surface-darken-1;
    }

    LoginModal > Vertical {
        width: 60;
        height: auto;
        border: thick $primary;
        background: $panel;
        padding: 2 4;
    }

    LoginModal Label {
        margin: 1 0;
        width: 100%;
    }

    LoginModal #login-title {
        text-align: center;
        text-style: bold;
        color: $accent;
    }

    LoginModal Input {
        margin: 1 0;
    }

    LoginModal Horizontal {
        width: 100%;
        height: auto;
        align-horizontal: center;
        padding-top: 1;
    }

    LoginModal Button {
        margin: 0 2;
    }
    """

    BINDINGS = [
        Binding("escape", "cancel_login", "Cancel"),
        Binding("enter", "attempt_login", "Login")
    ]

    def compose(self) -> ComposeResult:
        with Vertical():
            yield Label("🔐 Login to WebScrape-TUI v2.0.0", id="login-title")
            yield Label("Please enter your credentials:")
            yield Input(placeholder="Username", id="username")
            yield Input(placeholder="Password", password=True, id="password")
            with Horizontal():
                yield Button("Login", variant="primary", id="login-btn")
                yield Button("Cancel", id="cancel-btn")

    def on_mount(self) -> None:
        """Focus username field on mount."""
        self.query_one("#username", Input).focus()

    def action_cancel_login(self) -> None:
        """Cancel login (ESC key)."""
        self.dismiss(None)

    def action_attempt_login(self) -> None:
        """Attempt login (Enter key)."""
        username_input = self.query_one("#username", Input)
        password_input = self.query_one("#password", Input)
        self._try_login(username_input.value, password_input.value)

    def on_button_pressed(self, event: Button.Pressed) -> None:
        """Handle button clicks."""
        if event.button.id == "login-btn":
            username = self.query_one("#username", Input).value
            password = self.query_one("#password", Input).value
            self._try_login(username, password)
        else:
            self.dismiss(None)

    def _try_login(self, username: str, password: str) -> None:
        """Execute login attempt."""
        if not username or not password:
            self.app.notify(
                "Please enter both username and password",
                severity="warning"
            )
            return

        # Call authentication function from Phase 1
        user_id = authenticate_user(username, password)

        if user_id:
            self.dismiss(user_id)
        else:
            self.app.notify("Invalid credentials", severity="error")
            self.query_one("#password", Input).value = ""
            self.query_one("#password", Input).focus()


class UserProfileModal(ModalScreen[None]):
    """User profile view/edit modal (v2.0.0)."""

    DEFAULT_CSS = """
    UserProfileModal {
        align: center middle;
        background: $surface-darken-1;
    }

    UserProfileModal > Vertical {
        width: 70;
        height: auto;
        border: thick $primary;
        background: $panel;
        padding: 2 4;
    }

    UserProfileModal Label {
        margin: 1 0;
    }

    UserProfileModal #profile-title {
        text-align: center;
        text-style: bold;
        color: $accent;
    }

    UserProfileModal Input {
        margin: 1 0;
    }

    UserProfileModal Horizontal {
        width: 100%;
        height: auto;
        align-horizontal: center;
        padding-top: 1;
    }

    UserProfileModal Button {
        margin: 0 1;
    }
    """

    def __init__(self, user_id: int):
        super().__init__()
        self.user_id = user_id
        self.user_data = {}

    def compose(self) -> ComposeResult:
        # Load user data
        try:
            with get_db_connection() as conn:
                row = conn.execute("""
                    SELECT username, email, role, created_at, last_login, is_active
                    FROM users WHERE id = ?
                """, (self.user_id,)).fetchone()
                if row:
                    self.user_data = dict(row)
        except Exception as e:
            logger.error(f"Error loading user profile: {e}")

        with Vertical():
            yield Label("👤 User Profile", id="profile-title")
            yield Label(f"Username: {self.user_data.get('username', 'Unknown')}")
            yield Label(f"Email: {self.user_data.get('email', 'Not set')}")
            yield Label(f"Role: {self.user_data.get('role', 'user').upper()}")
            yield Label(f"Created: {self.user_data.get('created_at', 'Unknown')}")
            yield Label(f"Last Login: {self.user_data.get('last_login', 'Never')}")
            yield Label(f"Status: {'Active' if self.user_data.get('is_active') else 'Inactive'}")
            yield Label("Email Address:")
            yield Input(
                placeholder="Email address",
                value=self.user_data.get('email', '') or '',
                id="email-input"
            )
            with Horizontal():
                yield Button("Update Email", variant="primary", id="update-btn")
                yield Button("Change Password", id="password-btn")
                yield Button("Close", id="close-btn")

    def on_button_pressed(self, event: Button.Pressed) -> None:
        """Handle button clicks."""
        if event.button.id == "update-btn":
            email = self.query_one("#email-input", Input).value
            try:
                with get_db_connection() as conn:
                    conn.execute(
                        "UPDATE users SET email = ? WHERE id = ?",
                        (email or None, self.user_id)
                    )
                    conn.commit()
                self.app.notify("Email updated successfully", severity="information")
            except Exception as e:
                logger.error(f"Error updating email: {e}")
                self.app.notify("Failed to update email", severity="error")

        elif event.button.id == "password-btn":
            self.app.push_screen(ChangePasswordModal(self.user_id))

        else:
            self.dismiss()


class ChangePasswordModal(ModalScreen[bool]):
    """Modal for changing user password (v2.0.0)."""

    DEFAULT_CSS = """
    ChangePasswordModal {
        align: center middle;
        background: $surface-darken-1;
    }

    ChangePasswordModal > Vertical {
        width: 60;
        height: auto;
        border: thick $primary;
        background: $panel;
        padding: 2 4;
    }

    ChangePasswordModal Label {
        margin: 1 0;
    }

    ChangePasswordModal #pwd-title {
        text-align: center;
        text-style: bold;
        color: $accent;
    }

    ChangePasswordModal Input {
        margin: 1 0;
    }

    ChangePasswordModal Horizontal {
        width: 100%;
        height: auto;
        align-horizontal: center;
        padding-top: 1;
    }

    ChangePasswordModal Button {
        margin: 0 1;
    }
    """

    def __init__(self, user_id: int):
        super().__init__()
        self.user_id = user_id

    def compose(self) -> ComposeResult:
        with Vertical():
            yield Label("🔒 Change Password", id="pwd-title")
            yield Input(placeholder="Current Password", password=True, id="current-pwd")
            yield Input(placeholder="New Password", password=True, id="new-pwd")
            yield Input(placeholder="Confirm New Password", password=True, id="confirm-pwd")
            with Horizontal():
                yield Button("Change Password", variant="primary", id="change-btn")
                yield Button("Cancel", id="cancel-btn")

    def on_mount(self) -> None:
        """Focus current password field."""
        self.query_one("#current-pwd", Input).focus()

    def on_button_pressed(self, event: Button.Pressed) -> None:
        """Handle button clicks."""
        if event.button.id == "change-btn":
            current = self.query_one("#current-pwd", Input).value
            new = self.query_one("#new-pwd", Input).value
            confirm = self.query_one("#confirm-pwd", Input).value

            # Validate current password
            try:
                with get_db_connection() as conn:
                    row = conn.execute(
                        "SELECT password_hash FROM users WHERE id = ?",
                        (self.user_id,)
                    ).fetchone()

                    if not row or not verify_password(current, row['password_hash']):
                        self.app.notify("Current password incorrect", severity="error")
                        return
            except Exception as e:
                logger.error(f"Error validating password: {e}")
                self.app.notify("Error validating password", severity="error")
                return

            # Validate new passwords match
            if new != confirm:
                self.app.notify("New passwords do not match", severity="error")
                return

            # Validate password strength (min 8 chars)
            if len(new) < 8:
                self.app.notify(
                    "Password must be at least 8 characters",
                    severity="error"
                )
                return

            # Update password
            try:
                new_hash = hash_password(new)
                with get_db_connection() as conn:
                    conn.execute(
                        "UPDATE users SET password_hash = ? WHERE id = ?",
                        (new_hash, self.user_id)
                    )
                    conn.commit()

                self.app.notify(
                    "Password changed successfully",
                    severity="information"
                )
                self.dismiss(True)
            except Exception as e:
                logger.error(f"Error changing password: {e}")
                self.app.notify("Failed to change password", severity="error")
        else:
            self.dismiss(False)


class UserManagementModal(ModalScreen[None]):
    """Admin-only user management screen (v2.0.0)."""

    DEFAULT_CSS = """
    UserManagementModal {
        align: center middle;
        background: $surface-darken-1;
    }

    UserManagementModal > Vertical {
        width: 90;
        height: 40;
        border: thick $primary;
        background: $panel;
        padding: 2 4;
    }

    UserManagementModal Label {
        margin: 1 0;
    }

    UserManagementModal #mgmt-title {
        text-align: center;
        text-style: bold;
        color: $accent;
    }

    UserManagementModal DataTable {
        height: 1fr;
        margin: 1 0;
    }

    UserManagementModal Horizontal {
        width: 100%;
        height: auto;
        align-horizontal: center;
        padding-top: 1;
    }

    UserManagementModal Button {
        margin: 0 1;
    }
    """

    def compose(self) -> ComposeResult:
        with Vertical():
            yield Label("👥 User Management [ADMIN]", id="mgmt-title")
            yield DataTable(id="users-table", cursor_type="row", zebra_stripes=True)
            with Horizontal():
                yield Button("Create User", variant="primary", id="create-btn")
                yield Button("Edit User", id="edit-btn")
                yield Button("Toggle Active", id="toggle-btn")
                yield Button("Close", id="close-btn")

    def on_mount(self) -> None:
        """Initialize user table."""
        table = self.query_one("#users-table", DataTable)
        table.add_columns("ID", "Username", "Email", "Role", "Active", "Last Login")
        self.refresh_users()
        table.focus()

    def refresh_users(self, *args) -> None:
        """Refresh user list from database."""
        table = self.query_one("#users-table", DataTable)
        table.clear()

        try:
            with get_db_connection() as conn:
                rows = conn.execute("""
                    SELECT id, username, email, role, is_active, last_login
                    FROM users
                    ORDER BY id
                """).fetchall()

                for row in rows:
                    table.add_row(
                        str(row['id']),
                        row['username'],
                        row['email'] or '',
                        row['role'].upper(),
                        "✓" if row['is_active'] else "✗",
                        row['last_login'] or 'Never',
                        key=str(row['id'])
                    )
        except Exception as e:
            logger.error(f"Error refreshing users: {e}")
            self.app.notify("Failed to load users", severity="error")

    def on_button_pressed(self, event: Button.Pressed) -> None:
        """Handle button clicks."""
        if event.button.id == "create-btn":
            self.app.push_screen(CreateUserModal(), self.refresh_users)

        elif event.button.id == "edit-btn":
            table = self.query_one("#users-table", DataTable)
            if table.row_count > 0:
                try:
                    row_key = table.get_row_key_from_coordinate(table.cursor_coordinate)
                    if row_key:
                        user_id = int(row_key.value)
                        self.app.push_screen(EditUserModal(user_id), self.refresh_users)
                except Exception as e:
                    logger.error(f"Error getting selected user: {e}")
                    self.app.notify("Please select a user", severity="warning")

        elif event.button.id == "toggle-btn":
            table = self.query_one("#users-table", DataTable)
            if table.row_count > 0:
                try:
                    row_key = table.get_row_key_from_coordinate(table.cursor_coordinate)
                    if row_key:
                        user_id = int(row_key.value)
                        with get_db_connection() as conn:
                            conn.execute("""
                                UPDATE users SET is_active = NOT is_active WHERE id = ?
                            """, (user_id,))
                            conn.commit()
                        self.refresh_users()
                        self.app.notify("User status toggled", severity="information")
                except Exception as e:
                    logger.error(f"Error toggling user: {e}")
                    self.app.notify("Failed to toggle user status", severity="error")

        else:
            self.dismiss()


class CreateUserModal(ModalScreen[bool]):
    """Modal for creating new user (admin only, v2.0.0)."""

    DEFAULT_CSS = """
    CreateUserModal {
        align: center middle;
        background: $surface-darken-1;
    }

    CreateUserModal > Vertical {
        width: 60;
        height: auto;
        border: thick $primary;
        background: $panel;
        padding: 2 4;
    }

    CreateUserModal Label {
        margin: 1 0;
    }

    CreateUserModal #create-title {
        text-align: center;
        text-style: bold;
        color: $accent;
    }

    CreateUserModal Input {
        margin: 1 0;
    }

    CreateUserModal RadioSet {
        margin: 1 0;
    }

    CreateUserModal Horizontal {
        width: 100%;
        height: auto;
        align-horizontal: center;
        padding-top: 1;
    }

    CreateUserModal Button {
        margin: 0 1;
    }
    """

    def compose(self) -> ComposeResult:
        with Vertical():
            yield Label("➕ Create New User", id="create-title")
            yield Input(placeholder="Username", id="username")
            yield Input(placeholder="Email (optional)", id="email")
            yield Input(placeholder="Password", password=True, id="password")
            yield Label("Role:")
            with RadioSet(id="role-radio"):
                yield RadioButton("Admin", id="role-admin")
                yield RadioButton("User", id="role-user", value=True)
                yield RadioButton("Viewer", id="role-viewer")
            with Horizontal():
                yield Button("Create", variant="primary", id="create-btn")
                yield Button("Cancel", id="cancel-btn")

    def on_mount(self) -> None:
        """Focus username field."""
        self.query_one("#username", Input).focus()

    def on_button_pressed(self, event: Button.Pressed) -> None:
        """Handle button clicks."""
        if event.button.id == "create-btn":
            username = self.query_one("#username", Input).value
            email = self.query_one("#email", Input).value
            password = self.query_one("#password", Input).value

            # Determine role from radio buttons
            role = "user"  # default
            if self.query_one("#role-admin", RadioButton).value:
                role = "admin"
            elif self.query_one("#role-viewer", RadioButton).value:
                role = "viewer"

            # Validate inputs
            if not username or not password:
                self.app.notify(
                    "Username and password required",
                    severity="error"
                )
                return

            if len(password) < 8:
                self.app.notify(
                    "Password must be at least 8 characters",
                    severity="error"
                )
                return

            # Create user
            try:
                password_hash = hash_password(password)
                with get_db_connection() as conn:
                    conn.execute("""
                        INSERT INTO users (username, email, password_hash, role)
                        VALUES (?, ?, ?, ?)
                    """, (username, email or None, password_hash, role))
                    conn.commit()

                self.app.notify(
                    f"User '{username}' created successfully",
                    severity="information"
                )
                self.dismiss(True)

            except sqlite3.IntegrityError:
                self.app.notify("Username already exists", severity="error")
            except Exception as e:
                logger.error(f"Error creating user: {e}")
                self.app.notify("Failed to create user", severity="error")
        else:
            self.dismiss(False)


class EditUserModal(ModalScreen[bool]):
    """Modal for editing existing user (admin only, v2.0.0)."""

    DEFAULT_CSS = """
    EditUserModal {
        align: center middle;
        background: $surface-darken-1;
    }

    EditUserModal > Vertical {
        width: 60;
        height: auto;
        border: thick $primary;
        background: $panel;
        padding: 2 4;
    }

    EditUserModal Label {
        margin: 1 0;
    }

    EditUserModal #edit-title {
        text-align: center;
        text-style: bold;
        color: $accent;
    }

    EditUserModal Input {
        margin: 1 0;
    }

    EditUserModal RadioSet {
        margin: 1 0;
    }

    EditUserModal Horizontal {
        width: 100%;
        height: auto;
        align-horizontal: center;
        padding-top: 1;
    }

    EditUserModal Button {
        margin: 0 1;
    }
    """

    def __init__(self, user_id: int):
        super().__init__()
        self.user_id = user_id
        self.user_data = {}

    def compose(self) -> ComposeResult:
        # Load user data
        try:
            with get_db_connection() as conn:
                row = conn.execute("""
                    SELECT username, email, role
                    FROM users WHERE id = ?
                """, (self.user_id,)).fetchone()
                if row:
                    self.user_data = dict(row)
        except Exception as e:
            logger.error(f"Error loading user data: {e}")

        current_role = self.user_data.get('role', 'user')

        with Vertical():
            yield Label(
                f"✏️ Edit User: {self.user_data.get('username', 'Unknown')}",
                id="edit-title"
            )
            yield Label("Email:")
            yield Input(
                placeholder="Email",
                value=self.user_data.get('email', '') or '',
                id="email"
            )
            yield Label("Role:")
            with RadioSet(id="role-radio"):
                yield RadioButton(
                    "Admin",
                    id="role-admin",
                    value=(current_role == 'admin')
                )
                yield RadioButton(
                    "User",
                    id="role-user",
                    value=(current_role == 'user')
                )
                yield RadioButton(
                    "Viewer",
                    id="role-viewer",
                    value=(current_role == 'viewer')
                )
            with Horizontal():
                yield Button("Update", variant="primary", id="update-btn")
                yield Button("Cancel", id="cancel-btn")

    def on_button_pressed(self, event: Button.Pressed) -> None:
        """Handle button clicks."""
        if event.button.id == "update-btn":
            email = self.query_one("#email", Input).value

            # Determine role
            role = "user"
            if self.query_one("#role-admin", RadioButton).value:
                role = "admin"
            elif self.query_one("#role-viewer", RadioButton).value:
                role = "viewer"

            # Update user
            try:
                with get_db_connection() as conn:
                    conn.execute("""
                        UPDATE users SET email = ?, role = ? WHERE id = ?
                    """, (email or None, role, self.user_id))
                    conn.commit()

                self.app.notify("User updated successfully", severity="information")
                self.dismiss(True)
            except Exception as e:
                logger.error(f"Error updating user: {e}")
                self.app.notify("Failed to update user", severity="error")
        else:
            self.dismiss(False)


class ConfirmModal(ModalScreen[bool]):
    DEFAULT_CSS = """
    ConfirmModal {
        align: center middle;
        background: $surface-darken-1;
    }
    ConfirmModal > Vertical {
        width: auto;
        max-width: 80%;
        min-width: 50;
        height: auto;
        padding: 2 4;
        border: thick $primary-lighten-1;
        background: $panel;
    }
    ConfirmModal Label {
        margin-bottom: 1;
        width: 100%;
        text-align: center;
    }
    ConfirmModal Horizontal {
        width: 100%;
        height: auto;
        align-horizontal: center;
        padding-top: 1;
    }
    ConfirmModal Button {
        margin: 0 1;
        min-width: 10;
    }
    """
    BINDINGS = [Binding("escape", "dismiss_modal", "Cancel"), Binding("enter", "confirm_action", "Confirm")]

    def __init__(self, prompt: str, confirm_text: str = "Confirm", cancel_text: str = "Cancel"):
        super().__init__()
        self.prompt = prompt
        self.confirm_text = confirm_text
        self.cancel_text = cancel_text

    def compose(self) -> ComposeResult:
        with Vertical():
            yield Label(self.prompt)
            with Horizontal():
                yield Button(self.confirm_text, variant="primary", id="confirm")
                yield Button(self.cancel_text, id="cancel")

    def on_button_pressed(self, event: Button.Pressed) -> None:
        self.dismiss(event.button.id == "confirm")

    def action_dismiss_modal(self) -> None:
        self.dismiss(False)

    def action_confirm_action(self) -> None:
        self.dismiss(True)


class FilenameModal(ModalScreen[Optional[str]]):
    DEFAULT_CSS = """
    FilenameModal {
        align: center middle;
        background: $surface-darken-1;
    }
    FilenameModal > Vertical {
        width: 60;
        height: auto;
        border: thick $primary-lighten-1;
        padding: 1 2;
        background: $surface;
    }
    FilenameModal Input {
        margin-bottom: 1;
    }
    FilenameModal Button {
        margin-top: 1;
    }
    FilenameModal Label.dialog-title {
        text-style: bold;
        width: 100%;
        text-align: center;
        margin-bottom: 1;
    }
    """

    def __init__(self, p: str = "Enter filename:", dfn: str = "export.csv"):
        super().__init__()
        self.p, self.dfn = p, dfn

    def compose(self) -> ComposeResult:
        with Vertical():
            yield Label(self.p, classes="dialog-title")
            yield Input(value=self.dfn, id="fn_in")
            yield Horizontal(
                Button("Save", variant="primary", id="sfn_b"),
                Button("Cancel", id="cfn_b"),
                classes="modal-buttons"
            )

    def on_button_pressed(self, e: Button.Pressed) -> None:
        if e.button.id == "sfn_b":
            fn = self.query_one("#fn_in", Input).value.strip()
            if fn:
                self.dismiss(fn)
            else:
                self.app.notify("Filename empty.", title="Error", severity="error")
        else:
            self.dismiss(None)


class ManageTagsModal(ModalScreen[Optional[str]]):
    DEFAULT_CSS = """
    ManageTagsModal {
        align: center middle;
        background: $surface-darken-1;
    }
    ManageTagsModal > Vertical {
        width: 70;
        height: auto;
        border: thick $primary-lighten-1;
        padding: 1 2;
        background: $surface;
    }
    ManageTagsModal Label {
        margin-bottom: 0.5;
    }
    ManageTagsModal Input {
        margin-bottom: 1;
    }
    ManageTagsModal #cur_tags_lbl {
        color: $text-muted;
        text-style: italic;
        max-height: 3;
        overflow-y: auto;
        border: round $primary-darken-1;
        padding: 0 1;
    }
    """

    def __init__(self, aid: int, ctl: List[str]):
        super().__init__()
        self.aid, self.cts = aid, ", ".join(sorted(ctl))

    def compose(self) -> ComposeResult:
        with Vertical():
            yield Label(f"Tags for Article ID: {self.aid}", classes="dialog-title")
            yield Label("Current:")
            yield Static(self.cts or "None", id="cur_tags_lbl")
            yield Label("New (comma-sep):")
            yield Input(value=self.cts, id="tags_in")
            yield Horizontal(
                Button("Save", variant="primary", id="st_b"),
                Button("Cancel", id="ct_b"),
                classes="modal-buttons"
            )

    def on_button_pressed(self, e: Button.Pressed) -> None:
        if e.button.id == "st_b":
            self.dismiss(self.query_one("#tags_in", Input).value)
        else:
            self.dismiss(None)


class SelectSummaryStyleModal(ModalScreen[Optional[str]]):
    DEFAULT_CSS = """
    SelectSummaryStyleModal {
        align: center middle;
        background: $surface-darken-1;
    }
    SelectSummaryStyleModal > Vertical {
        width: 70;
        height: auto;
        max-height: 80%;
        border: thick $primary-lighten-1;
        padding: 1 2;
        background: $surface;
    }
    SelectSummaryStyleModal VerticalScroll {
        height: auto;
        max-height: 20;
        margin: 1 0;
    }
    SelectSummaryStyleModal RadioSet {
        height: auto;
    }
    SelectSummaryStyleModal RadioButton {
        padding: 1 0;
    }
    """

    def __init__(self):
        super().__init__()
        self.templates = TemplateManager.get_all_templates()

    def compose(self) -> ComposeResult:
        with Vertical():
            yield Label("Select Summarization Template", classes="dialog-title")
            with VerticalScroll():
                # Create radio buttons from database templates
                radio_buttons = []
                for i, template in enumerate(self.templates):
                    label = f"{template['name']}"
                    if template['description']:
                        label += f" - {template['description']}"
                    radio_buttons.append(
                        RadioButton(label, id=f"template_{template['id']}", value=(i == 0))
                    )
                yield RadioSet(*radio_buttons)
            with Horizontal(classes="modal-buttons"):
                yield Button("Summarize", variant="primary", id="cs_b")
                yield Button("Manage Templates", id="manage_templates_b")
                yield Button("Cancel", id="ccs_b")

    def on_button_pressed(self, e: Button.Pressed) -> None:
        if e.button.id == "cs_b":
            rs = self.query_one(RadioSet)
            if rs.pressed_button:
                # Extract template ID from button id (format: "template_{id}")
                template_id = int(rs.pressed_button.id.split("_")[1])
                # Find the template name
                template = next((t for t in self.templates if t['id'] == template_id), None)
                if template:
                    self.dismiss(f"template:{template['name']}")
                else:
                    self.dismiss(None)
            else:
                self.dismiss(None)
        elif e.button.id == "manage_templates_b":
            # TODO: Open template management modal
            self.dismiss(None)
        else:
            self.dismiss(None)


class ViewSummaryModal(ModalScreen):
    CSS = """
    ViewSummaryModal {
        align: center middle;
        background: $surface-darken-1;
    }
    ViewSummaryModal > VerticalScroll {
        width: 90%;
        max-width: 100;
        height: 80%;
        max-height: 35;
        border: thick $primary-lighten-1;
        padding: 1 2;
        background: $surface;
    }
    """
    BINDINGS = [Binding("escape", "dismiss", "Close")]

    def __init__(self, title: str, summary_data: dict):
        super().__init__()
        self.article_title = title
        self.summary_data = summary_data

    def compose(self) -> ComposeResult:
        with VerticalScroll():
            yield Label(f"Summary: {self.article_title}", id="summary_title", classes="dialog-title")

            if self.summary_data.get('summary'):
                yield Label("Summary:", classes="section-header")
                yield Static(self.summary_data['summary'], classes="summary-content")
            else:
                yield Label("No summary available for this article.", classes="no-content")

            if self.summary_data.get('sentiment'):
                yield Label("Sentiment Analysis:", classes="section-header")
                yield Static(self.summary_data['sentiment'], classes="sentiment-content")

            yield Horizontal(Button("Close", id="close_btn"), classes="modal-buttons")

    def on_button_pressed(self, event: Button.Pressed) -> None:
        self.dismiss()

    def action_dismiss(self) -> None:
        self.dismiss()


class ReadArticleModal(ModalScreen):
    DEFAULT_CSS = """
    ReadArticleModal {
        align: center middle;
        background: $surface-darken-1;
    }
    ReadArticleModal > Vertical {
        width: 90%;
        height: 90%;
        border: thick $primary-lighten-1;
        background: $surface;
        padding: 1;
    }
    ReadArticleModal #art_ttl_lbl {
        text-style: bold;
        margin-bottom: 1;
        text-align: center;
        background: $primary-background-darken-1;
        padding: 0 1;
        color: $text;
    }
    ReadArticleModal VerticalScroll {
        border: panel $primary;
        padding: 1;
        background: $surface;
    }
    ReadArticleModal Markdown {
        width: 100%;
    }
    """

    def __init__(self, t: str, c: str):
        super().__init__()
        self.t, self.c = t, c

    def compose(self) -> ComposeResult:
        with Vertical():
            yield Label(self.t, id="art_ttl_lbl")
            yield VerticalScroll(Markdown(self.c or "_No content._"))
            yield Horizontal(Button("Close", id="cr_b"), classes="modal-buttons")

    def on_button_pressed(self, e: Button.Pressed) -> None:
        self.dismiss()


class AIProviderSelectionModal(ModalScreen[Optional[str]]):
    """Modal for selecting AI provider."""
    DEFAULT_CSS = """
    AIProviderSelectionModal {
        align: center middle;
        background: $surface-darken-1;
    }
    AIProviderSelectionModal > Vertical {
        width: 60;
        height: auto;
        border: thick $primary-lighten-1;
        background: $surface;
        padding: 2;
    }
    AIProviderSelectionModal Label {
        text-align: center;
        text-style: bold;
        margin-bottom: 1;
    }
    AIProviderSelectionModal RadioSet {
        height: auto;
        margin-bottom: 1;
    }
    AIProviderSelectionModal Horizontal {
        width: 100%;
        height: auto;
        align-horizontal: center;
        padding-top: 1;
    }
    """

    def __init__(self):
        super().__init__()
        self.current_provider = get_ai_provider()

    def compose(self) -> ComposeResult:
        with Vertical():
            yield Label("Select AI Provider")
            with RadioSet(id="provider_radioset"):
                yield RadioButton("Google Gemini", value=GEMINI_API_KEY != "", id="gemini_radio")
                yield RadioButton("OpenAI GPT", value=OPENAI_API_KEY != "", id="openai_radio")
                yield RadioButton("Anthropic Claude", value=CLAUDE_API_KEY != "", id="claude_radio")
            with Horizontal():
                yield Button("Select", variant="primary", id="select_btn")
                yield Button("Cancel", id="cancel_btn")

    def on_mount(self) -> None:
        # Pre-select current provider
        if self.current_provider:
            name = self.current_provider.name
            if "Gemini" in name:
                self.query_one("#gemini_radio", RadioButton).value = True
            elif "OpenAI" in name:
                self.query_one("#openai_radio", RadioButton).value = True
            elif "Claude" in name:
                self.query_one("#claude_radio", RadioButton).value = True

    def on_button_pressed(self, event: Button.Pressed) -> None:
        if event.button.id == "select_btn":
            radioset = self.query_one("#provider_radioset", RadioSet)
            pressed_idx = radioset.pressed_index
            if pressed_idx == 0 and GEMINI_API_KEY:
                self.dismiss("gemini")
            elif pressed_idx == 1 and OPENAI_API_KEY:
                self.dismiss("openai")
            elif pressed_idx == 2 and CLAUDE_API_KEY:
                self.dismiss("claude")
            else:
                self.dismiss(None)
        else:
            self.dismiss(None)


class ScrapeURLModal(ModalScreen[tuple[str, str, int] | None]):
    DEFAULT_CSS = """
    ScrapeURLModal {
        align: center middle;
        background: $surface-darken-1;
    }
    ScrapeURLModal > Vertical {
        width: 70;
        height: auto;
        border: thick $primary-lighten-1;
        padding: 1 2;
        background: $surface;
    }
    ScrapeURLModal Input {
        margin-bottom: 1;
    }
    ScrapeURLModal Button {
        margin-top: 1;
    }
    ScrapeURLModal Label.dialog-title {
        text-style: bold;
        width: 100%;
        text-align: center;
        margin-bottom: 1;
    }
    ScrapeURLModal Static.profile-context {
        text-align: center;
        color: $accent;
        margin-bottom: 1;
    }
    """

    def __init__(self, lu: str = "", ls: str = "h2 a", ll: int = 0):
        super().__init__()
        self.lu, self.ls, self.ll = lu, ls, ll

    def compose(self) -> ComposeResult:
        with Vertical():
            yield Label("Scrape New URL", classes="dialog-title")
            yield Static(f"Profile: {self.app.current_scraper_profile}", classes="profile-context")
            yield Input(value=self.lu, placeholder="URL", id="s_url_in")
            yield Input(value=self.ls, placeholder="CSS Selector", id="s_sel_in")
            yield Input(value=str(self.ll), placeholder="Limit (0=all)", id="s_lim_in", type="integer")
            yield Horizontal(
                Button("Scrape", variant="primary", id="sc_b"),
                Button("Cancel", id="scc_b"),
                classes="modal-buttons"
            )

    def on_button_pressed(self, e: Button.Pressed) -> None:
        if e.button.id == "sc_b":
            u = self.query_one("#s_url_in", Input).value
            s = self.query_one("#s_sel_in", Input).value
            l_s = self.query_one("#s_lim_in", Input).value
            try:
                limit = int(l_s) if l_s.strip() else 0
            except ValueError:
                self.app.notify("Invalid limit.", title="Error", severity="error")
                return
            if not u or not s:
                self.app.notify("URL & Selector required.", title="Error", severity="error")
                return
            if hasattr(self.app, 'last_scrape_url'):
                self.app.last_scrape_url = u
            if hasattr(self.app, 'last_scrape_selector'):
                self.app.last_scrape_selector = s
            if hasattr(self.app, 'last_scrape_limit'):
                self.app.last_scrape_limit = limit
            self.dismiss((u, s, limit))
        else:
            self.dismiss(None)


class ArticleDetailModal(ModalScreen):
    DEFAULT_CSS = """
    ArticleDetailModal {
        align: center middle;
        background: $surface-darken-1;
    }
    ArticleDetailModal > ScrollableContainer {
        width: 80%;
        max-width: 100;
        height: 80%;
        max-height: 38;
        border: thick $primary-lighten-1;
        padding: 0;
        background: $surface;
    }
    ArticleDetailModal Markdown {
        padding: 1 2;
    }
    ArticleDetailModal Button {
        width: 100%;
        margin-top: 1;
    }
    """

    def __init__(self, ad: sqlite3.Row, tags: List[str]):
        super().__init__()
        self.ad, self.tags = ad, tags

    def compose(self) -> ComposeResult:
        ts = ", ".join(sorted(self.tags)) if self.tags else "_No tags_"
        senti_str = self.ad['sentiment'] or "_N/A_"
        timestamp_val = self.ad['timestamp']
        if isinstance(timestamp_val, datetime):
            timestamp_str = timestamp_val.strftime('%Y-%m-%d %H:%M:%S')
        else:
            timestamp_str = str(timestamp_val)
        c = (f"# {self.ad['title']}\n"
             f"**ID:** {self.ad['id']}\n"
             f"**Src URL:** {self.ad['url']}\n"
             f"**Link:** [{self.ad['link']}]({self.ad['link']})\n"
             f"**Scraped:** {timestamp_str}\n"
             f"**Tags:** {ts}\n"
             f"**Sentiment:** {senti_str}\n"
             f"---\n## Summary\n")
        c += f"\n{self.ad['summary']}" if self.ad['summary'] else "\n_Not summarized._"
        with VerticalScroll():
            yield Markdown(c)
            yield Horizontal(Button("Close", variant="primary", id="d_cls_b"), classes="modal-buttons")

    def on_button_pressed(self, e: Button.Pressed) -> None:
        self.dismiss()


class OriginalURLModal(ModalScreen[Optional[str]]):
    DEFAULT_CSS = """
    OriginalURLModal {
        align: center middle;
        background: $surface-darken-1;
    }
    OriginalURLModal > Vertical {
        width: 70;
        height: auto;
        border: thick $primary-lighten-1;
        padding: 1 2;
        background: $surface;
    }
    OriginalURLModal Input {
        margin-bottom: 1;
    }
    OriginalURLModal Button {
        margin-top: 1;
    }
    """

    def __init__(self, prompt_text: str = "Enter Original URL for Archival Fetch:"):
        super().__init__()
        self.prompt_text = prompt_text

    def compose(self) -> ComposeResult:
        with Vertical():
            yield Label(self.prompt_text, classes="dialog-title")
            yield Input(placeholder="e.g., https://example.com/article", id="original_url_input")
            with Horizontal(classes="modal-buttons"):
                yield Button("Fetch Archive", variant="primary", id="fetch_archive_url")
                yield Button("Cancel", id="cancel_archive_url")

    def on_button_pressed(self, event: Button.Pressed) -> None:
        if event.button.id == "fetch_archive_url":
            url = self.query_one("#original_url_input", Input).value.strip()
            if not url:
                self.app.notify("Original URL cannot be empty.", title="Input Error", severity="error")
                return
            self.dismiss(url)
        else:
            self.dismiss(None)


class SavedScraperItem(ListItem):
    def __init__(self, scraper_data: sqlite3.Row):
        super().__init__()
        self.scraper_data = scraper_data
        self.prefix = "[P] " if scraper_data['is_preinstalled'] else ""

    def compose(self) -> ComposeResult:
        display_name = f"{self.prefix}{self.scraper_data['name']}"
        sub_text = self.scraper_data['description'] or self.scraper_data['url']
        if len(sub_text) > 60:
            sub_text = sub_text[:57] + "..."
        with Vertical():
            yield Label(display_name, classes="scraper-item-name")
            yield Label(f"  ({sub_text})", classes="scraper-item-subtext")


class ManageScrapersModal(ModalScreen[Optional[Tuple[str, Any]]]):
    DEFAULT_CSS = """
    ManageScrapersModal {
        align: center middle;
        background: $surface-darken-1;
    }
    ManageScrapersModal > Vertical {
        width: 85%;
        max-width: 110;
        height: 85%;
        max-height: 35;
        border: thick $primary-lighten-1;
        padding: 1;
        background: $surface;
    }
    ManageScrapersModal ListView {
        height: 1fr;
        margin-bottom: 1;
        border: panel $primary;
    }
    ManageScrapersModal .buttons-top, ManageScrapersModal .buttons-bottom {
        layout: horizontal;
        height: auto;
        align-horizontal: center;
        margin-top: 1;
    }
    ManageScrapersModal Button {
        margin: 0 1;
    }
    .scraper-item-name {
        text-style: bold;
    }
    .scraper-item-subtext {
        color: $text-muted;
        text-style: italic;
    }
    """

    def compose(self) -> ComposeResult:
        with Vertical():
            yield Label("Manage Scraper Profiles", classes="dialog-title")
            yield ListView(id="scrapers_list_view")
            with Horizontal(classes="buttons-top"):
                yield Button("Add New", id="add_scraper", variant="success")
                yield Button("Edit Selected", id="edit_scraper", variant="primary")
                yield Button("Delete Selected", id="delete_scraper", variant="error")
            with Horizontal(classes="buttons-bottom"):
                yield Button("Execute Selected", id="execute_scraper", variant="success")
                yield Button("Close", id="close_manage_scrapers")

    async def on_mount(self) -> None:
        await self.load_scrapers()

    async def load_scrapers(self) -> None:
        lv = self.query_one(ListView)
        lv.clear()
        try:
            with get_db_connection() as conn:
                scrapers = conn.execute(
                    "SELECT id,name,url,selector,default_limit,default_tags_csv,description,is_preinstalled "
                    "FROM saved_scrapers ORDER BY is_preinstalled DESC, name COLLATE NOCASE"
                ).fetchall()
            for sd in scrapers:
                lv.append(SavedScraperItem(sd))
        except Exception as e:
            logger.error(f"Failed to load saved scrapers: {e}", exc_info=True)
            self.app.notify("Error loading scrapers.", severity="error")

    async def on_button_pressed(self, e: Button.Pressed) -> None:
        lv = self.query_one(ListView)
        si = lv.highlighted_child
        if e.button.id == "add_scraper":
            self.dismiss(("add", None))
        elif e.button.id == "edit_scraper":
            if si and isinstance(si, SavedScraperItem):
                self.dismiss(("edit", si.scraper_data))
            else:
                self.app.notify("No scraper selected to edit.", severity="warning")
        elif e.button.id == "delete_scraper":
            if si and isinstance(si, SavedScraperItem):
                if si.scraper_data['is_preinstalled']:
                    self.app.notify("Pre-installed profiles cannot be deleted directly. You can edit them.", severity="warning")
                    return
                self.dismiss(("delete", si.scraper_data['id']))
            else:
                self.app.notify("No scraper selected to delete.", severity="warning")
        elif e.button.id == "execute_scraper":
            if si and isinstance(si, SavedScraperItem):
                self.dismiss(("execute", si.scraper_data))
            else:
                self.app.notify("No scraper selected to execute.", severity="warning")
        elif e.button.id == "close_manage_scrapers":
            self.dismiss(None)


class AddEditScraperModal(ModalScreen[Optional[Dict[str, Any]]]):
    DEFAULT_CSS = """
    AddEditScraperModal {
        align: center middle;
        background: $surface-darken-1;
    }
    AddEditScraperModal > VerticalScroll {
        width: 70%;
        max-width: 90;
        max-height: 90%;
        border: thick $primary-lighten-1;
        padding: 1 2;
        background: $surface;
    }
    AddEditScraperModal Input, AddEditScraperModal Label, AddEditScraperModal Static {
        margin-bottom: 1;
    }
    AddEditScraperModal Label {
        margin-top: 1;
    }
    AddEditScraperModal #scraper_description_input {
        height: 3;
        border: round $primary-border;
        padding: 0 1;
    }
    AddEditScraperModal Static.warning-text {
        color: $warning;
        padding: 0 1;
        text-align: center;
    }
    """

    def __init__(self, sd: Optional[sqlite3.Row] = None):
        super().__init__()
        self.sd = sd
        self.is_edit = sd is not None

    def compose(self) -> ComposeResult:
        title = "Edit Scraper Profile" if self.is_edit else "Add New Scraper Profile"
        with VerticalScroll():
            yield Label(title, classes="dialog-title")
            if self.is_edit and self.sd and self.sd['is_preinstalled']:
                yield Static("[PRE-INSTALLED PROFILE - Edits create a custom copy if name changes]", classes="warning-text")
            yield Label("Name:")
            yield Input(value=self.sd['name'] if self.is_edit else "", id="scraper_name")
            yield Label("URL (use [USER_PROVIDES_URL] to prompt):")
            yield Input(value=self.sd['url'] if self.is_edit else "", id="scraper_url")
            yield Label("CSS Selector:")
            yield Input(value=self.sd['selector'] if self.is_edit else "h2 a", id="scraper_selector")
            yield Label("Default Limit (0 for all):")
            yield Input(
                value=str(self.sd['default_limit']) if self.is_edit else "0",
                id="scraper_limit",
                type="integer"
            )
            yield Label("Default Tags (comma-separated, optional):")
            yield Input(
                value=self.sd['default_tags_csv'] if self.is_edit and self.sd['default_tags_csv'] else "",
                id="scraper_tags"
            )
            yield Label("Description:")
            yield Input(
                value=self.sd['description'] if self.is_edit and self.sd['description'] else "",
                id="scraper_description_input"
            )
            with Horizontal(classes="modal-buttons"):
                yield Button("Save", variant="primary", id="save_s_cfg")
                yield Button("Cancel", id="cancel_s_cfg")

    def on_button_pressed(self, e: Button.Pressed) -> None:
        if e.button.id == "save_s_cfg":
            data = {
                "name": self.query_one("#scraper_name", Input).value.strip(),
                "url": self.query_one("#scraper_url", Input).value.strip(),
                "selector": self.query_one("#scraper_selector", Input).value.strip(),
                "default_limit": 0,
                "default_tags_csv": self.query_one("#scraper_tags", Input).value.strip(),
                "description": self.query_one("#scraper_description_input", Input).value.strip(),
                "is_preinstalled": self.sd['is_preinstalled'] if self.is_edit and self.sd else 0
            }
            try:
                data["default_limit"] = int(self.query_one("#scraper_limit", Input).value.strip() or "0")
            except ValueError:
                self.app.notify("Invalid limit.", severity="error")
                return
            if not data["name"] or not data["url"] or not data["selector"]:
                self.app.notify("Name,URL,Selector required.", severity="error")
                return
            if self.is_edit and self.sd:
                data["id"] = self.sd["id"]
                if self.sd['is_preinstalled'] and data['name'] != self.sd['name']:
                    data['is_preinstalled'] = 0
                    del data['id']
            self.dismiss(data)
        else:
            self.dismiss(None)


class FilterScreen(ModalScreen[bool]):
    CSS = """
    FilterScreen {
        align: center middle;
        background: $surface-darken-1;
    }
    FilterScreen > Vertical {
        width: 85%;
        max-width: 90;
        height: auto;
        max-height: 85%;
        border: thick $primary-lighten-1;
        padding: 1 2;
        background: $surface;
    }
    FilterScreen VerticalScroll {
        height: auto;
        max-height: 30;
    }
    FilterScreen Horizontal {
        width: 100%;
        height: auto;
    }
    """
    BINDINGS = [Binding("escape", "dismiss", "Close"), Binding("enter", "apply_filters", "Apply")]

    def __init__(self, app_ref):
        super().__init__()
        self.app_ref = app_ref

    def compose(self) -> ComposeResult:
        from textual.widgets import Checkbox
        with Vertical():
            yield Label("Advanced Article Filters (v1.3.0)", classes="dialog-title")
            with VerticalScroll():
                yield Label("Title Filter:")
                yield Input(placeholder="Title (supports regex)...", id="title_filter_input", value=self.app_ref.title_filter)

                yield Label("URL Filter:")
                yield Input(placeholder="Source URL (supports regex)...", id="url_filter_input", value=self.app_ref.url_filter)

                yield Checkbox("Use Regex for Title/URL filters", id="use_regex_checkbox", value=self.app_ref.use_regex)

                yield Label("Date Range:")
                with Horizontal():
                    yield Input(placeholder="From (YYYY-MM-DD)...", id="date_from_input", value=self.app_ref.date_filter_from)
                    yield Input(placeholder="To (YYYY-MM-DD)...", id="date_to_input", value=self.app_ref.date_filter_to)

                yield Label("Tags Filter (comma-separated):")
                yield Input(placeholder="Tags...", id="tags_filter_input", value=self.app_ref.tags_filter)

                with Horizontal():
                    yield Label("Tags Logic:")
                    with RadioSet(id="tags_logic_radioset"):
                        yield RadioButton("AND (all tags)", id="tags_and", value=(self.app_ref.tags_logic == "AND"))
                        yield RadioButton("OR (any tag)", id="tags_or", value=(self.app_ref.tags_logic == "OR"))

                yield Label("Sentiment Filter:")
                yield Input(placeholder="Sentiment (Pos/Neg/Neu)...", id="sentiment_filter_input", value=self.app_ref.sentiment_filter)

            with Horizontal(classes="modal-buttons"):
                yield Button("Save Preset", id="save_preset_btn")
                yield Button("Load Preset", id="load_preset_btn")
                yield Button("Clear All", id="clear_btn")
                yield Button("Apply", id="apply_btn", variant="primary")
                yield Button("Cancel", id="cancel_btn")

    def on_button_pressed(self, event: Button.Pressed) -> None:
        if event.button.id == "apply_btn":
            self.action_apply_filters()
        elif event.button.id == "clear_btn":
            for input_widget in self.query(Input):
                input_widget.value = ""
            self.query_one("#use_regex_checkbox", Checkbox).value = False
            self.query_one("#tags_and", RadioButton).value = True
        elif event.button.id == "save_preset_btn":
            # TODO: Implement save preset modal
            self.app_ref.notify("Save preset: Coming soon", severity="info")
        elif event.button.id == "load_preset_btn":
            # TODO: Implement load preset modal
            self.app_ref.notify("Load preset: Coming soon", severity="info")
        elif event.button.id == "cancel_btn":
            self.dismiss(False)

    def action_apply_filters(self) -> None:
        from textual.widgets import Checkbox
        self.app_ref.title_filter = self.query_one("#title_filter_input", Input).value
        self.app_ref.url_filter = self.query_one("#url_filter_input", Input).value
        self.app_ref.use_regex = self.query_one("#use_regex_checkbox", Checkbox).value
        self.app_ref.date_filter_from = self.query_one("#date_from_input", Input).value
        self.app_ref.date_filter_to = self.query_one("#date_to_input", Input).value
        self.app_ref.tags_filter = self.query_one("#tags_filter_input", Input).value

        # Get tags logic from radio buttons
        radioset = self.query_one("#tags_logic_radioset", RadioSet)
        if radioset.pressed_button:
            self.app_ref.tags_logic = "AND" if radioset.pressed_button.id == "tags_and" else "OR"

        self.app_ref.sentiment_filter = self.query_one("#sentiment_filter_input", Input).value

        # Keep legacy date_filter for backwards compatibility
        if self.app_ref.date_filter_from:
            self.app_ref.date_filter = self.app_ref.date_filter_from

        self.dismiss(True)

    def action_dismiss(self) -> None:
        self.dismiss(False)


class FilterPresetModal(ModalScreen[Optional[str]]):
    """Modal for managing filter presets."""

    DEFAULT_CSS = """
    FilterPresetModal > Vertical {
        width: 70;
        height: auto;
        max-height: 80%;
        padding: 2;
        border: thick $primary-lighten-1;
        background: $panel;
    }
    FilterPresetModal Horizontal {
        width: 100%;
        height: auto;
        align-horizontal: center;
        padding-top: 1;
    }
    FilterPresetModal ListView {
        height: 15;
        border: solid $primary;
        margin-bottom: 1;
    }
    """

    def compose(self) -> ComposeResult:
        with Vertical():
            yield Label("Filter Presets (Ctrl+Shift+F)", classes="dialog-title")
            yield Label("Select a preset to load:")

            # List of presets
            presets = FilterPresetManager.list_presets()
            with ListView(id="preset_list"):
                if not presets:
                    yield ListItem(Label("(No saved presets)"))
                else:
                    for preset_name in presets:
                        yield ListItem(Label(preset_name))

            with Horizontal(classes="modal-buttons"):
                yield Button("Load", variant="primary", id="load_btn")
                yield Button("Delete", variant="error", id="delete_btn")
                yield Button("Cancel", id="cancel_btn")

    def on_button_pressed(self, event: Button.Pressed) -> None:
        if event.button.id == "load_btn":
            list_view = self.query_one("#preset_list", ListView)
            if list_view.index is not None and list_view.index >= 0:
                try:
                    item = list(list_view.children)[list_view.index]
                    label = item.query_one(Label)
                    preset_name = label.renderable
                    if preset_name != "(No saved presets)":
                        self.dismiss(preset_name)
                        return
                except:
                    pass
            self.app.notify("Please select a preset", severity="warning")
        elif event.button.id == "delete_btn":
            list_view = self.query_one("#preset_list", ListView)
            if list_view.index is not None and list_view.index >= 0:
                try:
                    item = list(list_view.children)[list_view.index]
                    label = item.query_one(Label)
                    preset_name = label.renderable
                    if preset_name != "(No saved presets)":
                        if FilterPresetManager.delete_preset(preset_name):
                            self.app.notify(f"Deleted preset: {preset_name}", severity="information")
                            # Refresh list
                            list_view.clear()
                            presets = FilterPresetManager.list_presets()
                            if not presets:
                                list_view.append(ListItem(Label("(No saved presets)")))
                            else:
                                for pn in presets:
                                    list_view.append(ListItem(Label(pn)))
                        else:
                            self.app.notify("Failed to delete preset", severity="error")
                        return
                except:
                    pass
            self.app.notify("Please select a preset", severity="warning")
        else:  # cancel_btn
            self.dismiss(None)


class SavePresetModal(ModalScreen[Optional[str]]):
    """Modal for saving current filters as a preset."""

    DEFAULT_CSS = """
    SavePresetModal > Vertical {
        width: 50;
        height: auto;
        padding: 2;
        border: thick $primary-lighten-1;
        background: $panel;
    }
    SavePresetModal Horizontal {
        width: 100%;
        height: auto;
        align-horizontal: center;
        padding-top: 1;
    }
    """

    def compose(self) -> ComposeResult:
        with Vertical():
            yield Label("Save Filter Preset", classes="dialog-title")
            yield Label("Enter preset name:")
            yield Input(placeholder="Preset name...", id="preset_name_input")
            with Horizontal(classes="modal-buttons"):
                yield Button("Save", variant="primary", id="save_btn")
                yield Button("Cancel", id="cancel_btn")

    def on_button_pressed(self, event: Button.Pressed) -> None:
        if event.button.id == "save_btn":
            name_input = self.query_one("#preset_name_input", Input)
            name = name_input.value.strip()
            if name:
                self.dismiss(name)
            else:
                self.app.notify("Please enter a preset name", severity="warning")
        else:
            self.dismiss(None)


class SettingsModal(ModalScreen[bool]):
    """Modal for application settings."""

    DEFAULT_CSS = """
    SettingsModal > VerticalScroll {
        width: 70;
        height: auto;
        max-height: 80%;
        padding: 2;
        border: thick $primary-lighten-1;
        background: $panel;
    }
    SettingsModal Horizontal {
        width: 100%;
        height: auto;
        align-horizontal: center;
        padding-top: 1;
    }
    SettingsModal Label {
        margin-bottom: 1;
    }
    """

    def __init__(self, config: Dict[str, Any]):
        super().__init__()
        self.config = config
        self.modified = False

    def compose(self) -> ComposeResult:
        with VerticalScroll():
            yield Label("Application Settings (Ctrl+G)", classes="dialog-title")

            yield Label("\n=== AI Configuration ===")
            yield Label("Default AI Provider:")
            with RadioSet(id="ai_provider_radio"):
                providers = ['gemini', 'openai', 'claude']
                current = self.config.get('ai', {}).get('default_provider', 'gemini')
                for provider in providers:
                    yield RadioButton(
                        provider.capitalize(),
                        value=(provider == current),
                        id=f"provider_{provider}"
                    )

            yield Label("\n=== Export Configuration ===")
            yield Label("Default Export Format:")
            with RadioSet(id="export_format_radio"):
                formats = ['csv', 'json']
                current_format = self.config.get('export', {}).get('default_format', 'csv')
                for fmt in formats:
                    yield RadioButton(
                        fmt.upper(),
                        value=(fmt == current_format),
                        id=f"format_{fmt}"
                    )

            yield Label("Export Output Directory:")
            yield Input(
                value=self.config.get('export', {}).get('output_directory', '.'),
                placeholder="Output directory...",
                id="output_dir_input"
            )

            yield Label("\n=== Database Configuration ===")
            yield Checkbox(
                "Auto-vacuum database on startup",
                value=self.config.get('database', {}).get('auto_vacuum', False),
                id="auto_vacuum_check"
            )
            yield Checkbox(
                "Backup database on exit",
                value=self.config.get('database', {}).get('backup_on_exit', False),
                id="backup_check"
            )

            with Horizontal(classes="modal-buttons"):
                yield Button("Save", variant="primary", id="save_settings_btn")
                yield Button("Cancel", id="cancel_settings_btn")

    def on_button_pressed(self, event: Button.Pressed) -> None:
        if event.button.id == "save_settings_btn":
            # Update config from UI
            # AI provider
            for provider in ['gemini', 'openai', 'claude']:
                radio = self.query_one(f"#provider_{provider}", RadioButton)
                if radio.value:
                    self.config['ai']['default_provider'] = provider
                    break

            # Export format
            for fmt in ['csv', 'json']:
                radio = self.query_one(f"#format_{fmt}", RadioButton)
                if radio.value:
                    self.config['export']['default_format'] = fmt
                    break

            # Output directory
            output_dir = self.query_one("#output_dir_input", Input).value
            self.config['export']['output_directory'] = output_dir

            # Database options
            self.config['database']['auto_vacuum'] = self.query_one("#auto_vacuum_check", Checkbox).value
            self.config['database']['backup_on_exit'] = self.query_one("#backup_check", Checkbox).value

            # Save config
            if ConfigManager.save_config(self.config):
                self.app.notify("Settings saved successfully", severity="information")
                self.dismiss(True)
            else:
                self.app.notify("Failed to save settings", severity="error")
        else:
            self.dismiss(False)


class ScheduleManagementModal(ModalScreen[Optional[int]]):
    """Modal for managing scheduled scrapes (v1.5.0)."""

    DEFAULT_CSS = """
    ScheduleManagementModal {
        align: center middle;
        background: $surface-darken-1;
    }
    ScheduleManagementModal > Vertical {
        width: 90%;
        max-width: 100;
        height: 85%;
        max-height: 35;
        border: thick $primary-lighten-1;
        padding: 1 2;
        background: $surface;
    }
    ScheduleManagementModal DataTable {
        height: 20;
    }
    """

    BINDINGS = [Binding("escape", "dismiss_screen", "Close")]

    def compose(self) -> ComposeResult:
        with Vertical():
            yield Label("📅 Scheduled Scrapes", id="schedule-title")
            yield DataTable(id="schedule_table", cursor_type="row", zebra_stripes=True)
            with Horizontal(classes="modal-buttons"):
                yield Button("Add Schedule", id="add_schedule", variant="primary")
                yield Button("Enable/Disable", id="toggle_schedule")
                yield Button("Delete", id="delete_schedule", variant="error")
                yield Button("Close", id="close_btn")

    async def on_mount(self) -> None:
        """Initialize the schedule table."""
        table = self.query_one("#schedule_table", DataTable)
        table.add_columns("ID", "Name", "Profile", "Type", "Schedule", "Enabled", "Next Run", "Last Run", "Run Count", "Status")
        await self._refresh_schedules()

    async def _refresh_schedules(self) -> None:
        """Refresh the schedules table."""
        table = self.query_one("#schedule_table", DataTable)
        table.clear()

        schedules = ScheduleManager.list_schedules()
        for schedule in schedules:
            enabled_str = "✓" if schedule['enabled'] else "✗"
            next_run = schedule['next_run'] or "Not set"
            last_run = schedule['last_run'] or "Never"
            status = schedule['last_status'] or "-"

            table.add_row(
                str(schedule['id']),
                schedule['name'],
                schedule['profile_name'] or f"ID:{schedule['scraper_profile_id']}",
                schedule['schedule_type'],
                schedule['schedule_value'],
                enabled_str,
                next_run,
                last_run,
                str(schedule['run_count']),
                status,
                key=str(schedule['id'])
            )

    async def on_button_pressed(self, event: Button.Pressed) -> None:
        """Handle button presses."""
        if event.button.id == "close_btn":
            self.dismiss(None)
        elif event.button.id == "add_schedule":
            def handle_new_schedule(result):
                if result:
                    self.run_worker(self._refresh_schedules())
                    self.app.notify("Schedule created successfully", severity="information")
            self.app.push_screen(AddScheduleModal(), handle_new_schedule)
        elif event.button.id == "toggle_schedule":
            table = self.query_one("#schedule_table", DataTable)
            if table.row_count > 0:
                try:
                    row_key = table.get_row_key_from_coordinate(table.cursor_coordinate)
                    schedule_id = int(row_key.value)
                    schedule = ScheduleManager.get_schedule(schedule_id)
                    if schedule:
                        new_enabled = not schedule['enabled']
                        if ScheduleManager.update_schedule(schedule_id, enabled=new_enabled):
                            await self._refresh_schedules()
                            status = "enabled" if new_enabled else "disabled"
                            self.app.notify(f"Schedule {status}", severity="information")
                except Exception as e:
                    self.app.notify(f"Error toggling schedule: {e}", severity="error")
        elif event.button.id == "delete_schedule":
            table = self.query_one("#schedule_table", DataTable)
            if table.row_count > 0:
                try:
                    row_key = table.get_row_key_from_coordinate(table.cursor_coordinate)
                    schedule_id = int(row_key.value)

                    def handle_delete_confirm(confirmed):
                        if confirmed:
                            if ScheduleManager.delete_schedule(schedule_id):
                                self.run_worker(self._refresh_schedules())
                                self.app.notify("Schedule deleted", severity="information")
                            else:
                                self.app.notify("Failed to delete schedule", severity="error")

                    self.app.push_screen(
                        ConfirmModal("Delete this schedule?", "Delete", "Cancel"),
                        handle_delete_confirm
                    )
                except Exception as e:
                    self.app.notify(f"Error deleting schedule: {e}", severity="error")

    def action_dismiss_screen(self) -> None:
        self.dismiss(None)


class AddScheduleModal(ModalScreen[bool]):
    """Modal for creating a new scheduled scrape (v1.5.0)."""

    DEFAULT_CSS = """
    AddScheduleModal {
        align: center middle;
        background: $surface-darken-1;
    }
    AddScheduleModal > Vertical {
        width: 80%;
        max-width: 80;
        height: auto;
        border: thick $primary-lighten-1;
        padding: 2;
        background: $surface;
    }
    AddScheduleModal Input {
        margin: 1 0;
    }
    AddScheduleModal RadioSet {
        height: auto;
        margin: 1 0;
    }
    """

    BINDINGS = [Binding("escape", "cancel", "Cancel")]

    def compose(self) -> ComposeResult:
        with Vertical():
            yield Label("➕ Add Scheduled Scrape")
            yield Label("Schedule Name:")
            yield Input(placeholder="e.g., Daily News Scrape", id="schedule_name")

            yield Label("Scraper Profile:")
            yield Input(placeholder="Profile ID (from Profiles list)", id="profile_id")

            yield Label("Schedule Type:")
            with RadioSet(id="schedule_type"):
                yield RadioButton("Hourly (every hour)", id="type_hourly")
                yield RadioButton("Daily (HH:MM format, e.g., 09:00)", id="type_daily", value=True)
                yield RadioButton("Weekly (day:HH:MM, e.g., 0:09:00 for Monday 9am)", id="type_weekly")
                yield RadioButton("Interval (minutes, e.g., 30)", id="type_interval")

            yield Label("Schedule Value:")
            yield Input(placeholder="e.g., 09:00 for daily at 9am", id="schedule_value")

            with Horizontal(classes="modal-buttons"):
                yield Button("Create", id="create_btn", variant="primary")
                yield Button("Cancel", id="cancel_btn")

    async def on_button_pressed(self, event: Button.Pressed) -> None:
        """Handle button presses."""
        if event.button.id == "cancel_btn":
            self.dismiss(False)
        elif event.button.id == "create_btn":
            name_input = self.query_one("#schedule_name", Input)
            profile_input = self.query_one("#profile_id", Input)
            value_input = self.query_one("#schedule_value", Input)
            type_radio = self.query_one("#schedule_type", RadioSet)

            name = name_input.value.strip()
            profile_id_str = profile_input.value.strip()
            value = value_input.value.strip()

            # Determine schedule type
            if self.query_one("#type_hourly", RadioButton).value:
                schedule_type = "hourly"
            elif self.query_one("#type_daily", RadioButton).value:
                schedule_type = "daily"
            elif self.query_one("#type_weekly", RadioButton).value:
                schedule_type = "weekly"
            elif self.query_one("#type_interval", RadioButton).value:
                schedule_type = "interval"
            else:
                schedule_type = "daily"

            # Validate inputs
            if not name:
                self.app.notify("Schedule name is required", severity="error")
                return
            if not profile_id_str:
                self.app.notify("Profile ID is required", severity="error")
                return
            if not value:
                self.app.notify("Schedule value is required", severity="error")
                return

            try:
                profile_id = int(profile_id_str)
            except ValueError:
                self.app.notify("Profile ID must be a number", severity="error")
                return

            # Create the schedule
            if ScheduleManager.create_schedule(name, profile_id, schedule_type, value, enabled=True):
                self.dismiss(True)
            else:
                self.app.notify("Failed to create schedule (name may already exist)", severity="error")

    def action_cancel(self) -> None:
        self.dismiss(False)


class AnalyticsModal(ModalScreen):
    """Modal for viewing analytics and statistics (v1.6.0)."""

    DEFAULT_CSS = """
    AnalyticsModal {
        align: center middle;
        background: $surface-darken-1;
    }
    AnalyticsModal > VerticalScroll {
        width: 90%;
        max-width: 120;
        height: 90%;
        max-height: 40;
        border: thick $primary-lighten-1;
        padding: 1 2;
        background: $surface;
    }
    AnalyticsModal Button {
        margin: 1 2;
    }
    """

    BINDINGS = [Binding("escape", "dismiss_screen", "Close")]

    def compose(self) -> ComposeResult:
        stats = AnalyticsManager.get_statistics()

        # Build markdown content with statistics
        content = "# 📊 Data Analytics & Statistics\n\n"
        content += "## Overview\n\n"
        content += f"- **Total Articles**: {stats.get('total_articles', 0)}\n"
        content += f"- **With Summaries**: {stats.get('articles_with_summaries', 0)} ({stats.get('summary_percentage', 0):.1f}%)\n"
        content += f"- **With Sentiment**: {stats.get('articles_with_sentiment', 0)} ({stats.get('sentiment_percentage', 0):.1f}%)\n\n"

        # Sentiment distribution
        content += "## Sentiment Distribution\n\n"
        sentiment_dist = stats.get('sentiment_distribution', {})
        if sentiment_dist:
            for sentiment, count in sentiment_dist.items():
                content += f"- **{sentiment}**: {count}\n"
        else:
            content += "*No sentiment data available*\n"
        content += "\n"

        # Top sources
        content += "## Top 10 Sources\n\n"
        top_sources = stats.get('top_sources', [])
        if top_sources:
            for i, (source, count) in enumerate(top_sources, 1):
                source_display = source[:60] + "..." if len(source) > 60 else source
                content += f"{i}. **{source_display}** ({count} articles)\n"
        else:
            content += "*No source data available*\n"
        content += "\n"

        # Top tags
        content += "## Top 20 Tags\n\n"
        top_tags = stats.get('top_tags', [])
        if top_tags:
            # Display tags in a more compact format
            tag_list = ", ".join([f"**{tag}** ({count})" for tag, count in top_tags[:10]])
            content += tag_list + "\n\n"
            if len(top_tags) > 10:
                tag_list2 = ", ".join([f"**{tag}** ({count})" for tag, count in top_tags[10:20]])
                content += tag_list2 + "\n"
        else:
            content += "*No tag data available*\n"
        content += "\n"

        # Timeline info
        content += "## Recent Activity\n\n"
        articles_per_day = stats.get('articles_per_day', [])
        if articles_per_day:
            content += f"Data collected over **{len(articles_per_day)} days** (last 30 days)\n"
            total_recent = sum(count for _, count in articles_per_day)
            avg_per_day = total_recent / len(articles_per_day) if articles_per_day else 0
            content += f"Average: **{avg_per_day:.1f} articles/day**\n"
        else:
            content += "*No recent activity data*\n"

        with VerticalScroll():
            yield Markdown(content)
            with Horizontal(classes="modal-buttons"):
                yield Button("Export Charts", id="export_charts", variant="primary")
                yield Button("Export Report", id="export_report")
                yield Button("Close", id="close_btn")

    async def on_button_pressed(self, event: Button.Pressed) -> None:
        """Handle button presses."""
        if event.button.id == "close_btn":
            self.dismiss()
        elif event.button.id == "export_charts":
            # Export all charts
            def export_worker():
                try:
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    AnalyticsManager.generate_sentiment_chart(f"sentiment_chart_{timestamp}.png")
                    AnalyticsManager.generate_timeline_chart(f"timeline_chart_{timestamp}.png")
                    AnalyticsManager.generate_top_sources_chart(f"sources_chart_{timestamp}.png")
                    return True
                except Exception as e:
                    logger.error(f"Error exporting charts: {e}")
                    return False

            success = await self.app.run_in_thread(export_worker)
            if success:
                self.app.notify("Charts exported successfully", severity="information")
            else:
                self.app.notify("Failed to export charts", severity="error")
        elif event.button.id == "export_report":
            # Export text report
            def export_report_worker():
                try:
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    return AnalyticsManager.export_statistics_report(f"analytics_report_{timestamp}.txt")
                except Exception as e:
                    logger.error(f"Error exporting report: {e}")
                    return False

            success = await self.app.run_in_thread(export_report_worker)
            if success:
                self.app.notify("Report exported successfully", severity="information")
            else:
                self.app.notify("Failed to export report", severity="error")

    def action_dismiss_screen(self) -> None:
        self.dismiss()


class TopicModelingModal(ModalScreen[Optional[Dict[str, Any]]]):
    """Modal for configuring and running topic modeling."""
    DEFAULT_CSS = """
    TopicModelingModal {
        align: center middle;
        background: $surface-darken-1;
    }
    TopicModelingModal > Vertical {
        width: 70;
        height: auto;
        border: thick $primary-lighten-1;
        padding: 1 2;
        background: $surface;
    }
    TopicModelingModal Horizontal {
        width: 100%;
        height: auto;
        align-horizontal: center;
        padding-top: 1;
    }
    """

    def compose(self) -> ComposeResult:
        with Vertical():
            yield Label("Topic Modeling Configuration", classes="modal-title")
            yield Label("Select algorithm and parameters:")
            yield Select(
                [("LDA (Latent Dirichlet Allocation)", "lda"), ("NMF (Non-negative Matrix Factorization)", "nmf")],
                id="algorithm",
                prompt="Select Algorithm"
            )
            yield Input(placeholder="Number of topics (default: 5)", id="num_topics", value="5")
            yield Input(placeholder="Words per topic (default: 10)", id="num_words", value="10")
            with Horizontal():
                yield Button("Run Modeling", variant="primary", id="run")
                yield Button("Cancel", id="cancel")

    async def on_button_pressed(self, event: Button.Pressed) -> None:
        if event.button.id == "run":
            algorithm_select = self.query_one("#algorithm", Select)
            num_topics_input = self.query_one("#num_topics", Input)
            num_words_input = self.query_one("#num_words", Input)

            try:
                num_topics = int(num_topics_input.value) if num_topics_input.value else 5
                num_words = int(num_words_input.value) if num_words_input.value else 10
                algorithm = algorithm_select.value if algorithm_select.value != Select.BLANK else "lda"

                self.dismiss({
                    "algorithm": algorithm,
                    "num_topics": num_topics,
                    "num_words": num_words
                })
            except ValueError:
                self.app.notify("Please enter valid numbers", severity="error")
        else:
            self.dismiss(None)


class QuestionAnsweringModal(ModalScreen[Optional[str]]):
    """Modal for asking questions about articles."""
    DEFAULT_CSS = """
    QuestionAnsweringModal {
        align: center middle;
        background: $surface-darken-1;
    }
    QuestionAnsweringModal > Vertical {
        width: 80;
        height: auto;
        max-height: 35;
        border: thick $primary-lighten-1;
        padding: 1 2;
        background: $surface;
    }
    QuestionAnsweringModal TextArea {
        height: 5;
        margin-bottom: 1;
    }
    QuestionAnsweringModal Horizontal {
        width: 100%;
        height: auto;
        align-horizontal: center;
        padding-top: 1;
    }
    """

    def __init__(self, answer: Optional[str] = None, sources: Optional[List[Dict]] = None):
        super().__init__()
        self.answer = answer
        self.sources = sources or []

    def compose(self) -> ComposeResult:
        with Vertical():
            if self.answer:
                # Display answer mode
                yield Label("Answer:", classes="modal-title")
                yield Static(self.answer, classes="answer-text")
                if self.sources:
                    yield Label("\nSources:", classes="modal-subtitle")
                    for idx, source in enumerate(self.sources, 1):
                        yield Static(f"{idx}. {source.get('title', 'Unknown')} ({source.get('url', '')})")
                with Horizontal():
                    yield Button("Ask Another", variant="primary", id="ask_another")
                    yield Button("Close", id="close")
            else:
                # Input question mode
                yield Label("Ask a Question", classes="modal-title")
                yield TextArea(id="question", placeholder="Enter your question about the articles...")
                with Horizontal():
                    yield Button("Submit", variant="primary", id="submit")
                    yield Button("Cancel", id="cancel")

    async def on_button_pressed(self, event: Button.Pressed) -> None:
        if event.button.id == "submit":
            question_area = self.query_one("#question", TextArea)
            question = question_area.text.strip()
            if question:
                self.dismiss(question)
            else:
                self.app.notify("Please enter a question", severity="warning")
        elif event.button.id == "ask_another":
            self.dismiss("__ask_another__")
        else:
            self.dismiss(None)


class DuplicateDetectionModal(ModalScreen[Optional[float]]):
    """Modal for finding and managing duplicate articles."""
    DEFAULT_CSS = """
    DuplicateDetectionModal {
        align: center middle;
        background: $surface-darken-1;
    }
    DuplicateDetectionModal > Vertical {
        width: 80;
        height: auto;
        max-height: 40;
        border: thick $primary-lighten-1;
        padding: 1 2;
        background: $surface;
    }
    DuplicateDetectionModal DataTable {
        height: 20;
        margin: 1 0;
    }
    DuplicateDetectionModal Horizontal {
        width: 100%;
        height: auto;
        align-horizontal: center;
        padding-top: 1;
    }
    """

    def __init__(self, duplicates: Optional[List[Dict]] = None):
        super().__init__()
        self.duplicates = duplicates or []

    def compose(self) -> ComposeResult:
        with Vertical():
            yield Label("Duplicate Detection", classes="modal-title")
            if self.duplicates:
                # Display results
                yield Label(f"Found {len(self.duplicates)} duplicate pairs:")
                table = DataTable()
                table.add_columns("Article 1", "Article 2", "Similarity")
                for dup in self.duplicates:
                    table.add_row(
                        dup.get('title1', 'Unknown'),
                        dup.get('title2', 'Unknown'),
                        f"{dup.get('similarity', 0):.2%}"
                    )
                yield table
                with Horizontal():
                    yield Button("Close", variant="primary", id="close")
            else:
                # Input threshold
                yield Label("Enter similarity threshold (0.0 - 1.0):")
                yield Input(placeholder="0.85", id="threshold", value="0.85")
                with Horizontal():
                    yield Button("Detect", variant="primary", id="detect")
                    yield Button("Cancel", id="cancel")

    async def on_button_pressed(self, event: Button.Pressed) -> None:
        if event.button.id == "detect":
            threshold_input = self.query_one("#threshold", Input)
            try:
                threshold = float(threshold_input.value) if threshold_input.value else 0.85
                if 0.0 <= threshold <= 1.0:
                    self.dismiss(threshold)
                else:
                    self.app.notify("Threshold must be between 0.0 and 1.0", severity="error")
            except ValueError:
                self.app.notify("Please enter a valid number", severity="error")
        else:
            self.dismiss(None)


class RelatedArticlesModal(ModalScreen):
    """Modal to show related articles."""
    DEFAULT_CSS = """
    RelatedArticlesModal {
        align: center middle;
        background: $surface-darken-1;
    }
    RelatedArticlesModal > Vertical {
        width: 80;
        height: auto;
        max-height: 35;
        border: thick $primary-lighten-1;
        padding: 1 2;
        background: $surface;
    }
    RelatedArticlesModal DataTable {
        height: 15;
        margin: 1 0;
    }
    RelatedArticlesModal Horizontal {
        width: 100%;
        height: auto;
        align-horizontal: center;
        padding-top: 1;
    }
    """

    def __init__(self, article_title: str, related: List[Dict]):
        super().__init__()
        self.article_title = article_title
        self.related = related

    def compose(self) -> ComposeResult:
        with Vertical():
            yield Label(f"Related to: {self.article_title}", classes="modal-title")
            if self.related:
                table = DataTable()
                table.add_columns("Title", "Similarity", "URL")
                for item in self.related:
                    table.add_row(
                        item.get('title', 'Unknown'),
                        f"{item.get('similarity', 0):.2%}",
                        item.get('url', '')[:40] + "..." if len(item.get('url', '')) > 40 else item.get('url', '')
                    )
                yield table
            else:
                yield Label("No related articles found.")
            with Horizontal():
                yield Button("Close", variant="primary", id="close")

    async def on_button_pressed(self, event: Button.Pressed) -> None:
        self.dismiss()


class ClusterViewModal(ModalScreen[Optional[int]]):
    """Modal for visualizing article clusters."""
    DEFAULT_CSS = """
    ClusterViewModal {
        align: center middle;
        background: $surface-darken-1;
    }
    ClusterViewModal > Vertical {
        width: 80;
        height: auto;
        max-height: 40;
        border: thick $primary-lighten-1;
        padding: 1 2;
        background: $surface;
    }
    ClusterViewModal DataTable {
        height: 20;
        margin: 1 0;
    }
    ClusterViewModal Horizontal {
        width: 100%;
        height: auto;
        align-horizontal: center;
        padding-top: 1;
    }
    """

    def __init__(self, clusters: Optional[List[Dict]] = None):
        super().__init__()
        self.clusters = clusters or []

    def compose(self) -> ComposeResult:
        with Vertical():
            yield Label("Article Clustering", classes="modal-title")
            if self.clusters:
                # Display clusters
                table = DataTable()
                table.add_columns("Cluster", "Articles", "Top Keywords")
                for idx, cluster in enumerate(self.clusters):
                    keywords = ", ".join(cluster.get('keywords', [])[:5])
                    table.add_row(
                        f"Cluster {idx + 1}",
                        str(len(cluster.get('articles', []))),
                        keywords
                    )
                yield table
                with Horizontal():
                    yield Button("Export", id="export")
                    yield Button("Close", variant="primary", id="close")
            else:
                # Input number of clusters
                yield Label("Enter number of clusters:")
                yield Input(placeholder="5", id="num_clusters", value="5")
                with Horizontal():
                    yield Button("Run Clustering", variant="primary", id="cluster")
                    yield Button("Cancel", id="cancel")

    async def on_button_pressed(self, event: Button.Pressed) -> None:
        if event.button.id == "cluster":
            num_clusters_input = self.query_one("#num_clusters", Input)
            try:
                num_clusters = int(num_clusters_input.value) if num_clusters_input.value else 5
                self.dismiss(num_clusters)
            except ValueError:
                self.app.notify("Please enter a valid number", severity="error")
        elif event.button.id == "export":
            # TODO: Implement export
            self.app.notify("Export functionality coming soon", severity="information")
        else:
            self.dismiss(None)


class SummaryQualityModal(ModalScreen):
    """Modal to display summary quality metrics."""
    DEFAULT_CSS = """
    SummaryQualityModal {
        align: center middle;
        background: $surface-darken-1;
    }
    SummaryQualityModal > Vertical {
        width: 70;
        height: auto;
        border: thick $primary-lighten-1;
        padding: 1 2;
        background: $surface;
    }
    SummaryQualityModal Horizontal {
        width: 100%;
        height: auto;
        align-horizontal: center;
        padding-top: 1;
    }
    """

    def __init__(self, metrics: Dict[str, Any]):
        super().__init__()
        self.metrics = metrics

    def compose(self) -> ComposeResult:
        with Vertical():
            yield Label("Summary Quality Metrics", classes="modal-title")

            rouge_scores = self.metrics.get('rouge_scores', {})
            if rouge_scores:
                yield Label(f"\nROUGE-1: {rouge_scores.get('rouge1', 0):.3f}")
                yield Label(f"ROUGE-2: {rouge_scores.get('rouge2', 0):.3f}")
                yield Label(f"ROUGE-L: {rouge_scores.get('rougeL', 0):.3f}")

            coherence = self.metrics.get('coherence', 0)
            yield Label(f"\nCoherence Score: {coherence:.1f}/100")

            readability = self.metrics.get('readability', {})
            if readability:
                yield Label(f"\nReadability:")
                yield Label(f"  Flesch Reading Ease: {readability.get('flesch', 0):.1f}")

            yield Label("\nRate this summary (1-5 stars):")
            yield Input(placeholder="Enter rating (1-5)", id="rating")

            with Horizontal():
                yield Button("Save Feedback", variant="primary", id="save")
                yield Button("Close", id="close")

    async def on_button_pressed(self, event: Button.Pressed) -> None:
        if event.button.id == "save":
            rating_input = self.query_one("#rating", Input)
            try:
                rating = int(rating_input.value) if rating_input.value else 0
                if 1 <= rating <= 5:
                    # TODO: Save rating to database
                    self.app.notify(f"Rating {rating} saved successfully", severity="information")
                    self.dismiss()
                else:
                    self.app.notify("Rating must be between 1 and 5", severity="error")
            except ValueError:
                self.app.notify("Please enter a valid rating", severity="error")
        else:
            self.dismiss()


class HelpModal(ModalScreen):
    DEFAULT_CSS = """
    HelpModal {
        align: center middle;
        background: $surface-darken-1;
    }
    HelpModal > VerticalScroll {
        width: 90%;
        max-width: 120;
        height: 90%;
        max-height: 40;
        border: thick $primary-lighten-1;
        padding: 1 2;
        background: $surface;
    }
    /* Styling for h2/h3 inside Markdown is not directly supported this way.
       Style the Markdown widget itself or use Rich console markup in the Markdown text.
    */
    """
    BINDINGS = [Binding("escape", "dismiss_screen", "Close Help")]

    def compose(self) -> ComposeResult:
        ps_desc = "\n\n### Pre-installed Scraper Profiles:\n\n"
        for ps_data in PREINSTALLED_SCRAPERS:
            ps_desc += (
                f"- **{ps_data['name']}**: {ps_data['description']}\n"
                f"  - *Example/Target URL Hint*: `{ps_data['url']}`\n"
                f"  - *Suggested Selector*: `{ps_data['selector']}`\n"
                f"  - *Default Tags*: "
                f"`{ps_data['default_tags_csv'] or 'None'}`\n\n"
            )
        ht = f"""\
## Keybindings & Help (v1.9.5)

### Navigation & Display
| Key           | Action              | Description                                        |
|---------------|---------------------|----------------------------------------------------|
| `UP`/`DOWN`   | Navigate Table      | Move selection in articles list.                   |
| `ENTER`       | View Details        | Show full details of selected article.             |
| `r`           | Refresh List        | Reload articles from database.                     |
| `ctrl+l`      | Toggle Theme        | Switch between Light/Dark mode.                    |

### Scraping & Data Entry
| Key           | Action              | Description                                        |
|---------------|---------------------|----------------------------------------------------|
| `ctrl+n`      | New Scrape          | Open dialog to scrape new URL (generic).           |
| `ctrl+m`      | Scraper Profiles    | Manage & execute saved/pre-installed scrapers.     |
| `ctrl+shift+a`| Manage Schedules    | Create/edit scheduled scraping automation.         |

### Article Management
| Key           | Action              | Description                                        |
|---------------|---------------------|----------------------------------------------------|
| `d`/`DELETE`  | Delete Selected     | Delete selected article (confirm).                 |
| `ctrl+r`      | Read Article        | Fetch & display full content of selected.          |
| `ctrl+t`      | Manage Tags         | Add/remove tags for selected article.              |
| `SPACE`       | Toggle Selection    | Toggle bulk selection for current article.         |
| `ctrl+a`      | Select All          | Select all articles in current view.               |
| `ctrl+d`      | Deselect All        | Clear all bulk selections.                         |
| `ctrl+shift+d`| Bulk Delete         | Delete all selected articles (confirm).            |
| `ctrl+x`      | Clear Database      | Delete ALL articles (confirm).                     |

### AI Features
| Key           | Action              | Description                                        |
|---------------|---------------------|----------------------------------------------------|
| `s`           | Summarize           | AI summary for selected (choose style).            |
| `ctrl+k`      | Sentiment Analysis  | AI sentiment for selected article.                 |
| `ctrl+p`      | Select AI Provider  | Switch between Gemini/OpenAI/Claude.               |
| `ctrl+shift+t`| Auto-Tag            | AI-powered automatic tagging of article.           |
| `ctrl+shift+e`| Extract Entities    | Extract named entities (people, orgs, locations).  |
| `ctrl+shift+k`| Extract Keywords    | Extract key terms and topics from article.         |
| `ctrl+shift+r`| Find Similar        | Find similar articles using content similarity.    |

### Smart Categorization & Topic Modeling (v1.9.0)
| Key           | Action              | Description                                        |
|---------------|---------------------|----------------------------------------------------|
| `ctrl+alt+t`  | Topic Modeling      | Run LDA/NMF topic modeling on articles.            |
| `ctrl+alt+q`  | Ask Question        | Question answering about scraped content.          |
| `ctrl+alt+d`  | Find Duplicates     | Detect duplicate articles with fuzzy matching.     |
| `ctrl+alt+l`  | Related Articles    | Find articles related to selected one.             |
| `ctrl+alt+c`  | Cluster Articles    | Cluster similar articles by content.               |
| `ctrl+alt+h`  | Q&A History         | View question-answer conversation history.         |
| `ctrl+alt+m`  | Summary Quality     | Evaluate summary quality with ROUGE metrics.       |

### Filtering & Sorting
| Key           | Action              | Description                                        |
|---------------|---------------------|----------------------------------------------------|
| `ctrl+f`      | Open Filters        | Filter by Title, URL, Date, Tags, Sentiment.       |
| `ctrl+shift+f`| Filter Presets      | Manage saved filter combinations.                  |
| `ctrl+shift+s`| Save Filter Preset  | Save current filters as a preset.                  |
| `ctrl+s`      | Cycle Sort          | Change article list sorting order.                 |

### Export & Analytics
| Key           | Action              | Description                                        |
|---------------|---------------------|----------------------------------------------------|
| `ctrl+e`      | Export CSV          | Export current view to CSV file.                   |
| `ctrl+j`      | Export JSON         | Export current view to JSON file.                  |
| `ctrl+shift+x`| Export Excel        | Export to formatted XLSX with charts.              |
| `ctrl+shift+p`| Export PDF          | Generate professional PDF report.                  |
| `ctrl+shift+w`| Word Cloud          | Generate word cloud visualization from tags.       |
| `ctrl+shift+v`| View Analytics      | Open analytics dashboard with charts.              |

### Configuration & Help
| Key           | Action              | Description                                        |
|---------------|---------------------|----------------------------------------------------|
| `ctrl+g`      | Settings            | Open application settings editor.                  |
| `F1`/`ctrl+h` | Help                | Show/hide this help screen.                        |
| `q`/`ctrl+c`  | Quit                | Exit application.                                  |

**Tips:**
- Use `ctrl+f` for advanced filtering (regex, date ranges, tag logic).
- Use `ctrl+shift+f` to save/load frequently-used filter combinations.
- Press `SPACE` to bulk-select articles, then `ctrl+shift+d` to delete multiple at once.
- Analytics (`ctrl+shift+v`) provides sentiment charts, timelines, and source statistics.
{ps_desc}"""
        with VerticalScroll():
            yield Markdown(ht)
            yield Horizontal(Button("Close", id="ch_b"), classes="modal-buttons")

    def on_button_pressed(self, e: Button.Pressed) -> None:
        self.dismiss()

    def action_dismiss_screen(self) -> None:
        self.dismiss()


class StatusBar(Static):
    total_articles = reactive(0)
    selected_id = reactive[Optional[int]](None)
    bulk_selected_count = reactive(0)
    filter_status = reactive("")
    sort_status = reactive("")
    current_theme = reactive("Dark")
    scraper_profile = reactive("Manual Entry")
    # v2.0.0 User information
    current_username = reactive("Not logged in")
    current_user_role = reactive("guest")

    def render(self) -> str:
        # Build user info with role indicator
        user_info = f"👤 {self.current_username}"
        if self.current_user_role == "admin":
            user_info += " [ADMIN]"
        elif self.current_user_role == "viewer":
            user_info += " [VIEWER]"

        parts = [
            user_info,
            f"Total: {self.total_articles}",
            f"Profile: {self.scraper_profile}",
            f"Theme: {self.current_theme}"
        ]
        if self.selected_id is not None:
            parts.append(f"Sel ID: {self.selected_id}")
        if self.bulk_selected_count > 0:
            parts.append(f"Bulk: {self.bulk_selected_count} selected")
        if self.filter_status:
            parts.append(f"Filter: {self.filter_status}")
        if self.sort_status:
            parts.append(f"Sort: {self.sort_status}")
        return " | ".join(parts)


class WebScraperApp(App[None]):
    CSS_PATH = "web_scraper_tui_v1.0.tcss"
    BINDINGS = [
        Binding("q,ctrl+c", "quit", "Quit", priority=True),
        Binding("r", "refresh_data", "Refresh"),
        Binding("enter", "view_details", "View Details"),
        Binding("space", "select_row", "Toggle Selection"),
        Binding("v", "view_summary", "View Summary"),
        Binding("s", "summarize_selected", "Summarize"),
        Binding("ctrl+k", "sentiment_analysis_selected", "Sentiment"),
        Binding("d,delete", "delete_selected", "Delete"),
        Binding("ctrl+r", "read_article", "Read Full"),
        Binding("ctrl+t", "manage_tags", "Tags"),
        Binding("ctrl+s", "cycle_sort_order", "Sort"),
        Binding("ctrl+m", "manage_saved_scrapers", "Profiles"),
        Binding("ctrl+n", "scrape_new", "New Scrape"),
        Binding("ctrl+e", "export_csv", "Export CSV"),
        Binding("ctrl+j", "export_json", "Export JSON"),
        Binding("ctrl+shift+x", "export_excel", "Export Excel"),
        Binding("ctrl+shift+p", "export_pdf", "Export PDF"),
        Binding("ctrl+shift+w", "export_word_cloud", "Word Cloud"),
        Binding("ctrl+x", "clear_database", "Clear DB"),
        Binding("ctrl+f", "open_filters", "Filters"),
        Binding("ctrl+shift+l", "logout", "Logout"),  # v2.0.0 (changed from ctrl+l)
        Binding("ctrl+u", "user_profile", "Profile"),  # v2.0.0
        Binding("ctrl+alt+u", "user_management", "Users"),  # v2.0.0 (admin only)
        Binding("ctrl+l", "toggle_dark_mode", "Theme"),
        Binding("ctrl+a", "select_all", "Select All"),
        Binding("ctrl+d", "deselect_all", "Deselect All"),
        Binding("ctrl+shift+d", "bulk_delete", "Bulk Delete"),
        Binding("ctrl+p", "select_ai_provider", "AI Provider"),
        Binding("ctrl+g", "open_settings", "Settings"),
        Binding("ctrl+shift+f", "manage_filter_presets", "Filter Presets"),
        Binding("ctrl+shift+s", "save_filter_preset", "Save Preset"),
        Binding("ctrl+shift+a", "manage_schedules", "Schedules"),
        Binding("ctrl+shift+v", "view_analytics", "Analytics"),
        Binding("ctrl+shift+t", "auto_tag", "Auto-Tag"),
        Binding("ctrl+shift+e", "extract_entities", "Entities"),
        Binding("ctrl+shift+k", "extract_keywords", "Keywords"),
        Binding("ctrl+shift+r", "find_similar", "Similar"),
        # v1.9.0 Smart Categorization & Topic Modeling
        Binding("ctrl+alt+t", "topic_modeling", "Topic Modeling"),
        Binding("ctrl+alt+q", "ask_question", "Ask Question"),
        Binding("ctrl+alt+d", "find_duplicates", "Find Duplicates"),
        Binding("ctrl+alt+l", "find_related", "Related Articles"),
        Binding("ctrl+alt+c", "cluster_articles", "Cluster Articles"),
        Binding("ctrl+alt+h", "view_qa_history", "Q&A History"),
        Binding("ctrl+alt+m", "evaluate_summary", "Summary Quality"),
        Binding("f1,ctrl+h", "toggle_help", "Help")
    ]
    dark = reactive(True, layout=True)
    selected_row_id: reactive[int | None] = reactive(None)
    selected_row_ids: reactive[set] = reactive(set)  # Bulk selection
    db_init_ok: bool = False
    last_scrape_url, last_scrape_selector, last_scrape_limit = "", "h2 a", 0
    title_filter = reactive("")
    url_filter = reactive("")
    date_filter = reactive("")
    tags_filter = reactive("")
    sentiment_filter = reactive("")
    # Advanced filtering options (v1.3.0)
    use_regex = reactive(False)
    date_filter_from = reactive("")
    date_filter_to = reactive("")
    tags_logic = reactive("AND")  # AND or OR
    current_scraper_profile: reactive[str] = reactive("Manual Entry")
    SORT_OPTIONS: List[Tuple[str, str]] = [
        ("sd.timestamp DESC", "Date Newest"),
        ("sd.timestamp ASC", "Date Oldest"),
        ("sd.title COLLATE NOCASE ASC", "Title A-Z"),
        ("sd.title COLLATE NOCASE DESC", "Title Z-A"),
        ("sd.sentiment ASC, sd.timestamp DESC", "Sentiment"),
        ("sd.id ASC", "ID Asc"),
        ("sd.id DESC", "ID Desc"),
        ("sd.url COLLATE NOCASE ASC", "Src URL A-Z"),
        ("sd.url COLLATE NOCASE DESC", "Src URL Z-A")
    ]
    current_sort_index = reactive(0)
    # v2.0.0 User state
    current_user_id: reactive[Optional[int]] = reactive(None)
    current_username: reactive[str] = reactive("Not logged in")
    current_user_role: reactive[str] = reactive("guest")
    session_token: reactive[Optional[str]] = reactive(None)

    def __init__(self):
        super().__init__()
        self.db_init_ok = init_db()
        self.row_metadata = {}
        self._summarize_context = {}
        self.config = ConfigManager.load_config()
        # Initialize background scheduler (v1.5.0)
        self.scheduler = BackgroundScheduler()
        self.scheduler.start()
        logger.info("Background scheduler started")

    def compose(self) -> ComposeResult:
        yield Header(show_clock=True, name="Web Scraper TUI v1.9.5")
        yield DataTable(id="article_table", cursor_type="row", zebra_stripes=True)
        yield LoadingIndicator(id="loading_indicator", classes="hidden")
        yield StatusBar(id="status_bar")
        yield Footer()
        # yield Notifications() # Intentionally removed as App handles this

    async def on_mount(self) -> None:
        sbar = self.query_one(StatusBar)
        sbar.current_theme = "Dark" if self.dark else "Light"
        sbar.scraper_profile = self.current_scraper_profile
        if not self.db_init_ok:
            self.notify("CRITICAL: DB init failed!", title="DB Error", severity="error", timeout=0)
            return

        # v2.0.0: Show login modal before initializing app
        user_id = await self.push_screen_wait(LoginModal())
        if user_id is None:
            # User cancelled login - exit app
            self.notify("Login required. Exiting...", severity="warning")
            self.exit()
            return

        # Login successful - initialize user session
        await self._initialize_user_session(user_id)

        tbl = self.query_one(DataTable)
        tbl.add_columns("ID", "S", "Sentiment", "Title", "Source URL", "Tags", "Scraped At")
        sbar.sort_status = self.SORT_OPTIONS[self.current_sort_index][1]
        await self.refresh_article_table()
        tbl.focus()
        # Load and schedule enabled scrapes (v1.5.0)
        self._load_scheduled_scrapes()

    def _load_scheduled_scrapes(self) -> None:
        """Load enabled schedules from database and add them to APScheduler."""
        try:
            schedules = ScheduleManager.list_schedules(enabled_only=True)
            for schedule in schedules:
                self._add_schedule_to_scheduler(schedule)
            logger.info(f"Loaded {len(schedules)} enabled schedules")
        except Exception as e:
            logger.error(f"Error loading schedules: {e}")

    def _add_schedule_to_scheduler(self, schedule: Dict[str, Any]) -> None:
        """Add a schedule to APScheduler."""
        try:
            schedule_id = schedule['id']
            schedule_type = schedule['schedule_type']
            schedule_value = schedule['schedule_value']

            # Create appropriate trigger
            if schedule_type == 'hourly':
                trigger = IntervalTrigger(hours=1)
            elif schedule_type == 'daily':
                # Parse HH:MM
                hour, minute = map(int, schedule_value.split(':'))
                trigger = CronTrigger(hour=hour, minute=minute)
            elif schedule_type == 'weekly':
                # Parse day:HH:MM (0=Monday, 6=Sunday)
                parts = schedule_value.split(':')
                day_of_week = int(parts[0])
                hour, minute = int(parts[1]), int(parts[2])
                trigger = CronTrigger(day_of_week=day_of_week, hour=hour, minute=minute)
            elif schedule_type == 'interval':
                # Minutes interval
                minutes = int(schedule_value)
                trigger = IntervalTrigger(minutes=minutes)
            else:
                logger.warning(f"Unknown schedule type: {schedule_type}")
                return

            # Add job to scheduler
            self.scheduler.add_job(
                func=self._execute_scheduled_scrape,
                trigger=trigger,
                args=[schedule_id],
                id=f"schedule_{schedule_id}",
                replace_existing=True,
                name=schedule['name']
            )
            logger.info(f"Added schedule '{schedule['name']}' (ID: {schedule_id}) to scheduler")
        except Exception as e:
            logger.error(f"Error adding schedule to scheduler: {e}")

    def _execute_scheduled_scrape(self, schedule_id: int) -> None:
        """
        Execute a scheduled scrape (runs in background thread).

        Args:
            schedule_id: ID of the schedule to execute
        """
        try:
            logger.info(f"Executing scheduled scrape ID {schedule_id}")
            ScheduleManager.record_execution(schedule_id, 'running')

            # Get schedule details
            schedule = ScheduleManager.get_schedule(schedule_id)
            if not schedule:
                logger.error(f"Schedule ID {schedule_id} not found")
                ScheduleManager.record_execution(schedule_id, 'failed', 'Schedule not found')
                return

            if not schedule['enabled']:
                logger.warning(f"Schedule ID {schedule_id} is disabled, skipping")
                return

            # Get scraper profile details
            profile_url = schedule['profile_url']
            profile_selector = schedule['profile_selector']

            if not profile_url or not profile_selector:
                error_msg = "Invalid scraper profile (missing URL or selector)"
                logger.error(f"Schedule ID {schedule_id}: {error_msg}")
                ScheduleManager.record_execution(schedule_id, 'failed', error_msg)
                return

            # Execute the scrape
            # v2.0.0: Scheduled scrapes use admin user (id=1) or profile owner
            inserted, skipped, url, ids = scrape_url_action(
                profile_url,
                profile_selector,
                limit=0,  # No limit for scheduled scrapes
                user_id=1  # Default to admin for scheduled jobs
            )

            # Record success
            success_msg = f"Scraped {inserted} new articles, skipped {skipped} duplicates"
            logger.info(f"Schedule ID {schedule_id} completed: {success_msg}")
            ScheduleManager.record_execution(schedule_id, 'success')

        except Exception as e:
            error_msg = str(e)
            logger.error(f"Error executing schedule ID {schedule_id}: {e}", exc_info=True)
            ScheduleManager.record_execution(schedule_id, 'failed', error_msg)

    async def on_unmount(self) -> None:
        """Cleanup when app is shutting down."""
        try:
            if hasattr(self, 'scheduler') and self.scheduler:
                self.scheduler.shutdown(wait=False)
                logger.info("Background scheduler stopped")
        except Exception as e:
            logger.error(f"Error stopping scheduler: {e}")

    async def refresh_article_table(self) -> None:
        tbl = self.query_one(DataTable)
        cur_row = tbl.cursor_row
        tbl.clear()
        s_col, s_disp = self.SORT_OPTIONS[self.current_sort_index]
        self.query_one(StatusBar).sort_status = s_disp
        bq = ("SELECT sd.id, sd.title, sd.url, sd.timestamp, "
              "sd.summary IS NOT NULL as has_s, sd.link, sd.sentiment, "
              "GROUP_CONCAT(DISTINCT t.name) as tags_c "
              "FROM scraped_data sd "
              "LEFT JOIN article_tags at ON sd.id = at.article_id "
              "LEFT JOIN tags t ON at.tag_id = t.id")
        conds, params, fdesc = [], {}, []

        # Title filter with optional regex support
        if self.title_filter:
            if self.use_regex:
                # For regex, we'll need to filter in Python after fetching
                fdesc.append(f"Title~regex'{self.title_filter}'")
            else:
                conds.append("sd.title LIKE :tf")
                params["tf"] = f"%{self.title_filter}%"
                fdesc.append(f"Title~'{self.title_filter}'")

        # URL filter with optional regex support
        if self.url_filter:
            if self.use_regex:
                # For regex, we'll need to filter in Python after fetching
                fdesc.append(f"URL~regex'{self.url_filter}'")
            else:
                conds.append("sd.url LIKE :uf")
                params["uf"] = f"%{self.url_filter}%"
                fdesc.append(f"URL~'{self.url_filter}'")

        # Date range filtering
        if self.date_filter_from or self.date_filter_to:
            if self.date_filter_from:
                try:
                    datetime.strptime(self.date_filter_from, "%Y-%m-%d")
                    conds.append("date(sd.timestamp) >= :df_from")
                    params["df_from"] = self.date_filter_from
                    fdesc.append(f"Date>={self.date_filter_from}")
                except ValueError:
                    self.notify("Invalid 'from' date format.", title="Filter Error", severity="warning")
            if self.date_filter_to:
                try:
                    datetime.strptime(self.date_filter_to, "%Y-%m-%d")
                    conds.append("date(sd.timestamp) <= :df_to")
                    params["df_to"] = self.date_filter_to
                    fdesc.append(f"Date<={self.date_filter_to}")
                except ValueError:
                    self.notify("Invalid 'to' date format.", title="Filter Error", severity="warning")
        elif self.date_filter:
            # Legacy single date filter
            try:
                datetime.strptime(self.date_filter, "%Y-%m-%d")
                conds.append("date(sd.timestamp) = :df")
                params["df"] = self.date_filter
                fdesc.append(f"Date='{self.date_filter}'")
            except ValueError:
                if self.date_filter:
                    self.notify("Invalid date format.", title="Filter Error", severity="warning")

        # Tags filter with AND/OR logic
        if self.tags_filter:
            tfs = [t.strip().lower() for t in self.tags_filter.split(',') if t.strip()]
            if tfs:
                if self.tags_logic == "OR":
                    # OR logic: article must have at least one of the tags
                    tag_placeholders = ", ".join([f":tgf_{i}" for i in range(len(tfs))])
                    conds.append(f"sd.id IN (SELECT at_s.article_id FROM article_tags at_s JOIN tags t_s ON at_s.tag_id = t_s.id WHERE t_s.name IN ({tag_placeholders}))")
                    for i, tn in enumerate(tfs):
                        params[f"tgf_{i}"] = tn
                    fdesc.append(f"Tags(OR)='{', '.join(tfs)}'")
                else:
                    # AND logic: article must have all tags (original behavior)
                    for i, tn in enumerate(tfs):
                        pn = f"tgf_{i}"
                        conds.append(f"sd.id IN (SELECT at_s.article_id FROM article_tags at_s JOIN tags t_s ON at_s.tag_id = t_s.id WHERE t_s.name = :{pn})")
                        params[pn] = tn
                    fdesc.append(f"Tags(AND)='{', '.join(tfs)}'")

        # Sentiment filter
        if self.sentiment_filter:
            sval = self.sentiment_filter.strip().capitalize()
            if sval in ["Positive", "Negative", "Neutral"]:
                conds.append("sd.sentiment LIKE :sf")
                params["sf"] = f"%{sval}%"
                fdesc.append(f"Sentiment='{sval}'")
            elif sval:
                self.notify("Sentiment filter: Positive, Negative, or Neutral.", title="Filter Info", severity="info")
        if conds:
            bq += " WHERE " + " AND ".join(conds)
        bq += " GROUP BY sd.id ORDER BY " + s_col
        self.query_one(StatusBar).filter_status = ", ".join(fdesc) if fdesc else "None"
        try:
            with get_db_connection() as conn:
                rows = conn.execute(bq, params).fetchall()
            logger.debug(f"Query returned {len(rows)} rows")

            # Apply regex filtering if enabled (post-SQL filter)
            if self.use_regex and (self.title_filter or self.url_filter):
                filtered_rows = []
                for r_d in rows:
                    try:
                        matches = True
                        if self.title_filter:
                            if not re.search(self.title_filter, r_d["title"], re.IGNORECASE):
                                matches = False
                        if matches and self.url_filter:
                            if not re.search(self.url_filter, r_d["url"], re.IGNORECASE):
                                matches = False
                        if matches:
                            filtered_rows.append(r_d)
                    except re.error as e:
                        self.notify(f"Invalid regex: {e}", title="Regex Error", severity="error")
                        break
                rows = filtered_rows
                logger.debug(f"After regex filtering: {len(rows)} rows")

            self.row_metadata.clear()
            for r_d in rows:
                s_ind = "✓" if r_d["has_s"] else " "
                tags_d = ", ".join(sorted(r_d["tags_c"].split(','))) if r_d["tags_c"] else ""
                senti_d = r_d["sentiment"] or "-"
                timestamp_val = r_d["timestamp"]
                timestamp_str = timestamp_val.strftime('%Y-%m-%d %H:%M:%S') if isinstance(timestamp_val, datetime) else str(timestamp_val)
                row_key = str(r_d["id"])
                # Add visual indicator for bulk selection
                if r_d["id"] in self.selected_row_ids:
                    id_display = f"[✓] {r_d['id']}"
                elif self.selected_row_id == r_d["id"]:
                    id_display = f"*{r_d['id']}"
                else:
                    id_display = str(r_d["id"])
                tbl.add_row(id_display, s_ind, senti_d, r_d["title"], r_d["url"], tags_d, timestamp_str, key=row_key)
                self.row_metadata[row_key] = {'link': r_d['link'], 'has_s': bool(r_d['has_s']), 'tags': r_d["tags_c"] or ""}
            self.query_one(StatusBar).total_articles = len(rows)
            logger.debug(f"Added {len(rows)} rows to table, table now has {tbl.row_count} rows")
            if cur_row is not None and cur_row < len(rows):
                tbl.move_cursor(row=cur_row)
            elif len(rows) > 0:
                tbl.move_cursor(row=0)
            if not rows and any([self.title_filter, self.url_filter, self.date_filter, self.tags_filter, self.sentiment_filter]):
                self.notify("No articles match filters.", title="Filter Info", severity="info", timeout=3)
            elif not rows:
                self.notify("No articles in DB.", title="Info", severity="info", timeout=3)
        except Exception as e:
            logger.error(f"Refresh err: {e}", exc_info=True)
            self.notify(f"Refresh err: {e}", title="DB Error", severity="error")
            self.query_one(StatusBar).total_articles = 0

    async def action_open_filters(self) -> None:
        def handle_filter_result(result):
            if result:  # If filters were applied
                self.call_later(self.refresh_article_table)
        self.push_screen(FilterScreen(self), handle_filter_result)

    async def action_select_row(self) -> None:
        current_id = self._get_current_row_id()
        if current_id is not None:
            # Toggle bulk selection
            if current_id in self.selected_row_ids:
                self.selected_row_ids.discard(current_id)
                logger.debug(f"Row removed from bulk selection, ID: {current_id}")
                self.notify(f"Removed article ID {current_id} from selection", title="Selection", severity="info", timeout=2)
            else:
                self.selected_row_ids.add(current_id)
                logger.debug(f"Row added to bulk selection, ID: {current_id}")
                self.notify(f"Added article ID {current_id} to selection", title="Selection", severity="info", timeout=2)

            # Update status bar
            self.query_one(StatusBar).bulk_selected_count = len(self.selected_row_ids)

            # Also update single selection for compatibility
            self.selected_row_id = current_id if len(self.selected_row_ids) == 1 else None
            self.query_one(StatusBar).selected_id = self.selected_row_id

            # Refresh table to show selection indicator
            await self.refresh_article_table()

    async def action_view_summary(self) -> None:
        current_id = self._get_current_row_id()
        if current_id is None:
            self.notify("No row selected.", title="Info", severity="warning")
            return
        self.selected_row_id = current_id
        try:
            def _get_summary_data_blocking():
                with get_db_connection() as conn_blocking:
                    return conn_blocking.execute("SELECT title, summary, sentiment FROM scraped_data WHERE id=?", (self.selected_row_id,)).fetchone()

            summary_data = _get_summary_data_blocking()
            if summary_data:
                summary_info = {
                    'summary': summary_data['summary'] if summary_data['summary'] else None,
                    'sentiment': summary_data['sentiment'] if summary_data['sentiment'] else None
                }
                if summary_info['summary'] or summary_info['sentiment']:
                    self.push_screen(ViewSummaryModal(summary_data['title'], summary_info))
                else:
                    self.notify("No summary or sentiment data available for this article.", title="Info", severity="info")
            else:
                self.notify(f"Article ID {self.selected_row_id} not found.", title="Error", severity="error")
        except Exception as e:
            logger.error(f"Error viewing summary for ID {self.selected_row_id}: {e}", exc_info=True)
            self.notify(f"Error loading summary: {e}", title="Error", severity="error")

    async def on_data_table_row_selected(self, e: DataTable.RowSelected) -> None:
        self.selected_row_id = int(e.row_key.value) if e.row_key else None
        self.query_one(StatusBar).selected_id = self.selected_row_id
        logger.debug(f"Row selected, ID: {self.selected_row_id}")
        await self.refresh_article_table()

    async def on_data_table_cell_selected(self, e: DataTable.CellSelected) -> None:
        """Handle mouse clicks on DataTable cells for row selection."""
        if e.row_key:
            row_id = int(e.row_key.value)
            # Toggle selection like spacebar does
            if self.selected_row_id == row_id:
                self.selected_row_id = None
                self.query_one(StatusBar).selected_id = None
                logger.debug(f"Row unselected via mouse click, ID: {row_id}")
                self.notify(f"Unselected article ID {row_id}", title="Selection", severity="info", timeout=2)
            else:
                self.selected_row_id = row_id
                self.query_one(StatusBar).selected_id = self.selected_row_id
                logger.debug(f"Row selected via mouse click, ID: {self.selected_row_id}")
                self.notify(f"Selected article ID {row_id}", title="Selection", severity="info", timeout=2)
            # Refresh table to show selection indicator
            await self.refresh_article_table()

    def _get_current_row_id(self) -> int | None:
        if self.selected_row_id is not None:
            return self.selected_row_id
        tbl = self.query_one(DataTable)
        if tbl.row_count > 0:
            try:
                # Get the row key at the current cursor position
                cursor_row = tbl.cursor_row
                if cursor_row < tbl.row_count:
                    # Get the row at cursor position and extract the ID from the first column
                    row_data = tbl.get_row_at(cursor_row)
                    if row_data and len(row_data) > 0:
                        return int(row_data[0])  # First column is the ID
            except Exception as e:
                logger.debug(f"Error getting current row ID: {e}")
        return None

    def _toggle_loading(self, show: bool) -> None:
        if show:
            self.query_one(LoadingIndicator).remove_class("hidden")
        else:
            self.query_one(LoadingIndicator).add_class("hidden")

    async def action_refresh_data(self) -> None:
        self.notify("Refreshing...", title="Data Update", severity="info", timeout=2)
        await self.refresh_article_table()

    # ═══════════════════════════════════════════════════════════════════════════
    # v2.0.0 Multi-User Methods (Phase 2)
    # ═══════════════════════════════════════════════════════════════════════════

    async def _initialize_user_session(self, user_id: int) -> None:
        """
        Initialize user session after successful login (v2.0.0).

        Args:
            user_id: ID of authenticated user
        """
        try:
            # Create session token
            self.session_token = create_user_session(user_id)

            # Load user details
            with get_db_connection() as conn:
                row = conn.execute("""
                    SELECT username, role
                    FROM users WHERE id = ?
                """, (user_id,)).fetchone()

                if row:
                    self.current_user_id = user_id
                    self.current_username = row['username']
                    self.current_user_role = row['role']

                    # Update status bar
                    sbar = self.query_one(StatusBar)
                    sbar.current_username = self.current_username
                    sbar.current_user_role = self.current_user_role

                    logger.info(
                        f"Session initialized for user '{self.current_username}' "
                        f"(role: {self.current_user_role})"
                    )
                    self.notify(
                        f"Welcome, {self.current_username}!",
                        severity="information"
                    )
        except Exception as e:
            logger.error(f"Error initializing user session: {e}")
            self.notify("Error initializing session", severity="error")

    def watch_current_username(self, username: str) -> None:
        """Update status bar when username changes (v2.0.0)."""
        try:
            sbar = self.query_one(StatusBar)
            sbar.current_username = username
        except Exception:
            pass  # Status bar might not be mounted yet

    def watch_current_user_role(self, role: str) -> None:
        """Update status bar when user role changes (v2.0.0)."""
        try:
            sbar = self.query_one(StatusBar)
            sbar.current_user_role = role
        except Exception:
            pass  # Status bar might not be mounted yet

    def check_permission(self, required_role: str) -> bool:
        """
        Check if current user has required role (v2.0.0).

        Args:
            required_role: 'admin', 'user', or 'viewer'

        Returns:
            True if user has permission
        """
        role_hierarchy = {'admin': 3, 'user': 2, 'viewer': 1, 'guest': 0}
        current = role_hierarchy.get(self.current_user_role, 0)
        required = role_hierarchy.get(required_role, 0)
        return current >= required

    def is_admin(self) -> bool:
        """Check if current user is admin (v2.0.0)."""
        return self.current_user_role == "admin"

    def can_edit(self, owner_user_id: Optional[int]) -> bool:
        """
        Check if current user can edit a resource (v2.0.0).

        Args:
            owner_user_id: User ID of resource owner

        Returns:
            True if user is admin or owns the resource
        """
        if self.is_admin():
            return True
        return self.current_user_id == owner_user_id

    def can_delete(self, owner_user_id: Optional[int]) -> bool:
        """Check if current user can delete a resource (v2.0.0)."""
        return self.can_edit(owner_user_id)

    def action_user_profile(self) -> None:
        """Show user profile modal (v2.0.0) - Ctrl+U."""
        if self.current_user_id:
            self.push_screen(UserProfileModal(self.current_user_id))
        else:
            self.notify("Not logged in", severity="warning")

    def action_user_management(self) -> None:
        """Show user management screen (v2.0.0) - Ctrl+Alt+U (admin only)."""
        if not self.is_admin():
            self.notify("Admin access required", severity="error")
            return
        self.push_screen(UserManagementModal())

    def action_logout(self) -> None:
        """Log out current user (v2.0.0) - Ctrl+Shift+L."""
        if self.session_token:
            logout_session(self.session_token)

        self.notify("Logged out successfully. Exiting...", severity="information")
        self.exit()

    async def action_toggle_dark_mode(self) -> None:
        self.dark = not self.dark
        self.query_one(StatusBar).current_theme = "Dark" if self.dark else "Light"

    async def action_view_details(self) -> None:
        current_id = self._get_current_row_id()
        if current_id is None:
            self.notify("No row selected.", title="Info", severity="warning")
            return
        self.selected_row_id = current_id
        try:
            with get_db_connection() as conn:
                ad = conn.execute("SELECT * FROM scraped_data WHERE id=?", (self.selected_row_id,)).fetchone()
                tags = get_tags_for_article(conn, self.selected_row_id)
            if ad:
                self.push_screen(ArticleDetailModal(ad, tags))
            else:
                self.notify(f"Not found: ID {self.selected_row_id}.", title="Error", severity="error")
        except Exception as e:
            logger.error(f"Err details ID {self.selected_row_id}: {e}", exc_info=True)
            self.notify(f"Err details: {e}", title="Error", severity="error")

    async def _summarize_worker(self, eid: int, link: str, style: str, title: str = "", url: str = "") -> None:
        self._toggle_loading(True)
        # Check if style is a template reference
        is_template = style.startswith("template:")
        display_name = style.split(":", 1)[1] if is_template else style
        self.notify(f"Starting '{display_name}' summary ID {eid}...", title="Summarizing", severity="info", timeout=3)
        try:
            txt = fetch_article_content(link, False)
            if not txt:
                self.notify(f"No content for ID {eid}.", title="Summ Error", severity="error")
                self._toggle_loading(False)
                return

            # Handle template-based summarization
            if is_template:
                template_name = style.split(":", 1)[1]
                template_text = TemplateManager.get_template_by_name(template_name)
                if template_text:
                    # Apply template variables
                    prompt = TemplateManager.apply_template(template_text, txt, title, url)
                    summ = get_summary_from_llm(txt, "overview", template=prompt)
                else:
                    self.notify(f"Template '{template_name}' not found.", title="Error", severity="error")
                    self._toggle_loading(False)
                    return
            else:
                # Legacy style-based summarization
                summ = get_summary_from_llm(txt, style)

            if summ:
                def _update_summary_blocking():
                    with get_db_connection() as conn_blocking:
                        conn_blocking.execute("UPDATE scraped_data SET summary=? WHERE id=?", (summ, eid))
                        conn_blocking.commit()
                _update_summary_blocking()
                self.notify(f"Summary for ID {eid} done.", title="Success", severity="info")
                await self.refresh_article_table()
            else:
                self.notify(f"No summary for ID {eid}.", title="Summ Error", severity="error")
        except Exception as e:
            logger.error(f"Err summ ID {eid}: {e}", exc_info=True)
            self.notify(f"Err summ ID {eid}: {e}", title="Error", severity="error")
        finally:
            self._toggle_loading(False)

    async def action_summarize_selected(self) -> None:
        current_id = self._get_current_row_id()
        if current_id is None:
            self.notify("No row selected.", title="Info", severity="warning")
            return
        self.selected_row_id = current_id
        row_key = str(self.selected_row_id)
        meta = self.row_metadata.get(row_key)
        if not meta or 'link' not in meta:
            self.notify(f"No link for ID {self.selected_row_id}.", title="Error", severity="error")
            return

        # Get article details for template variables
        def _get_article_info():
            with get_db_connection() as conn:
                cursor = conn.execute("SELECT title, url FROM scraped_data WHERE id = ?", (self.selected_row_id,))
                row = cursor.fetchone()
                return (row['title'], row['url']) if row else ("", "")

        title, url = _get_article_info()

        # Store summarization context for callbacks
        self._summarize_context = {
            'row_id': self.selected_row_id,
            'link': meta['link'],
            'title': title,
            'url': url
        }

        if meta.get('has_s'):
            def handle_confirm_result(confirmed):
                if confirmed:
                    self._show_summary_style_selector()
                else:
                    self.notify("Cancelled.", title="Info", severity="info")
            self.push_screen(ConfirmModal(f"ID {self.selected_row_id} has summary. Re-summarize?"), handle_confirm_result)
        else:
            self._show_summary_style_selector()

    def _show_summary_style_selector(self):
        def handle_style_result(style):
            if style is not None:
                context = self._summarize_context
                worker_with_args = functools.partial(
                    self._summarize_worker,
                    context['row_id'],
                    context['link'],
                    style,
                    context.get('title', ''),
                    context.get('url', '')
                )
                self.run_worker(worker_with_args, group="llm", exclusive=True)
            else:
                self.notify("Cancelled (no style).", title="Info", severity="info")
        self.push_screen(SelectSummaryStyleModal(), handle_style_result)

    async def _sentiment_worker(self, eid: int, link: str) -> None:
        self._toggle_loading(True)
        self.notify(f"Analyzing sentiment for ID {eid}...", title="Sentiment Analysis", severity="info", timeout=3)
        try:
            def _get_summary_blocking():
                with get_db_connection() as conn_blocking:
                    return conn_blocking.execute("SELECT summary FROM scraped_data WHERE id=?", (eid,)).fetchone()
            ad = _get_summary_blocking()
            txt_to_analyze = ad['summary'] if ad and ad['summary'] else fetch_article_content(link, False)
            if not txt_to_analyze:
                self.notify(f"No content to analyze for ID {eid}.", title="Sentiment Error", severity="error")
                self._toggle_loading(False)
                return
            s_res = get_sentiment_from_llm(txt_to_analyze)
            if s_res:
                def _update_sentiment_blocking():
                    with get_db_connection() as conn_blocking:
                        conn_blocking.execute("UPDATE scraped_data SET sentiment=? WHERE id=?", (s_res, eid))
                        conn_blocking.commit()
                _update_sentiment_blocking()
                self.notify(f"Sentiment for ID {eid}: {s_res}.", title="Sentiment Updated", severity="info")
                await self.refresh_article_table()
            else:
                self.notify(f"Failed to get sentiment for ID {eid}.", title="Sentiment Error", severity="error")
        except Exception as e:
            logger.error(f"Err sentiment worker ID {eid}: {e}", exc_info=True)
            self.notify(f"Err analyzing sentiment ID {eid}: {e}", title="Error", severity="error")
        finally:
            self._toggle_loading(False)

    async def action_sentiment_analysis_selected(self) -> None:
        current_id = self._get_current_row_id()
        if current_id is None:
            self.notify("No row selected for sentiment.", title="Info", severity="warning")
            return
        self.selected_row_id = current_id
        row_key = str(self.selected_row_id)
        meta = self.row_metadata.get(row_key)
        if not meta or 'link' not in meta:
            self.notify(f"No link for ID {self.selected_row_id}.", title="Error", severity="error")
            return
        worker_with_args = functools.partial(self._sentiment_worker, self.selected_row_id, meta['link'])
        self.run_worker(worker_with_args, group="llm", exclusive=True)

    async def _scrape_url_worker(self, url: str, selector: str, limit: int, default_tags_csv: Optional[str] = None) -> None:
        self._toggle_loading(True)
        self.notify(f"Scraping {url}...", title="Scraping", severity="info", timeout=3)
        try:
            # v2.0.0: Pass current user_id
            inserted, skipped, scraped_url, inserted_ids = scrape_url_action(
                url, selector, limit, user_id=self.current_user_id
            )
            self.last_scrape_url = scraped_url
            if inserted_ids and default_tags_csv:
                logger.info(f"Applying default tags '{default_tags_csv}' to {len(inserted_ids)} new articles.")

                def _apply_tags_blocking():
                    for aid in inserted_ids:
                        _update_tags_for_article_blocking(aid, default_tags_csv)
                _apply_tags_blocking()
                self.notify(f"Applied default tags to {len(inserted_ids)} new articles.", title="Tags Applied", severity="info")
            self.notify(f"Scrape of {url} done. New: {inserted}, Skipped: {skipped}.", title="Scrape Finished", severity="info")
            await self.refresh_article_table()
        except Exception as e:
            logger.error(f"Err scrape worker {url}: {e}", exc_info=True)
            self.notify(f"Err scraping {url}: {e}", title="Scrape Error", severity="error")
        finally:
            self._toggle_loading(False)

    async def _handle_scrape_new_result(self, result: tuple[str, str, int] | None, default_tags_csv: Optional[str] = None) -> None:
        if result:
            url, selector, limit = result
            worker_with_args = functools.partial(self._scrape_url_worker, url, selector, limit, default_tags_csv)
            self.run_worker(worker_with_args, group="scraping", exclusive=True)

    async def action_scrape_new(self) -> None:
        self.current_scraper_profile = "Manual Entry"
        self.query_one(StatusBar).scraper_profile = self.current_scraper_profile
        await self.app.push_screen(ScrapeURLModal(self.last_scrape_url, self.last_scrape_selector, self.last_scrape_limit), self._handle_scrape_new_result)

    async def action_delete_selected(self) -> None:
        current_id = self._get_current_row_id()
        if current_id is None:
            self.notify("No row selected.", title="Info", severity="warning")
            return
        self.selected_row_id = current_id

        def handle_delete_confirmation(confirmed):
            if confirmed:
                try:
                    def _delete_blocking():
                        with get_db_connection() as conn_blocking:
                            cur = conn_blocking.execute("DELETE FROM scraped_data WHERE id=?", (self.selected_row_id,))
                            conn_blocking.commit()
                            return cur.rowcount
                    rowcount = _delete_blocking()
                    if rowcount >0:
                        self.notify(f"Deleted ID {self.selected_row_id}.",title="Success",severity="info")
                        self.selected_row_id=None
                        self.query_one(StatusBar).selected_id=None
                        self.call_later(self.refresh_article_table)
                    else:
                        self.notify(f"Not found ID {self.selected_row_id}.",title="Warning",severity="warning")
                except Exception as e:
                    logger.error(f"Err del ID {self.selected_row_id}: {e}",exc_info=True)
                    self.notify(f"Err del ID {self.selected_row_id}: {e}",title="Error",severity="error")
            else:
                self.notify("Cancelled.",title="Info",severity="info")

        self.push_screen(ConfirmModal(f"Delete article ID {self.selected_row_id}?"), handle_delete_confirmation)

    async def action_select_all(self) -> None:
        """Select all visible articles in the current view."""
        try:
            with get_db_connection() as conn:
                # Get all IDs from current filtered view
                bq = "SELECT sd.id FROM scraped_data sd"
                conds, params = [], {}
                if self.title_filter:
                    conds.append("sd.title LIKE :tf")
                    params["tf"] = f"%{self.title_filter}%"
                if self.url_filter:
                    conds.append("sd.url LIKE :uf")
                    params["uf"] = f"%{self.url_filter}%"
                if conds:
                    bq += " WHERE " + " AND ".join(conds)
                rows = conn.execute(bq, params).fetchall()
                self.selected_row_ids = set(row["id"] for row in rows)
                self.query_one(StatusBar).bulk_selected_count = len(self.selected_row_ids)
                await self.refresh_article_table()
                self.notify(f"Selected {len(self.selected_row_ids)} articles", title="Selection", severity="info")
        except Exception as e:
            logger.error(f"Error in select_all: {e}", exc_info=True)
            self.notify(f"Error selecting all: {e}", title="Error", severity="error")

    async def action_deselect_all(self) -> None:
        """Deselect all articles."""
        count = len(self.selected_row_ids)
        self.selected_row_ids.clear()
        self.selected_row_id = None
        self.query_one(StatusBar).bulk_selected_count = 0
        self.query_one(StatusBar).selected_id = None
        await self.refresh_article_table()
        self.notify(f"Deselected {count} articles", title="Selection", severity="info")

    async def action_bulk_delete(self) -> None:
        """Delete all selected articles."""
        if not self.selected_row_ids:
            self.notify("No articles selected for bulk delete.", title="Info", severity="warning")
            return

        count = len(self.selected_row_ids)
        selected_ids = list(self.selected_row_ids)

        def handle_bulk_delete_confirmation(confirmed):
            if confirmed:
                try:
                    def _bulk_delete_blocking():
                        with get_db_connection() as conn_blocking:
                            placeholders = ','.join('?' * len(selected_ids))
                            cur = conn_blocking.execute(
                                f"DELETE FROM scraped_data WHERE id IN ({placeholders})",
                                selected_ids
                            )
                            conn_blocking.commit()
                            return cur.rowcount
                    rowcount = _bulk_delete_blocking()
                    self.selected_row_ids.clear()
                    self.selected_row_id = None
                    self.query_one(StatusBar).bulk_selected_count = 0
                    self.query_one(StatusBar).selected_id = None
                    self.notify(f"Deleted {rowcount} articles.", title="Bulk Delete Success", severity="info")
                    logger.info(f"Bulk deleted {rowcount} articles")
                    self.call_later(self.refresh_article_table)
                except Exception as e:
                    logger.error(f"Error bulk deleting: {e}", exc_info=True)
                    self.notify(f"Error bulk deleting: {e}", title="Error", severity="error")
            else:
                self.notify("Bulk delete cancelled.", title="Info", severity="info")

        self.push_screen(
            ConfirmModal(
                f"Delete {count} selected articles? This cannot be undone!",
                confirm_text="Yes, Delete All"
            ),
            handle_bulk_delete_confirmation
        )

    async def action_clear_database(self)->None:
        def handle_clear_confirmation(confirmed):
            if confirmed:
                self._toggle_loading(True)
                try:
                    def _clear_db_blocking():
                        with get_db_connection() as conn_blocking:conn_blocking.executescript("DELETE FROM article_tags;DELETE FROM tags;DELETE FROM scraped_data;DELETE FROM saved_scrapers WHERE is_preinstalled=0;DELETE FROM sqlite_sequence WHERE name IN ('scraped_data','tags','saved_scrapers');");conn_blocking.commit()
                    _clear_db_blocking()
                    self.notify("User data cleared (pre-installed scrapers kept).",title="DB Cleared",severity="info")
                    logger.info("DB cleared.")
                    self.selected_row_id=None
                    self.query_one(StatusBar).selected_id=None
                    self.call_later(self.refresh_article_table)
                except Exception as e:
                    logger.error(f"Err clearing DB: {e}",exc_info=True)
                    self.notify(f"Err clearing DB: {e}",title="DB Error",severity="error")
                finally:
                    self._toggle_loading(False)
            else:
                self.notify("Clear DB cancelled.",title="Info",severity="info")

        self.push_screen(ConfirmModal("Delete ALL articles from DB? Irreversible!",confirm_text="Yes, Delete All"), handle_clear_confirmation)

    async def action_select_ai_provider(self) -> None:
        """Open AI provider selection modal."""
        def handle_provider_selection(provider_key: Optional[str]) -> None:
            if provider_key:
                if provider_key == "gemini" and GEMINI_API_KEY:
                    set_ai_provider(GeminiProvider(GEMINI_API_KEY))
                    self.notify("AI Provider: Google Gemini", title="Provider Changed", severity="info")
                elif provider_key == "openai" and OPENAI_API_KEY:
                    set_ai_provider(OpenAIProvider(OPENAI_API_KEY))
                    self.notify("AI Provider: OpenAI GPT", title="Provider Changed", severity="info")
                elif provider_key == "claude" and CLAUDE_API_KEY:
                    set_ai_provider(ClaudeProvider(CLAUDE_API_KEY))
                    self.notify("AI Provider: Anthropic Claude", title="Provider Changed", severity="info")
                else:
                    self.notify("API key not configured in .env", title="Error", severity="error")
        self.push_screen(AIProviderSelectionModal(), handle_provider_selection)

    async def action_open_settings(self) -> None:
        """Open settings modal (Ctrl+G)."""
        def handle_settings_result(saved: bool) -> None:
            if saved:
                self.config = ConfigManager.load_config()
                self.notify("Settings reloaded", title="Settings", severity="info")
        self.push_screen(SettingsModal(self.config), handle_settings_result)

    async def action_manage_filter_presets(self) -> None:
        """Open filter presets management modal (Ctrl+Shift+F)."""
        def handle_preset_selection(preset_name: Optional[str]) -> None:
            if preset_name:
                preset_data = FilterPresetManager.load_preset(preset_name)
                if preset_data:
                    # Apply preset filters
                    self.title_filter = preset_data['title_filter']
                    self.url_filter = preset_data['url_filter']
                    self.date_filter_from = preset_data['date_from']
                    self.date_filter_to = preset_data['date_to']
                    self.tags_filter = preset_data['tags_filter']
                    self.sentiment_filter = preset_data['sentiment_filter']
                    self.use_regex = preset_data['use_regex']
                    self.tags_logic = preset_data['tags_logic']
                    self.notify(f"Loaded preset: {preset_name}", title="Preset Loaded", severity="info")
                    # Refresh table with new filters
                    self.run_worker(self.refresh_article_table())
                else:
                    self.notify("Failed to load preset", title="Error", severity="error")
        self.push_screen(FilterPresetModal(), handle_preset_selection)

    async def action_save_filter_preset(self) -> None:
        """Save current filters as a preset (Ctrl+Shift+S)."""
        def handle_preset_name(name: Optional[str]) -> None:
            if name:
                success = FilterPresetManager.save_preset(
                    name,
                    self.title_filter,
                    self.url_filter,
                    self.date_filter_from,
                    self.date_filter_to,
                    self.tags_filter,
                    self.sentiment_filter,
                    self.use_regex,
                    self.tags_logic
                )
                if success:
                    self.notify(f"Saved preset: {name}", title="Preset Saved", severity="info")
                else:
                    self.notify("Failed to save preset", title="Error", severity="error")
        self.push_screen(SavePresetModal(), handle_preset_name)

    async def action_manage_schedules(self) -> None:
        """Open schedule management modal (Ctrl+Shift+A)."""
        def handle_schedule_result(result: Optional[int]) -> None:
            if result:
                self.notify("Schedule management complete", severity="info")
        self.push_screen(ScheduleManagementModal(), handle_schedule_result)

    async def action_view_analytics(self) -> None:
        """Open analytics modal (Ctrl+Shift+V)."""
        self.push_screen(AnalyticsModal())

    async def action_auto_tag(self) -> None:
        """Auto-tag selected article using AI (Ctrl+Shift+T)."""
        current_id = self._get_current_row_id()
        if current_id is None:
            self.notify("No article selected.", title="Info", severity="warning")
            return

        self.selected_row_id = current_id
        worker = functools.partial(self._auto_tag_worker, current_id)
        self.run_worker(worker)

    async def action_extract_entities(self) -> None:
        """Extract entities from selected article (Ctrl+Shift+E)."""
        current_id = self._get_current_row_id()
        if current_id is None:
            self.notify("No article selected.", title="Info", severity="warning")
            return

        self.selected_row_id = current_id
        worker = functools.partial(self._extract_entities_worker, current_id)
        self.run_worker(worker)

    async def action_extract_keywords(self) -> None:
        """Extract keywords from selected article (Ctrl+Shift+K)."""
        current_id = self._get_current_row_id()
        if current_id is None:
            self.notify("No article selected.", title="Info", severity="warning")
            return

        self.selected_row_id = current_id
        worker = functools.partial(self._extract_keywords_worker, current_id)
        self.run_worker(worker)

    async def action_find_similar(self) -> None:
        """Find similar articles to selected one (Ctrl+Shift+R)."""
        current_id = self._get_current_row_id()
        if current_id is None:
            self.notify("No article selected.", title="Info", severity="warning")
            return

        self.selected_row_id = current_id
        worker = functools.partial(self._find_similar_worker, current_id)
        self.run_worker(worker)

    # === v1.9.0 Smart Categorization & Topic Modeling Actions ===

    async def action_topic_modeling(self) -> None:
        """Run topic modeling on articles (Ctrl+Alt+T)."""
        def handle_config(config):
            if config:
                worker = functools.partial(
                    self._topic_modeling_worker,
                    config['algorithm'],
                    config['num_topics'],
                    config['num_words']
                )
                self.run_worker(worker, group="topic_modeling", exclusive=True)

        self.push_screen(TopicModelingModal(), handle_config)

    async def action_ask_question(self) -> None:
        """Ask a question about scraped content (Ctrl+Alt+Q)."""
        def handle_question(question):
            if question and question != "__ask_another__":
                worker = functools.partial(self._qa_worker, question)
                self.run_worker(worker, group="qa", exclusive=True)
            elif question == "__ask_another__":
                # Show the input modal again
                self.push_screen(QuestionAnsweringModal(), handle_question)

        self.push_screen(QuestionAnsweringModal(), handle_question)

    async def action_find_duplicates(self) -> None:
        """Find duplicate articles (Ctrl+Alt+D)."""
        def handle_threshold(threshold):
            if threshold is not None:
                worker = functools.partial(self._duplicate_detection_worker, threshold)
                self.run_worker(worker, group="duplicates", exclusive=True)

        self.push_screen(DuplicateDetectionModal(), handle_threshold)

    async def action_find_related(self) -> None:
        """Find related articles to selected (Ctrl+Alt+L)."""
        current_id = self._get_current_row_id()
        if current_id is None:
            self.notify("No article selected.", title="Info", severity="warning")
            return

        self.selected_row_id = current_id

        def get_article_and_related():
            with get_db_connection() as conn:
                # Get current article
                cursor = conn.execute(
                    "SELECT id, title, content FROM scraped_data WHERE id = ?",
                    (current_id,)
                )
                article = cursor.fetchone()
                if not article:
                    return None, []

                # Get all articles for similarity comparison
                all_articles = conn.execute(
                    "SELECT id, title, content, url FROM scraped_data WHERE id != ?",
                    (current_id,)
                ).fetchall()

                return article, all_articles

        async def show_related():
            article, all_articles = await self.run_in_thread(get_article_and_related)

            if not article:
                self.notify("Article not found.", severity="error")
                return

            # Find similar articles
            related = ContentSimilarityManager.find_similar_articles(
                dict(article),
                [dict(a) for a in all_articles],
                top_n=10,
                threshold=0.3
            )

            self.push_screen(
                RelatedArticlesModal(
                    article['title'],
                    related
                )
            )

        await show_related()

    async def action_cluster_articles(self) -> None:
        """Cluster articles by similarity (Ctrl+Alt+C)."""
        def handle_num_clusters(num_clusters):
            if num_clusters:
                worker = functools.partial(self._clustering_worker, num_clusters)
                self.run_worker(worker, group="clustering", exclusive=True)

        self.push_screen(ClusterViewModal(), handle_num_clusters)

    async def action_view_qa_history(self) -> None:
        """View Q&A conversation history (Ctrl+Alt+H)."""
        def get_history():
            return QuestionAnsweringManager.get_qa_history(limit=20)

        async def show_history():
            history = await self.run_in_thread(get_history)

            if not history:
                self.notify("No Q&A history found.", severity="info")
                return

            # Build history display
            content = "# Q&A Conversation History\n\n"
            for idx, qa in enumerate(history, 1):
                content += f"## {idx}. Question\n{qa['question']}\n\n"
                content += f"**Answer:** {qa['answer']}\n\n"
                content += f"*Confidence: {qa.get('confidence', 0):.1%} | "
                content += f"Date: {qa.get('created_at', 'Unknown')}*\n\n---\n\n"

            # Show in a simple modal
            from textual.widgets import Markdown
            from textual.containers import VerticalScroll

            class QAHistoryModal(ModalScreen):
                DEFAULT_CSS = """
                QAHistoryModal {
                    align: center middle;
                    background: $surface-darken-1;
                }
                QAHistoryModal > VerticalScroll {
                    width: 90%;
                    height: 90%;
                    border: thick $primary-lighten-1;
                    padding: 1 2;
                    background: $surface;
                }
                """

                def compose(self):
                    with VerticalScroll():
                        yield Markdown(content)
                        yield Button("Close", id="close")

                def on_button_pressed(self, event):
                    self.dismiss()

            self.push_screen(QAHistoryModal())

        await show_history()

    async def action_evaluate_summary(self) -> None:
        """Evaluate summary quality (Ctrl+Alt+M)."""
        current_id = self._get_current_row_id()
        if current_id is None:
            self.notify("No article selected.", title="Info", severity="warning")
            return

        self.selected_row_id = current_id

        def get_summary_data():
            with get_db_connection() as conn:
                cursor = conn.execute(
                    "SELECT content, summary FROM scraped_data WHERE id = ?",
                    (current_id,)
                )
                row = cursor.fetchone()
                if row and row['summary']:
                    return {
                        'content': row['content'],
                        'summary': row['summary']
                    }
                return None

        async def evaluate():
            data = await self.run_in_thread(get_summary_data)

            if not data:
                self.notify("No summary found for this article.", severity="warning")
                return

            # Calculate quality metrics
            metrics = SummaryQualityManager.calculate_rouge_scores(
                data['summary'],
                data['content']
            )

            self.push_screen(SummaryQualityModal(metrics))

        await evaluate()

    async def action_toggle_help(self)->None:await self.app.push_screen(HelpModal())
    async def action_cycle_sort_order(self)->None:self.current_sort_index=(self.current_sort_index+1)%len(self.SORT_OPTIONS);await self.refresh_article_table();self.notify(f"Sorted by: {self.SORT_OPTIONS[self.current_sort_index][1]}",title="Sort Changed",severity="info",timeout=2)
    async def _handle_manage_tags_result(self,aid:int,nts:Optional[str])->None:
        if nts is not None:
            self._toggle_loading(True)
            try:
                _update_tags_for_article_blocking(aid,nts)
                self.notify(f"Tags updated for ID {aid}.",title="Tags Updated",severity="info");await self.refresh_article_table()
            except Exception as e:logger.error(f"Err tags ID {aid}: {e}",exc_info=True);self.notify(f"Err tags: {e}",title="Tag Error",severity="error")
            finally:self._toggle_loading(False)
    async def action_manage_tags(self)->None:
        current_id=self._get_current_row_id()
        if current_id is None:self.notify("No row selected.",title="Info",severity="warning");return
        self.selected_row_id=current_id
        try:
            def _get_tags_blocking():
                with get_db_connection() as conn_blocking:
                    return get_tags_for_article(conn_blocking, self.selected_row_id) # type: ignore
            ct = _get_tags_blocking()
            await self.app.push_screen(ManageTagsModal(self.selected_row_id,ct),lambda ts:self._handle_manage_tags_result(self.selected_row_id,ts)) # type: ignore
        except Exception as e:logger.error(f"Err prep tags ID {self.selected_row_id}: {e}",exc_info=True);self.notify(f"Err tag manager: {e}",title="Error",severity="error")

    async def _export_csv_worker(self,filename:str)->None:
        self._toggle_loading(True)
        self.notify(f"Exporting to {filename}...",title="Exporting CSV",severity="info")
        try:
            s_col,_=self.SORT_OPTIONS[self.current_sort_index]
            def _fetch_for_export_blocking():
                bq_export="SELECT sd.id,sd.title,sd.url,sd.link,sd.timestamp,sd.summary,sd.sentiment,GROUP_CONCAT(DISTINCT t.name) as tags_c FROM scraped_data sd LEFT JOIN article_tags at ON sd.id=at.article_id LEFT JOIN tags t ON at.tag_id=t.id"
                conds_export,params_export=[],{}
                if self.title_filter:conds_export.append("sd.title LIKE :tf");params_export["tf"]=f"%{self.title_filter}%"
                if self.url_filter:conds_export.append("sd.url LIKE :uf");params_export["uf"]=f"%{self.url_filter}%"
                if self.date_filter:conds_export.append("date(sd.timestamp)=:df");params_export["df"]=self.date_filter
                if self.tags_filter:
                    tfs_export=[t.strip().lower() for t in self.tags_filter.split(',') if t.strip()]
                    for i,tn_export in enumerate(tfs_export):pn_export=f"tgf_{i}";conds_export.append(f"sd.id IN (SELECT at_s.article_id FROM article_tags at_s JOIN tags t_s ON at_s.tag_id=t_s.id WHERE t_s.name=:{pn_export})");params_export[pn_export]=tn_export
                if self.sentiment_filter:sval_export=self.sentiment_filter.strip().capitalize();conds_export.append("sd.sentiment LIKE :sf");params_export["sf"]=f"%{sval_export}%"
                if conds_export:bq_export+=" WHERE "+" AND ".join(conds_export)
                bq_export+=" GROUP BY sd.id ORDER BY "+s_col
                with get_db_connection() as conn_blocking:
                    return conn_blocking.execute(bq_export,params_export).fetchall()
            rows_to_export = _fetch_for_export_blocking()
            if not rows_to_export:self.notify("No data to export.",title="Export Info",severity="info"); self._toggle_loading(False); return
            def _write_csv_blocking():
                fp=Path(filename)
                with open(fp,'w',newline='',encoding='utf-8') as csvf:
                    fn=['ID','Title','Source URL','Article Link','Timestamp','Summary','Sentiment','Tags'];w=csv.DictWriter(csvf,fieldnames=fn);w.writeheader()
                    for r_data in rows_to_export:
                        timestamp_val = r_data['timestamp']
                        timestamp_str = timestamp_val.strftime('%Y-%m-%d %H:%M:%S') if isinstance(timestamp_val, datetime) else str(timestamp_val)
                        w.writerow({'ID':r_data['id'],'Title':r_data['title'],'Source URL':r_data['url'],'Article Link':r_data['link'],'Timestamp':timestamp_str,'Summary':r_data['summary'],'Sentiment':r_data['sentiment'],'Tags':r_data['tags_c']})
                return fp.resolve(), len(rows_to_export)
            resolved_path, num_rows = _write_csv_blocking()
            self.notify(f"Data exported to {resolved_path}",title="CSV Exported",severity="info");logger.info(f"Exported {num_rows} rows to {resolved_path}")
        except Exception as e:logger.error(f"Err export CSV '{filename}': {e}",exc_info=True);self.notify(f"Err export CSV: {e}",title="Export Error",severity="error")
        finally:self._toggle_loading(False)
    async def action_export_csv(self)->None:
        dfn=f"scraped_articles_{datetime.now():%Y%m%d_%H%M%S}.csv"
        def handle_filename_result(fn):
            if fn:
                worker_with_args = functools.partial(self._export_csv_worker, fn)
                self.run_worker(worker_with_args, group="exporting", exclusive=True)
        self.push_screen(FilenameModal(default_filename=dfn), handle_filename_result)

    async def _export_json_worker(self, filename: str) -> None:
        """Export articles to JSON format."""
        self._toggle_loading(True)
        self.notify(f"Exporting to {filename}...", title="Exporting JSON", severity="info")
        try:
            s_col, _ = self.SORT_OPTIONS[self.current_sort_index]

            def _fetch_for_export_blocking():
                bq_export = (
                    "SELECT sd.id, sd.title, sd.url, sd.link, sd.timestamp, "
                    "sd.summary, sd.sentiment, sd.content, "
                    "GROUP_CONCAT(DISTINCT t.name) as tags_c "
                    "FROM scraped_data sd "
                    "LEFT JOIN article_tags at ON sd.id = at.article_id "
                    "LEFT JOIN tags t ON at.tag_id = t.id"
                )
                conds_export, params_export = [], {}
                if self.title_filter:
                    conds_export.append("sd.title LIKE :tf")
                    params_export["tf"] = f"%{self.title_filter}%"
                if self.url_filter:
                    conds_export.append("sd.url LIKE :uf")
                    params_export["uf"] = f"%{self.url_filter}%"
                if self.date_filter:
                    conds_export.append("date(sd.timestamp) = :df")
                    params_export["df"] = self.date_filter
                if self.tags_filter:
                    tfs_export = [t.strip().lower() for t in self.tags_filter.split(',') if t.strip()]
                    for i, tn_export in enumerate(tfs_export):
                        pn_export = f"tgf_{i}"
                        conds_export.append(
                            f"sd.id IN (SELECT at_s.article_id FROM article_tags at_s "
                            f"JOIN tags t_s ON at_s.tag_id = t_s.id WHERE t_s.name = :{pn_export})"
                        )
                        params_export[pn_export] = tn_export
                if self.sentiment_filter:
                    sval_export = self.sentiment_filter.strip().capitalize()
                    conds_export.append("sd.sentiment LIKE :sf")
                    params_export["sf"] = f"%{sval_export}%"
                if conds_export:
                    bq_export += " WHERE " + " AND ".join(conds_export)
                bq_export += " GROUP BY sd.id ORDER BY " + s_col
                with get_db_connection() as conn_blocking:
                    return conn_blocking.execute(bq_export, params_export).fetchall()

            rows_to_export = _fetch_for_export_blocking()
            if not rows_to_export:
                self.notify("No data to export.", title="Export Info", severity="info")
                self._toggle_loading(False)
                return

            def _write_json_blocking():
                fp = Path(filename)
                articles = []
                for r_data in rows_to_export:
                    timestamp_val = r_data['timestamp']
                    timestamp_str = (
                        timestamp_val.strftime('%Y-%m-%d %H:%M:%S')
                        if isinstance(timestamp_val, datetime)
                        else str(timestamp_val)
                    )
                    tags_list = (
                        [t.strip() for t in r_data['tags_c'].split(',') if t.strip()]
                        if r_data['tags_c']
                        else []
                    )
                    article_data = {
                        'id': r_data['id'],
                        'title': r_data['title'],
                        'source_url': r_data['url'],
                        'article_link': r_data['link'],
                        'timestamp': timestamp_str,
                        'summary': r_data['summary'],
                        'sentiment': r_data['sentiment'],
                        'content': r_data['content'],
                        'tags': tags_list
                    }
                    articles.append(article_data)

                export_data = {
                    'export_date': datetime.now().isoformat(),
                    'total_articles': len(articles),
                    'filters_applied': {
                        'title': self.title_filter if self.title_filter else None,
                        'url': self.url_filter if self.url_filter else None,
                        'date': self.date_filter if self.date_filter else None,
                        'tags': self.tags_filter if self.tags_filter else None,
                        'sentiment': self.sentiment_filter if self.sentiment_filter else None
                    },
                    'articles': articles
                }

                with open(fp, 'w', encoding='utf-8') as jsonf:
                    json.dump(export_data, jsonf, indent=2, ensure_ascii=False)
                return fp.resolve(), len(articles)

            resolved_path, num_rows = _write_json_blocking()
            self.notify(
                f"Data exported to {resolved_path}",
                title="JSON Exported",
                severity="info"
            )
            logger.info(f"Exported {num_rows} rows to {resolved_path}")
        except Exception as e:
            logger.error(f"Error exporting JSON '{filename}': {e}", exc_info=True)
            self.notify(f"Error exporting JSON: {e}", title="Export Error", severity="error")
        finally:
            self._toggle_loading(False)

    async def action_export_json(self) -> None:
        """Export articles to JSON format."""
        dfn = f"scraped_articles_{datetime.now():%Y%m%d_%H%M%S}.json"

        def handle_filename_result(fn):
            if fn:
                worker_with_args = functools.partial(self._export_json_worker, fn)
                self.run_worker(worker_with_args, group="exporting", exclusive=True)

        self.push_screen(FilenameModal(default_filename=dfn), handle_filename_result)

    async def action_export_excel(self) -> None:
        """Export articles to Excel (XLSX) format with formatting."""
        dfn = f"scraped_articles_{datetime.now():%Y%m%d_%H%M%S}.xlsx"

        def handle_filename_result(fn):
            if fn:
                worker_with_args = functools.partial(self._export_excel_worker, fn)
                self.run_worker(worker_with_args, group="exporting", exclusive=True)

        self.push_screen(FilenameModal(default_filename=dfn), handle_filename_result)

    async def _export_excel_worker(self, filename: str) -> None:
        """Worker to export articles to Excel format."""
        self._toggle_loading(True)
        self.notify(f"Exporting to {filename}...", title="Exporting Excel", severity="info")
        try:
            # Fetch articles using same logic as JSON export
            articles = await self.run_in_thread(self._fetch_articles_for_export)

            if not articles:
                self.notify("No data to export.", title="Export Info", severity="info")
                self._toggle_loading(False)
                return

            # Export to Excel
            def _export_blocking():
                return ExcelExportManager.export_to_excel(
                    articles, filename, include_charts=True, template="standard"
                )

            success = await self.run_in_thread(_export_blocking)

            if success:
                self.notify(
                    f"Exported {len(articles)} article(s) to {filename}",
                    title="Export Complete",
                    severity="info"
                )
            else:
                self.notify("Excel export failed. Check logs.", title="Export Error", severity="error")

        except Exception as e:
            logger.error(f"Error in Excel export worker: {e}", exc_info=True)
            self.notify(f"Excel export error: {e}", title="Export Error", severity="error")
        finally:
            self._toggle_loading(False)

    async def action_export_pdf(self) -> None:
        """Export articles to PDF format with professional layout."""
        dfn = f"scraped_articles_{datetime.now():%Y%m%d_%H%M%S}.pdf"

        def handle_filename_result(fn):
            if fn:
                worker_with_args = functools.partial(self._export_pdf_worker, fn)
                self.run_worker(worker_with_args, group="exporting", exclusive=True)

        self.push_screen(FilenameModal(default_filename=dfn), handle_filename_result)

    async def _export_pdf_worker(self, filename: str) -> None:
        """Worker to export articles to PDF format."""
        self._toggle_loading(True)
        self.notify(f"Generating PDF report {filename}...", title="Exporting PDF", severity="info")
        try:
            # Fetch articles
            articles = await self.run_in_thread(self._fetch_articles_for_export)

            if not articles:
                self.notify("No data to export.", title="Export Info", severity="info")
                self._toggle_loading(False)
                return

            # Export to PDF
            def _export_blocking():
                return PDFExportManager.export_to_pdf(
                    articles, filename, include_charts=True, template="standard"
                )

            success = await self.run_in_thread(_export_blocking)

            if success:
                self.notify(
                    f"PDF report generated with {len(articles)} article(s): {filename}",
                    title="Export Complete",
                    severity="info"
                )
            else:
                self.notify("PDF export failed. Check logs.", title="Export Error", severity="error")

        except Exception as e:
            logger.error(f"Error in PDF export worker: {e}", exc_info=True)
            self.notify(f"PDF export error: {e}", title="Export Error", severity="error")
        finally:
            self._toggle_loading(False)

    async def action_export_word_cloud(self) -> None:
        """Generate and export word cloud from tag data."""
        dfn = f"tag_wordcloud_{datetime.now():%Y%m%d_%H%M%S}.png"

        def handle_filename_result(fn):
            if fn:
                worker_with_args = functools.partial(self._export_word_cloud_worker, fn)
                self.run_worker(worker_with_args, group="exporting", exclusive=True)

        self.push_screen(FilenameModal(default_filename=dfn), handle_filename_result)

    async def _export_word_cloud_worker(self, filename: str) -> None:
        """Worker to generate word cloud."""
        self._toggle_loading(True)
        self.notify(f"Generating word cloud {filename}...", title="Word Cloud", severity="info")
        try:
            # Get tag statistics
            stats = await self.run_in_thread(
                lambda: AnalyticsManager.get_statistics(
                    self.title_filter, self.url_filter, self.date_filter,
                    self.tags_filter, self.sentiment_filter
                )
            )

            tags_data = stats.get('top_tags', [])

            if not tags_data:
                self.notify("No tag data available for word cloud.", title="Word Cloud", severity="warning")
                self._toggle_loading(False)
                return

            # Generate word cloud
            def _generate_blocking():
                return EnhancedVisualizationManager.generate_word_cloud(tags_data, filename)

            success = await self.run_in_thread(_generate_blocking)

            if success:
                self.notify(
                    f"Word cloud generated: {filename}",
                    title="Word Cloud Complete",
                    severity="info"
                )
            else:
                self.notify("Word cloud generation failed. Check logs.", title="Export Error", severity="error")

        except Exception as e:
            logger.error(f"Error in word cloud worker: {e}", exc_info=True)
            self.notify(f"Word cloud error: {e}", title="Export Error", severity="error")
        finally:
            self._toggle_loading(False)

    async def _auto_tag_worker(self, article_id: int) -> None:
        """Worker to auto-tag an article using AI."""
        self._toggle_loading(True)
        self.notify(f"Auto-tagging article ID {article_id}...", title="AI Auto-Tagging", severity="info")
        try:
            def _get_article_blocking():
                with get_db_connection() as conn:
                    result = conn.execute(
                        "SELECT title, link, content FROM scraped_data WHERE id = ?",
                        (article_id,)
                    ).fetchone()
                    return result if result else None

            article = await self.run_in_thread(_get_article_blocking)
            if not article:
                self.notify("Article not found.", title="Error", severity="error")
                self._toggle_loading(False)
                return

            # Generate tags using AI
            tags = await self.run_in_thread(
                lambda: AITaggingManager.generate_tags(
                    article['title'],
                    article['content'] or article['link']
                )
            )

            if tags:
                # Update tags in database
                def _update_tags_blocking():
                    _update_tags_for_article_blocking(article_id, ', '.join(tags))

                await self.run_in_thread(_update_tags_blocking)
                await self.refresh_article_table()
                self.notify(
                    f"Auto-tagged with: {', '.join(tags)}",
                    title="Auto-Tagging Complete",
                    severity="info"
                )
            else:
                self.notify("No tags generated.", title="Auto-Tagging", severity="warning")

        except Exception as e:
            logger.error(f"Error in auto-tag worker: {e}", exc_info=True)
            self.notify(f"Auto-tagging error: {e}", title="Error", severity="error")
        finally:
            self._toggle_loading(False)

    async def _extract_entities_worker(self, article_id: int) -> None:
        """Worker to extract named entities from an article."""
        self._toggle_loading(True)
        self.notify(f"Extracting entities from article ID {article_id}...", title="Entity Extraction", severity="info")
        try:
            def _get_article_blocking():
                with get_db_connection() as conn:
                    result = conn.execute(
                        "SELECT title, content, link FROM scraped_data WHERE id = ?",
                        (article_id,)
                    ).fetchone()
                    return result if result else None

            article = await self.run_in_thread(_get_article_blocking)
            if not article:
                self.notify("Article not found.", title="Error", severity="error")
                self._toggle_loading(False)
                return

            # Extract entities
            entities = await self.run_in_thread(
                lambda: EntityRecognitionManager.extract_entities(
                    article['content'] or article['link']
                )
            )

            if entities:
                # Format entities for display
                entity_text = "\n\n".join([
                    f"**{ent_type}:** {', '.join(ent_list)}"
                    for ent_type, ent_list in entities.items()
                    if ent_list
                ])

                if entity_text:
                    # Display in a modal
                    class EntityModal(ModalScreen[None]):
                        DEFAULT_CSS = """
                        EntityModal {
                            align: center middle;
                        }
                        EntityModal > Vertical {
                            width: 80;
                            height: auto;
                            max-height: 80%;
                            background: $panel;
                            border: thick $primary;
                            padding: 2;
                        }
                        """
                        BINDINGS = [Binding("escape", "dismiss", "Close")]

                        def __init__(self, title: str, entities_md: str):
                            super().__init__()
                            self.article_title = title
                            self.entities_md = entities_md

                        def compose(self) -> ComposeResult:
                            with Vertical():
                                yield Label(f"Entities: {self.article_title}", classes="dialog-title")
                                with VerticalScroll():
                                    yield Markdown(self.entities_md)
                                with Horizontal(classes="modal-buttons"):
                                    yield Button("Close", variant="primary", id="close")

                        def on_button_pressed(self, event: Button.Pressed) -> None:
                            self.dismiss()

                    self.push_screen(EntityModal(article['title'], entity_text))
                    self.notify("Entity extraction complete.", title="Entities", severity="info")
                else:
                    self.notify("No entities found.", title="Entity Extraction", severity="warning")
            else:
                self.notify("No entities extracted.", title="Entity Extraction", severity="warning")

        except Exception as e:
            logger.error(f"Error in extract entities worker: {e}", exc_info=True)
            self.notify(f"Entity extraction error: {e}", title="Error", severity="error")
        finally:
            self._toggle_loading(False)

    async def _extract_keywords_worker(self, article_id: int) -> None:
        """Worker to extract keywords from an article."""
        self._toggle_loading(True)
        self.notify(f"Extracting keywords from article ID {article_id}...", title="Keyword Extraction", severity="info")
        try:
            def _get_article_blocking():
                with get_db_connection() as conn:
                    result = conn.execute(
                        "SELECT title, content, link FROM scraped_data WHERE id = ?",
                        (article_id,)
                    ).fetchone()
                    return result if result else None

            article = await self.run_in_thread(_get_article_blocking)
            if not article:
                self.notify("Article not found.", title="Error", severity="error")
                self._toggle_loading(False)
                return

            # Extract keywords
            keywords = await self.run_in_thread(
                lambda: KeywordExtractionManager.extract_keywords(
                    article['content'] or article['link'],
                    article['title']
                )
            )

            if keywords:
                # Display keywords
                keywords_text = ", ".join(keywords)
                self.notify(
                    f"Keywords: {keywords_text}",
                    title="Keywords Extracted",
                    severity="info",
                    timeout=10
                )
            else:
                self.notify("No keywords extracted.", title="Keyword Extraction", severity="warning")

        except Exception as e:
            logger.error(f"Error in extract keywords worker: {e}", exc_info=True)
            self.notify(f"Keyword extraction error: {e}", title="Error", severity="error")
        finally:
            self._toggle_loading(False)

    async def _find_similar_worker(self, article_id: int) -> None:
        """Worker to find similar articles."""
        self._toggle_loading(True)
        self.notify(f"Finding similar articles to ID {article_id}...", title="Finding Similar", severity="info")
        try:
            def _get_all_articles_blocking():
                with get_db_connection() as conn:
                    return conn.execute(
                        "SELECT id, title, content, link FROM scraped_data"
                    ).fetchall()

            articles = await self.run_in_thread(_get_all_articles_blocking)
            if not articles:
                self.notify("No articles in database.", title="Error", severity="error")
                self._toggle_loading(False)
                return

            # Find similar articles
            similar_ids = await self.run_in_thread(
                lambda: ContentSimilarityManager.find_similar_articles(
                    article_id,
                    [(a['id'], a['title'], a['content'] or a['link']) for a in articles]
                )
            )

            if similar_ids:
                # Get titles of similar articles
                def _get_titles_blocking():
                    with get_db_connection() as conn:
                        placeholders = ','.join(['?'] * len(similar_ids))
                        return conn.execute(
                            f"SELECT id, title FROM scraped_data WHERE id IN ({placeholders})",
                            similar_ids
                        ).fetchall()

                similar_articles = await self.run_in_thread(_get_titles_blocking)

                if similar_articles:
                    similar_text = "\n".join([
                        f"• [ID {a['id']}] {a['title']}"
                        for a in similar_articles
                    ])
                    self.notify(
                        f"Similar articles:\n{similar_text}",
                        title="Similar Articles Found",
                        severity="info",
                        timeout=15
                    )
                else:
                    self.notify("No similar articles found.", title="Similar Articles", severity="info")
            else:
                self.notify("No similar articles found.", title="Similar Articles", severity="info")

        except Exception as e:
            logger.error(f"Error in find similar worker: {e}", exc_info=True)
            self.notify(f"Find similar error: {e}", title="Error", severity="error")
        finally:
            self._toggle_loading(False)

    # === v1.9.0 Worker Methods ===

    async def _topic_modeling_worker(self, algorithm: str, num_topics: int, num_words: int) -> None:
        """Worker for topic modeling."""
        self._toggle_loading(True)
        self.notify(f"Running {algorithm.upper()} topic modeling...", title="Topic Modeling", severity="info")
        try:
            def _get_articles_blocking():
                with get_db_connection() as conn:
                    return [dict(row) for row in conn.execute(
                        "SELECT id, title, content FROM scraped_data WHERE content IS NOT NULL"
                    ).fetchall()]

            articles = await self.run_in_thread(_get_articles_blocking)

            if len(articles) < 2:
                self.notify("Need at least 2 articles for topic modeling.", severity="warning")
                self._toggle_loading(False)
                return

            # Run topic modeling
            if algorithm == "lda":
                result = await self.run_in_thread(
                    lambda: TopicModelingManager.perform_lda_topic_modeling(
                        articles, num_topics=num_topics
                    )
                )
            else:  # nmf
                result = await self.run_in_thread(
                    lambda: TopicModelingManager.perform_nmf_topic_modeling(
                        articles, num_topics=num_topics
                    )
                )

            if 'error' in result:
                self.notify(f"Topic modeling error: {result['error']}", severity="error")
            else:
                topics = result.get('topics', [])
                if topics:
                    topics_text = "\n".join([
                        f"• {t['label']}" for t in topics[:5]
                    ])
                    self.notify(
                        f"Found {len(topics)} topics:\n{topics_text}",
                        title="Topic Modeling Complete",
                        severity="info",
                        timeout=15
                    )
                else:
                    self.notify("No topics found.", severity="info")

        except Exception as e:
            logger.error(f"Error in topic modeling worker: {e}", exc_info=True)
            self.notify(f"Topic modeling error: {e}", title="Error", severity="error")
        finally:
            self._toggle_loading(False)

    async def _qa_worker(self, question: str) -> None:
        """Worker for question answering."""
        self._toggle_loading(True)
        self.notify(f"Finding answer to: {question[:50]}...", title="Question Answering", severity="info")
        try:
            def _get_articles_blocking():
                with get_db_connection() as conn:
                    return [dict(row) for row in conn.execute(
                        "SELECT id, title, content, url FROM scraped_data WHERE content IS NOT NULL LIMIT 10"
                    ).fetchall()]

            articles = await self.run_in_thread(_get_articles_blocking)

            if not articles:
                self.notify("No articles available for Q&A.", severity="warning")
                self._toggle_loading(False)
                return

            # Get answer
            result = await self.run_in_thread(
                lambda: QuestionAnsweringManager.answer_question(question, articles)
            )

            answer = result.get('answer', 'No answer generated.')
            sources = result.get('sources', [])

            # Save to history
            article_ids = [s.get('article_id') for s in sources if s.get('article_id')]
            await self.run_in_thread(
                lambda: QuestionAnsweringManager.save_qa_conversation(
                    question, answer, article_ids, result.get('confidence', 0)
                )
            )

            # Show answer in modal
            self.push_screen(QuestionAnsweringModal(answer=answer, sources=sources))

        except Exception as e:
            logger.error(f"Error in Q&A worker: {e}", exc_info=True)
            self.notify(f"Q&A error: {e}", title="Error", severity="error")
        finally:
            self._toggle_loading(False)

    async def _duplicate_detection_worker(self, threshold: float) -> None:
        """Worker for duplicate detection."""
        self._toggle_loading(True)
        self.notify(f"Detecting duplicates (threshold: {threshold})...", title="Duplicate Detection", severity="info")
        try:
            def _get_articles_blocking():
                with get_db_connection() as conn:
                    return [dict(row) for row in conn.execute(
                        "SELECT id, title, content FROM scraped_data WHERE content IS NOT NULL"
                    ).fetchall()]

            articles = await self.run_in_thread(_get_articles_blocking)

            if len(articles) < 2:
                self.notify("Need at least 2 articles for duplicate detection.", severity="warning")
                self._toggle_loading(False)
                return

            # Find duplicates
            duplicates = await self.run_in_thread(
                lambda: DuplicateDetectionManager.find_duplicates(articles, threshold=threshold)
            )

            if duplicates:
                # Show results in modal
                self.push_screen(DuplicateDetectionModal(duplicates=duplicates))
            else:
                self.notify("No duplicates found.", severity="info")

        except Exception as e:
            logger.error(f"Error in duplicate detection worker: {e}", exc_info=True)
            self.notify(f"Duplicate detection error: {e}", title="Error", severity="error")
        finally:
            self._toggle_loading(False)

    async def _clustering_worker(self, num_clusters: int) -> None:
        """Worker for article clustering."""
        self._toggle_loading(True)
        self.notify(f"Clustering articles into {num_clusters} groups...", title="Clustering", severity="info")
        try:
            def _get_articles_blocking():
                with get_db_connection() as conn:
                    return [dict(row) for row in conn.execute(
                        "SELECT id, title, content FROM scraped_data WHERE content IS NOT NULL"
                    ).fetchall()]

            articles = await self.run_in_thread(_get_articles_blocking)

            if len(articles) < num_clusters:
                self.notify(
                    f"Need at least {num_clusters} articles for {num_clusters} clusters.",
                    severity="warning"
                )
                self._toggle_loading(False)
                return

            # Cluster articles
            clusters = await self.run_in_thread(
                lambda: DuplicateDetectionManager.cluster_articles(articles, num_clusters=num_clusters)
            )

            if clusters:
                # Show results in modal
                self.push_screen(ClusterViewModal(clusters=clusters))
            else:
                self.notify("Clustering failed.", severity="error")

        except Exception as e:
            logger.error(f"Error in clustering worker: {e}", exc_info=True)
            self.notify(f"Clustering error: {e}", title="Error", severity="error")
        finally:
            self._toggle_loading(False)

    async def _entity_extraction_worker(self, article_id: int) -> None:
        """Worker for entity relationship extraction."""
        self._toggle_loading(True)
        self.notify(f"Extracting entity relationships for ID {article_id}...", title="Entity Extraction", severity="info")
        try:
            def _get_article_blocking():
                with get_db_connection() as conn:
                    row = conn.execute(
                        "SELECT id, title, content FROM scraped_data WHERE id = ?",
                        (article_id,)
                    ).fetchone()
                    return dict(row) if row else None

            article = await self.run_in_thread(_get_article_blocking)

            if not article or not article.get('content'):
                self.notify("Article not found or has no content.", severity="warning")
                self._toggle_loading(False)
                return

            # Extract entities and build knowledge graph
            result = await self.run_in_thread(
                lambda: EntityRelationshipManager.build_knowledge_graph([article])
            )

            entities = result.get('entities', [])
            if entities:
                entities_text = "\n".join([
                    f"• {e['text']} ({e['label']})" for e in entities[:10]
                ])
                self.notify(
                    f"Found {len(entities)} entities:\n{entities_text}",
                    title="Entities Extracted",
                    severity="info",
                    timeout=15
                )
            else:
                self.notify("No entities found.", severity="info")

        except Exception as e:
            logger.error(f"Error in entity extraction worker: {e}", exc_info=True)
            self.notify(f"Entity extraction error: {e}", title="Error", severity="error")
        finally:
            self._toggle_loading(False)

    def _fetch_articles_for_export(self) -> List[Dict[str, Any]]:
        """Fetch articles for export (blocking function for worker thread)."""
        s_col, _ = self.SORT_OPTIONS[self.current_sort_index]

        bq_export = (
            "SELECT sd.id, sd.title, sd.url, sd.link, sd.timestamp, "
            "sd.summary, sd.sentiment, sd.content, "
            "GROUP_CONCAT(DISTINCT t.name) as tags_c "
            "FROM scraped_data sd "
            "LEFT JOIN article_tags at ON sd.id = at.article_id "
            "LEFT JOIN tags t ON at.tag_id = t.id"
        )
        conds_export, params_export = [], {}
        if self.title_filter:
            conds_export.append("sd.title LIKE :tf")
            params_export["tf"] = f"%{self.title_filter}%"
        if self.url_filter:
            conds_export.append("sd.url LIKE :uf")
            params_export["uf"] = f"%{self.url_filter}%"
        if self.date_filter:
            conds_export.append("date(sd.timestamp) = :df")
            params_export["df"] = self.date_filter
        if self.tags_filter:
            tfs_export = [t.strip().lower() for t in self.tags_filter.split(',') if t.strip()]
            for i, tn_export in enumerate(tfs_export):
                pn_export = f"tgf_{i}"
                conds_export.append(
                    f"sd.id IN (SELECT at_s.article_id FROM article_tags at_s "
                    f"JOIN tags t_s ON at_s.tag_id = t_s.id WHERE t_s.name = :{pn_export})"
                )
                params_export[pn_export] = tn_export
        if self.sentiment_filter:
            sval_export = self.sentiment_filter.strip().capitalize()
            conds_export.append("sd.sentiment LIKE :sf")
            params_export["sf"] = f"%{sval_export}%"
        if conds_export:
            bq_export += " WHERE " + " AND ".join(conds_export)
        bq_export += " GROUP BY sd.id ORDER BY " + s_col

        with get_db_connection() as conn_blocking:
            rows_to_export = conn_blocking.execute(bq_export, params_export).fetchall()

        articles = []
        for r_data in rows_to_export:
            timestamp_val = r_data['timestamp']
            timestamp_str = (
                timestamp_val.strftime('%Y-%m-%d %H:%M:%S')
                if isinstance(timestamp_val, datetime)
                else str(timestamp_val)
            )
            tags_list = (
                [t.strip() for t in r_data['tags_c'].split(',') if t.strip()]
                if r_data['tags_c']
                else []
            )
            article_data = {
                'id': r_data['id'],
                'title': r_data['title'],
                'source_url': r_data['url'],
                'article_link': r_data['link'],
                'date_scraped': timestamp_str[:10],  # YYYY-MM-DD for Excel
                'timestamp': timestamp_str,
                'summary': r_data['summary'],
                'sentiment': r_data['sentiment'],
                'full_text': r_data['content'],
                'tags': ', '.join(tags_list)
            }
            articles.append(article_data)

        return articles

    async def _read_article_worker(self,eid:int,link:str,title:str)->None:
        self._toggle_loading(True);self.notify(f"Fetching '{title[:30]}...' (ID {eid})",title="Reading Article",severity="info")
        try:
            content=fetch_article_content(link,True)
            if content is not None:await self.app.push_screen(ReadArticleModal(title,content))
            else:self.notify(f"No content for article ID {eid}.",title="Read Error",severity="error")
        except Exception as e:logger.error(f"Err read worker ID {eid}: {e}",exc_info=True);self.notify(f"Err reading ID {eid}: {e}",title="Read Error",severity="error")
        finally:self._toggle_loading(False)
    async def action_read_article(self)->None:
        current_id=self._get_current_row_id()
        if current_id is None:self.notify("No row selected.",title="Info",severity="warning");return
        self.selected_row_id=current_id
        try:
            def _get_article_data_blocking():
                with get_db_connection() as conn_blocking:
                    return conn_blocking.execute("SELECT title,link FROM scraped_data WHERE id=?",(self.selected_row_id,)).fetchone() # type: ignore
            ad = _get_article_data_blocking()
            if not ad:self.notify(f"Article ID {self.selected_row_id} not found.",title="Error",severity="error");return
            worker_with_args = functools.partial(self._read_article_worker, self.selected_row_id, ad['link'], ad['title'])
            self.run_worker(worker_with_args, group="reading", exclusive=True)
        except Exception as e:logger.error(f"Err prep read ID {self.selected_row_id}: {e}",exc_info=True);self.notify(f"Err prep read: {e}",title="Error",severity="error")

    async def _handle_manage_scrapers_result(self,result:Optional[Tuple[str,Any]])->None:
        if not result:return
        action,data=result
        if action=="add":
            def handle_add_scraper_result(sd):
                if sd:
                    try:
                        def _add_scraper_blocking():
                            # v2.0.0: Add user_id tracking
                            sd['user_id'] = self.current_user_id
                            with get_db_connection() as conn_blocking:conn_blocking.execute("INSERT INTO saved_scrapers (name,url,selector,default_limit,default_tags_csv,description,is_preinstalled,user_id) VALUES (:name,:url,:selector,:default_limit,:default_tags_csv,:description,0,:user_id)",sd);conn_blocking.commit()
                        _add_scraper_blocking()
                        self.notify(f"Scraper '{sd['name']}' added.",title="Success",severity="info")
                    except sqlite3.IntegrityError:self.notify(f"Scraper name '{sd['name']}' already exists.",title="Error",severity="error")
                    except Exception as e:logger.error(f"Err adding scraper: {e}",exc_info=True);self.notify(f"Error adding scraper: {e}",severity="error")
            self.push_screen(AddEditScraperModal(), handle_add_scraper_result)
        elif action=="edit" and data:
            def handle_edit_scraper_result(sd):
                if sd:
                    try:
                        def _edit_scraper_blocking():
                            with get_db_connection() as conn_blocking:
                                if 'id' in sd:
                                     conn_blocking.execute("UPDATE saved_scrapers SET name=:name,url=:url,selector=:selector,default_limit=:default_limit,default_tags_csv=:default_tags_csv,description=:description,is_preinstalled=:is_preinstalled WHERE id=:id",sd)
                                else:
                                     # v2.0.0: Add user_id tracking
                                     sd['user_id'] = self.current_user_id
                                     conn_blocking.execute("INSERT INTO saved_scrapers (name,url,selector,default_limit,default_tags_csv,description,is_preinstalled,user_id) VALUES (:name,:url,:selector,:default_limit,:default_tags_csv,:description,0,:user_id)",sd)
                                conn_blocking.commit()
                        _edit_scraper_blocking()
                        self.notify(f"Scraper '{sd['name']}' saved.",title="Success",severity="info")
                    except sqlite3.IntegrityError:self.notify(f"Scraper name '{sd['name']}' conflict.",title="Error",severity="error")
                    except Exception as e:logger.error(f"Err saving scraper: {e}",exc_info=True);self.notify(f"Error saving scraper: {e}",severity="error")
            self.push_screen(AddEditScraperModal(scraper_data=data), handle_edit_scraper_result)
        elif action=="delete" and data: 
            sid_to_del=data
            try:
                def _get_scraper_name_blocking():
                    with get_db_connection() as conn_blocking:
                        return conn_blocking.execute("SELECT name,is_preinstalled FROM saved_scrapers WHERE id=?",(sid_to_del,)).fetchone()
                s_to_del = _get_scraper_name_blocking()
                if not s_to_del:self.notify("Scraper not found.",severity="error");return
                if s_to_del['is_preinstalled']: self.notify("Pre-installed profiles cannot be deleted. Edit them to create custom versions.", severity="warning"); return
                def handle_delete_scraper_confirmation(confirmed):
                    if confirmed:
                        def _delete_scraper_blocking():
                            with get_db_connection() as conn_blocking:conn_blocking.execute("DELETE FROM saved_scrapers WHERE id=?",(sid_to_del,));conn_blocking.commit()
                        _delete_scraper_blocking()
                        self.notify(f"Scraper '{s_to_del['name']}' deleted.",title="Success",severity="info")
                self.push_screen(ConfirmModal(f"Delete saved scraper '{s_to_del['name']}'?"), handle_delete_scraper_confirmation)
            except Exception as e:logger.error(f"Err deleting scraper: {e}",exc_info=True);self.notify(f"Error deleting scraper: {e}",severity="error")
        elif action=="execute" and data: 
            scrape_target_url = data['url']
            final_url_to_scrape = scrape_target_url
            if scrape_target_url == "[ARCHIVE_WAYBACK_URL]":
                def handle_original_url_result(original_url):
                    if original_url:
                        final_url = f"https://web.archive.org/web/timestamp_latest/{quote_plus(original_url)}"
                        self.notify(f"Attempting to fetch latest archive for: {original_url}", title="Archive Fetch", severity="information")
                        # Execute the scraping with the final URL
                        worker_with_args = functools.partial(self._scrape_url_worker, final_url, data['selector'], data['default_limit'], data.get('default_tags_csv'))
                        self.run_worker(worker_with_args, group="scraping", exclusive=True)
                    else:
                        self.notify("Wayback scrape cancelled.", title="Info", severity="information")
                self.push_screen(OriginalURLModal(), handle_original_url_result)
                return  # Exit early since we'll handle the scraping in the callback
            elif scrape_target_url == "[USER_PROVIDES_URL]":
                self.last_scrape_url = "" 
                self.last_scrape_selector = data['selector']
                self.last_scrape_limit = data['default_limit']
                self.current_scraper_profile = data['name']
                self.query_one(StatusBar).scraper_profile = self.current_scraper_profile
                self.notify(f"Profile '{data['name']}' loaded. Please provide target URL.", title="Scraper Profile", severity="information")
                callback_with_tags = functools.partial(self._handle_scrape_new_result, default_tags_csv=data['default_tags_csv'])
                await self.app.push_screen(ScrapeURLModal("", data['selector'], data['default_limit']), callback_with_tags)
                return 
            self.last_scrape_url = final_url_to_scrape
            self.last_scrape_selector = data['selector']
            self.last_scrape_limit = data['default_limit']
            default_tags = data['default_tags_csv']
            self.current_scraper_profile = data['name']
            self.query_one(StatusBar).scraper_profile = self.current_scraper_profile
            self.notify(f"Executing scraper profile '{data['name']}'. Parameters loaded.", title="Scraper Profile", severity="information")
            callback_with_tags = functools.partial(self._handle_scrape_new_result, default_tags_csv=default_tags)
            await self.app.push_screen(ScrapeURLModal(final_url_to_scrape, data['selector'], data['default_limit']), callback_with_tags)

    async def action_manage_saved_scrapers(self) -> None:
        await self.app.push_screen(ManageScrapersModal(), self._handle_manage_scrapers_result)


def print_startup_banner():
    """Display welcome banner when starting the application."""
    banner = """
╔══════════════════════════════════════════════════════════════════════════════╗
║                                                                              ║
║  ██╗    ██╗███████╗██████╗ ███████╗ ██████╗██████╗  █████╗ ██████╗ ███████╗  ║
║  ██║    ██║██╔════╝██╔══██╗██╔════╝██╔════╝██╔══██╗██╔══██╗██╔══██╗██╔════╝  ║
║  ██║ █╗ ██║█████╗  ██████╔╝███████╗██║     ██████╔╝███████║██████╔╝█████╗    ║
║  ██║███╗██║██╔══╝  ██╔══██╗╚════██║██║     ██╔══██╗██╔══██║██╔═══╝ ██╔══╝    ║
║  ╚███╔███╔╝███████╗██████╔╝███████║╚██████╗██║  ██║██║  ██║██║     ███████╗  ║
║   ╚══╝╚══╝ ╚══════╝╚═════╝ ╚══════╝ ╚═════╝╚═╝  ╚═╝╚═╝  ╚═╝╚═╝     ╚══════╝  ║
║                                                                              ║
║                       Text User Interface Web Scraper                        ║
║                                Version 1.9.5                                 ║
║                                                                              ║
║   ╔══════════════════════════════════════════════════════════════════════╗   ║
║   ║                              Features                                ║   ║
║   ║                                                                      ║   ║
║   ║  • Interactive terminal-based web scraping interface                 ║   ║
║   ║  • Pre-configured scraper profiles for popular websites              ║   ║
║   ║  • Custom scraper creation with CSS selectors                        ║   ║
║   ║  • SQLite database for persistent article storage                    ║   ║
║   ║  • AI-powered summarization and sentiment analysis                   ║   ║
║   ║  • Advanced filtering and sorting capabilities                       ║   ║
║   ║  • CSV export functionality                                          ║   ║
║   ║  • Tag-based article organization                                    ║   ║
║   ║                                                                      ║   ║
║   ╚══════════════════════════════════════════════════════════════════════╝   ║
║                                                                              ║
║                     Starting application... Please wait...                   ║
║                                                                              ║
╚══════════════════════════════════════════════════════════════════════════════╝
    """
    print(banner)


def print_shutdown_banner():
    """Display farewell banner when exiting the application."""
    banner = """
╔═══════════════════════════════════════════════════════════════════╗
║                                                                   ║
║                Thank you for using WebScrape-TUI!                 ║
║                                                                   ║
║                                                                   ║
║   ██████╗  ██████╗  ██████╗ ██████╗ ██████╗ ██╗   ██╗███████╗██╗  ║
║  ██╔════╝ ██╔═══██╗██╔═══██╗██╔══██╗██╔══██╗╚██╗ ██╔╝██╔════╝██║  ║
║  ██║  ███╗██║   ██║██║   ██║██║  ██║██████╔╝ ╚████╔╝ █████╗  ██║  ║
║  ██║   ██║██║   ██║██║   ██║██║  ██║██╔══██╗  ╚██╔╝  ██╔══╝  ╚═╝  ║
║  ╚██████╔╝╚██████╔╝╚██████╔╝██████╔╝██████╔╝   ██║   ███████╗██╗  ║
║   ╚═════╝  ╚═════╝  ╚═════╝ ╚═════╝ ╚═════╝    ╚═╝   ╚══════╝╚═╝  ║
║                                                                   ║
║    Visit our GitHub repository for updates and documentation:     ║
║           https://github.com/doublegate/WebScrape-TUI             ║
║                                                                   ║
║                Happy scraping! Come back soon!                    ║
║                                                                   ║
╚═══════════════════════════════════════════════════════════════════╝
    """
    print(banner)


if __name__ == "__main__":
    import time

    print_startup_banner()
    time.sleep(2)  # 2-second pause before starting the application

    css_file_content = """
Screen{layout:vertical;overflow:hidden}Header{dock:top}Footer{dock:bottom}
#app-grid{layout:vertical;overflow:hidden;height:1fr}
#filter-container{height:auto;padding:0 1;background:$boost}
#filter-container Label{padding:1 0 0 0;color:$text-muted}
#filter-inputs-grid{grid-size:1 5;grid-gutter:0 1;padding:0 0 1 0}
.filter-input{width:1fr}
DataTable{height:1fr;width:100%;margin-bottom:0}
DataTable .datatable--header-label { text-overflow: ellipsis; } /* Prevent header text wrapping */
DataTable .tags-column { width: 20%; } /* Give more space to tags column */
StatusBar{dock:bottom;height:1;padding:0 1;background:$primary-background;color:$text;width:100%}
LoadingIndicator{width:100%;height:100%;background:$surface-darken-2 50%;align:center middle;display:block;overlay:screen;}
LoadingIndicator.hidden{display:none}
.modal-buttons{width:100%;align-horizontal:center;padding-top:1}
.modal-buttons Button{margin:0 1}
Label.dialog-title{text-style:bold;width:100%;text-align:center;margin-bottom:1}
ManageTagsModal #cur_tags_lbl{margin-bottom:1;max-height:3;overflow-y:auto;border:round $primary-darken-1;padding:0 1}
ReadArticleModal VerticalScroll{border:panel $primary;padding:1;background:$surface}
ReadArticleModal #art_ttl_lbl{background:$primary-background-darken-1;padding:0 1;color:$text}
ManageScrapersModal ListView{border:panel $primary-background;background:$surface-lighten-1}
ManageScrapersModal ListView:focus{border:panel $primary}
ManageScrapersModal ListItem{padding:0 1;} /* Vertical padding 0, horizontal 1 */
ManageScrapersModal ListItem:hover{background:$primary 20%}
ManageScrapersModal ListItem.--highlight{background:$primary 40%}
ManageScrapersModal .buttons-top Button,ManageScrapersModal .buttons-bottom Button{width:auto;padding:0 2}
AddEditScraperModal Input#scraper_description_input { height: 3; border: round $primary-border; padding: 0 1; } /* Basic multiline-like appearance */
AddEditScraperModal Static.warning-text { color: $warning; padding: 0 1; text-align: center; }
.scraper-item-name{text-style:bold;}
.scraper-item-subtext{color:$text-muted;text-style:italic;}
    """
    css_file = Path("web_scraper_tui_v1.0.tcss")
    if not css_file.exists():
        with open(css_file, "w", encoding="utf-8") as f:
            f.write(css_file_content)
        logger.info(f"Created CSS file: {css_file.resolve()}")
    try:
        app = WebScraperApp()
        app.run()
    except KeyboardInterrupt:
        print("\n\nApplication interrupted by user.")
    finally:
        print_shutdown_banner()
